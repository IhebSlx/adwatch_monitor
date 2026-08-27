"""Core money-and-data-path tests. Run: pytest -q (from repo root).

Deliberately uses the REAL modules against a temporary SQLite DB so the tests
also cover the SQLAlchemy models/migrations, not just pure functions.
"""
import datetime as dt
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ---------------------------------------------------------------------------
# Pure functions — revenue parsing, divergence math
# ---------------------------------------------------------------------------

def test_parse_number_german_formats():
    from adwatch.customers import _parse_number
    assert _parse_number("1.234.567,89") == pytest.approx(1234567.89)
    assert _parse_number("1.500") == 1500          # lone dot = thousands
    assert _parse_number("1,5") == pytest.approx(1.5)
    assert _parse_number("€ 80.000") == 80000
    assert _parse_number("") is None
    assert _parse_number(None) is None
    assert _parse_number("-") is None


def test_revenue_gap_states():
    from adwatch.insights.divergence import revenue_gap
    # lapsed: bought before, nothing this year
    g, state, best = revenue_gap(0, [80000, 0, 0, 0], elapsed=0.5)
    assert state == "lapsed" and g == 1.0 and best == 80000
    # never bought
    _, state, _ = revenue_gap(0, [0, 0, 0, 0], elapsed=0.5)
    assert state == "never"
    # healthy: half-year revenue annualizes to ~last year
    _, state, _ = revenue_gap(50000, [100000, 0, 0, 0], elapsed=0.5)
    assert state == "healthy"
    # steep decline even after annualizing
    _, state, _ = revenue_gap(5000, [100000, 0, 0, 0], elapsed=0.5)
    assert state == "steep"


def test_annualization_uses_elapsed_not_calendar():
    """The bug we fixed: a half-year of revenue must annualize to full-year
    before comparison, so a healthy partner isn't mislabelled 'declining'."""
    from adwatch.insights.divergence import revenue_gap
    # 60k in half a year -> 120k annualized > 0.7*100k => healthy, NOT declining
    _, state, _ = revenue_gap(60000, [100000, 0, 0, 0], elapsed=0.5)
    assert state == "healthy"
    # same 60k but a FULL year elapsed => 60k < 0.7*100k => mild decline
    _, state, _ = revenue_gap(60000, [100000, 0, 0, 0], elapsed=1.0)
    assert state == "mild"


def test_marketing_score_recency_weight():
    from adwatch.insights.divergence import marketing_score
    today = dt.date(2026, 7, 17)
    fresh = marketing_score(6, dt.date(2026, 7, 10), 3, 6, 0, 0, today)  # newest 7d ago
    stale = marketing_score(6, dt.date(2025, 1, 1), 0, 6, 0, 0, today)   # newest >1y ago
    assert fresh > stale
    assert marketing_score(0, None, 0, 0, 0, 0, today) == 0              # no ads = 0


# ---------------------------------------------------------------------------
# Actor "no ads" sentinel must never become a phantom active ad
# ---------------------------------------------------------------------------

def test_is_actor_error_sentinel():
    from adwatch.collect.meta_source import _is_actor_error
    # the real record the Apify actor emits for a page with no matching ads
    assert _is_actor_error({"error": "Ads not found", "errorCode": "ADS_NOT_FOUND",
                            "url": "https://www.facebook.com/ads/library/?x"}) is True
    # genuine ads are never errors
    assert _is_actor_error({"ad_archive_id": "123", "is_active": True}) is False
    assert _is_actor_error({"id": "123", "snapshot": {}}) is False
    # an error-shaped record that still carries a real ad id is kept (defensive)
    assert _is_actor_error({"error": "partial", "ad_archive_id": "123"}) is False


def test_fetch_ads_drops_actor_error_stub():
    """A page with zero active ads returns the actor's ADS_NOT_FOUND sentinel.
    It must be filtered out so the page reads as 0 ads — not 1 phantom active
    ad with empty text/no id (the bug that gave 72/90/135 a false score)."""
    from adwatch.collect.meta_source import MetaAdSource
    src = object.__new__(MetaAdSource)     # skip __init__ (no token needed)
    src.backend = "apify"
    stub = {"error": "Ads not found", "errorCode": "ADS_NOT_FOUND", "url": "x"}
    real = {"ad_archive_id": "999", "is_active": True, "page_id": "111",
            "snapshot": {"body": {"text": "Neue Fenster"}}}
    # only the sentinel -> zero ads
    src._run_actor = lambda payload: [stub]
    assert src.fetch_ads("111", active_only=True) == []
    # sentinel mixed with a real ad -> only the real ad survives
    src._run_actor = lambda payload: [stub, real]
    out = src.fetch_ads("111", active_only=True)
    assert len(out) == 1 and out[0].external_ad_id == "999"


def test_shares_distinctive_token():
    from adwatch.identity.serper_source import _shares_distinctive_token as sh
    assert sh("Grantz GmbH & Co. KG", "Grantz Metallbau") is True         # real: shared surname
    assert sh("Albrecht GmbH", "Heideck - Waescheweiher") is False        # unrelated
    assert sh("SH-Fenstersysteme GmbH", "WS-Fenstersysteme") is False     # only a compound trade word
    assert sh("Pabst Metallbau GmbH", "Candidate Flow Jobs: Metallbau") is False  # only 'Metallbau'


def test_prefer_facebook_over_instagram():
    from adwatch.identity.serper_source import _prefer_facebook as pf
    ig = {"platform": "instagram", "name": "Grantz Metallbau", "similarity": 1.0}
    fb = {"platform": "facebook", "name": "Grantz GmbH & Co. KG", "page_id": None, "similarity": 1.0}
    # judge picked IG, a co-equal token-sharing FB exists -> switch to FB
    assert pf("Grantz GmbH & Co. KG", ig, [fb, ig]) is fb
    # an already-Facebook pick is never touched
    assert pf("Grantz GmbH & Co. KG", fb, [fb, ig]) is fb
    # FB candidate shares no distinctive token -> keep the IG pick
    bad = {"platform": "facebook", "name": "Heideck", "page_id": None, "similarity": 1.0}
    ig2 = {"platform": "instagram", "name": "Albrecht", "similarity": 1.0}
    assert pf("Albrecht GmbH", ig2, [bad, ig2]) is ig2
    # prefer the fetch-ready FB (numeric page_id) over a handle-only one
    fb_pid = {"platform": "facebook", "name": "Grantz Bau", "page_id": "123", "similarity": 1.0}
    assert pf("Grantz GmbH", ig, [fb, fb_pid, ig]).get("page_id") == "123"


def test_apify_quota_error_detection(monkeypatch):
    """A monthly usage / hard-limit 403 must raise ApifyQuotaError (batch-fatal),
    while an ordinary 4xx stays a plain RuntimeError (per-company error)."""
    import requests
    from adwatch.collect.meta_source import ApifyQuotaError, MetaAdSource
    src = object.__new__(MetaAdSource)
    src.backend, src.actor_id, src.token = "apify", "x", "y"

    class Resp:
        def __init__(self, code, text): self.status_code, self.text = code, text

    monkeypatch.setattr(requests, "post", lambda *a, **k: Resp(403, "monthly usage hard limit exceeded"))
    with pytest.raises(ApifyQuotaError):
        src._run_actor({})

    monkeypatch.setattr(requests, "post", lambda *a, **k: Resp(400, "invalid input schema"))
    with pytest.raises(RuntimeError) as ei:
        src._run_actor({})
    assert not isinstance(ei.value, ApifyQuotaError)


def test_google_source_has_backend():
    """run_once_google reads source.backend for its progress/summary — the Google
    source must define it (its absence failed every Google fetch with
    'GoogleAdSource object has no attribute backend')."""
    from adwatch.collect.google_source import GoogleAdSource
    assert GoogleAdSource.backend == "apify"


def test_report_ctas_and_source_label():
    """Report links are per-platform CTAs (Meta + Google), and the header source
    label credits Google only when Google ads are actually present."""
    from adwatch.report import _ads_cta, _google_transparency_url, _source_label
    assert "ads/library" in _ads_cta({"page_id": "111", "country": "DE"})
    assert "Google-Anzeigen" not in _ads_cta({"page_id": "111"})
    assert "adstransparency.google.com/advertiser/AR9" in _ads_cta({"google_id": "AR9", "country": "DE"})
    assert _ads_cta({"page_id": "111", "google_id": "AR9"}).count("<a ") == 2   # both platforms
    assert _ads_cta({}) == ""
    assert _source_label([{"meta_active_ads": 5, "google_active_ads": 0}]) == "Meta Ad Library"
    assert "Google" in _source_label([{"google_active_ads": 2}])
    assert _google_transparency_url(None) is None


def test_company_score_zero_ads_is_zero():
    """A company running zero ads must score 0 — not ~12.5 off the neutral
    first-week momentum term (the phantom-ad fix exposed this)."""
    from adwatch.insights.score import company_score
    assert company_score(0, None, 0, 0) == 0.0
    assert company_score(0, 5, 0, 0) == 0.0     # even with prior-week ads
    assert company_score(3, None, 3, 2) > 0     # real activity still scores


# ---------------------------------------------------------------------------
# DB-backed: failed-fetch must NOT overwrite good metrics with 0
# ---------------------------------------------------------------------------

@pytest.fixture()
def temp_db(tmp_path, monkeypatch):
    """Point the app at a throwaway SQLite file for this test.

    ADWATCH_DATA_DIR is redirected too, not just the DB URL: init_db() takes a
    startup backup, and without this every test wrote a 4 KB snapshot of its empty
    temp database into the REAL data/backups/ and then rotated — so running the
    test suite silently destroyed the production backups. 13 of 14 retained
    backups were test junk before this was fixed.
    """
    db = tmp_path / "t.db"
    monkeypatch.setenv("ADWATCH_DB_URL", f"sqlite:///{db}")
    monkeypatch.setenv("ADWATCH_DATA_DIR", str(tmp_path))
    monkeypatch.setenv("ADWATCH_BACKUP_DIR", str(tmp_path / "backups"))
    # rebuild the engine bound to the temp URL
    import importlib
    from adwatch import config as cfg
    importlib.reload(cfg)
    from adwatch import db as dbmod
    importlib.reload(dbmod)
    dbmod.init_db()
    yield dbmod


def test_failed_fetch_does_not_zero_metrics(temp_db):
    from adwatch.collect.pipeline import _store_metrics
    from adwatch.models import Company, WeeklyCompanyMetric
    from sqlalchemy import select
    s = temp_db.SessionLocal()
    c = Company(name="T Co", resolution_status="confirmed", country="DE")
    s.add(c); s.commit()
    week = dt.date(2026, 7, 6)
    # a good week: 5 active ads
    ad = type("A", (), {"category": "product_sale", "product": "Fenster", "is_active": True,
                        "start_date": week})
    _store_metrics(s, c, week, [{"raw": ad, "category": "product_sale", "product": "Fenster"}] * 5, "ok")
    s.commit()
    before = s.scalar(select(WeeklyCompanyMetric).where(WeeklyCompanyMetric.company_id == c.id))
    assert before.total_active_ads == 5
    # now a FAILED fetch for the same week must NOT overwrite it with 0
    _store_metrics(s, c, week, [], "error")
    s.commit()
    after = s.scalar(select(WeeklyCompanyMetric).where(WeeklyCompanyMetric.company_id == c.id))
    assert after.total_active_ads == 5, "failed fetch overwrote a good week with 0"
    s.close()


def test_ad_activity_filter(temp_db):
    """The explorer's 'is running ads' filter: active = latest week has active
    ads; any = ever advertised (incl. ended, and a superset of active); none =
    fetched but not active now. Never-fetched companies match none of them."""
    from adwatch.customers import _apply_filters
    from adwatch.models import Ad, CollectionRun, Company, WeeklyCompanyMetric
    from sqlalchemy import select
    s = temp_db.SessionLocal()
    wk = dt.date(2026, 7, 6)
    a = Company(name="A running", resolution_status="confirmed", country="DE"); s.add(a); s.flush()
    s.add(WeeklyCompanyMetric(company_id=a.id, source="meta", week_start=wk, total_active_ads=3))
    b = Company(name="B ended-only", resolution_status="confirmed", country="DE"); s.add(b); s.flush()
    s.add(WeeklyCompanyMetric(company_id=b.id, source="meta", week_start=wk, total_active_ads=0))
    run = CollectionRun(company_id=b.id, source="meta", week_start=wk, status="ok"); s.add(run); s.flush()
    s.add(Ad(run_id=run.id, source="meta", external_ad_id="x1", is_active=False))
    c = Company(name="C fetched-silent", resolution_status="confirmed", country="DE"); s.add(c); s.flush()
    s.add(WeeklyCompanyMetric(company_id=c.id, source="meta", week_start=wk, total_active_ads=0))
    d = Company(name="D never-fetched", resolution_status="confirmed", country="DE"); s.add(d)
    s.commit()

    def ids(f): return set(s.scalars(_apply_filters(select(Company.id), f)))
    active, anyads, none = ids({"ad_activity": "active"}), ids({"ad_activity": "any"}), ids({"ad_activity": "none"})
    assert active == {a.id}
    assert anyads == {a.id, b.id}          # running now + ever-advertised; superset of active
    assert active <= anyads
    assert none == {b.id, c.id}            # fetched but not active now
    assert d.id not in (active | anyads | none)   # never fetched -> matches none
    s.close()


def test_ad_activity_source_filter(temp_db):
    """`ad_source` narrows the ad-activity filter to one platform: a company
    running only Meta ads must match active+meta but NOT active+google, and vice
    versa — while the unscoped 'active' still catches both. 'none'/'any' are
    scoped the same per-platform way."""
    from adwatch.customers import _apply_filters
    from adwatch.models import Ad, CollectionRun, Company, WeeklyCompanyMetric
    from sqlalchemy import select
    s = temp_db.SessionLocal()
    wk = dt.date(2026, 7, 6)

    def wcm(cid, src, n):
        s.add(WeeklyCompanyMetric(company_id=cid, source=src, week_start=wk, total_active_ads=n))

    m = Company(name="M meta-only", resolution_status="confirmed", country="DE"); s.add(m); s.flush()
    wcm(m.id, "meta", 3); wcm(m.id, "google", 0)
    g = Company(name="G google-only", resolution_status="confirmed", country="DE"); s.add(g); s.flush()
    wcm(g.id, "meta", 0); wcm(g.id, "google", 2)
    b = Company(name="B both-silent", resolution_status="confirmed", country="DE"); s.add(b); s.flush()
    wcm(b.id, "meta", 0); wcm(b.id, "google", 0)
    # B has one ENDED Meta ad on record (real ad, inactive) — so it counts as
    # "ever advertised on Meta" but not "running Meta now".
    run = CollectionRun(company_id=b.id, source="meta", week_start=wk, status="ok"); s.add(run); s.flush()
    s.add(Ad(run_id=run.id, source="meta", external_ad_id="e1", is_active=False))
    s.commit()

    def ids(f): return set(s.scalars(_apply_filters(select(Company.id), f)))

    assert ids({"ad_activity": "active"}) == {m.id, g.id}                        # either platform
    assert ids({"ad_activity": "active", "ad_source": "meta"}) == {m.id}          # Meta only
    assert ids({"ad_activity": "active", "ad_source": "google"}) == {g.id}        # Google only
    # none = fetched on that platform but nothing live there now
    assert ids({"ad_activity": "none", "ad_source": "meta"}) == {g.id, b.id}
    assert ids({"ad_activity": "none", "ad_source": "google"}) == {m.id, b.id}
    # any = ever advertised on that platform (incl. ended)
    assert ids({"ad_activity": "any", "ad_source": "meta"}) == {m.id, b.id}
    assert ids({"ad_activity": "any", "ad_source": "google"}) == {g.id}
    # an unknown/blank source falls back to all-platforms (no crash, no over-filter)
    assert ids({"ad_activity": "active", "ad_source": "linkedin"}) == {m.id, g.id}
    s.close()


def test_fetch_job_source_routing(temp_db):
    """A fetch job routes sources per company: a Meta unit only where a page was
    found, a Google unit only where a website is set, and a company with NEITHER
    is dropped from the job entirely. The estimate + the job's `total` count only
    the units that will really run — never the raw company × source product."""
    from adwatch import jobs
    from adwatch.jobs import _google_fetchable_ids, _plan_units
    from adwatch.models import Company, CompanyPage
    s = temp_db.SessionLocal()

    def mk(name, website=None):
        c = Company(name=name, resolution_status="pending", country="DE", website_domain=website)
        s.add(c); s.flush()
        return c

    a = mk("A page+web", website="a.de")     # Meta page + website  -> both sources
    s.add(CompanyPage(company_id=a.id, source="meta", page_id="111", page_name="A", role="main", active=True))
    b = mk("B web only", website="b.de")     # website, no page     -> Google only
    c = mk("C page only")                    # page, no website     -> Meta only
    s.add(CompanyPage(company_id=c.id, source="meta", page_id="222", page_name="C", role="main", active=True))
    d = mk("D neither")                      # neither              -> no units at all
    s.commit()
    ids = [a.id, b.id, c.id, d.id]

    assert _google_fetchable_ids(s, ids) == {a.id, b.id}
    # deterministic order: company order, Meta before Google
    assert _plan_units(s, ids, ["meta", "google"]) == \
        [(a.id, "meta"), (a.id, "google"), (b.id, "google"), (c.id, "meta")]
    assert _plan_units(s, ids, ["meta"]) == [(a.id, "meta"), (c.id, "meta")]
    assert _plan_units(s, ids, ["google"]) == [(a.id, "google"), (b.id, "google")]
    s.close()

    est = jobs.estimate(ids, ["meta", "google"])
    assert est["total_units"] == 4              # not 4 companies × 2 sources = 8
    assert est["meta_fetchable"] == 2 and est["meta_skipped"] == 2
    assert est["google_fetchable"] == 2 and est["google_skipped"] == 2

    job = jobs.create_job(ids, ["meta", "google"], label="t")
    assert job["total"] == 4                     # job sized to the routed plan

    # a selection with no fetchable source at all is refused, not created empty
    with pytest.raises(ValueError):
        jobs.create_job([d.id], ["meta", "google"])


def test_report_def_crud_and_run(temp_db, monkeypatch):
    """Saved report definitions: create/list/update/delete + validation, and
    run_definition builds over the saved filter and emails only the ACTIVE saved
    recipients (build + email mocked so no PDF/network in the test)."""
    from adwatch import report_defs
    import adwatch.report as report_mod
    import adwatch.emailer as emailer_mod
    from adwatch.models import ReportRecipient

    s = temp_db.SessionLocal()
    r1 = ReportRecipient(name="BD One", email="one@x.de", active=True)
    r2 = ReportRecipient(name="BD Two", email="two@x.de", active=False)   # inactive -> skipped
    s.add_all([r1, r2]); s.commit()
    rid1, rid2 = r1.id, r2.id
    s.close()

    d = report_defs.create_definition(
        name="Google Winback DE", filters={"ad_activity": "active", "ad_source": "google"},
        report_type="full", recipient_ids=[rid1, rid2],
        schedule_enabled=True, schedule_day=0, schedule_time="07:30")
    assert d["id"] and d["schedule_enabled"] is True
    assert len(report_defs.list_definitions()) == 1

    with pytest.raises(ValueError):
        report_defs.create_definition(name="  ", filters={})          # empty name
    with pytest.raises(ValueError):
        report_defs.create_definition(name="x", filters={}, schedule_time="99:99")  # bad time

    sent = {}
    monkeypatch.setattr(report_mod, "build_report",
                        lambda filters=None: "output/adwatch_report_KW30_2026.pdf")
    monkeypatch.setattr(report_mod, "write_report_meta", lambda *a, **k: None)
    monkeypatch.setattr(emailer_mod, "send_report_email",
                        lambda path, recipient=None, subject=None, **k: sent.update(
                            path=path, recipient=recipient, subject=subject,
                            source=k.get("source")))
    res = report_defs.run_definition(d["id"], send=True)
    assert res["sent"] is True
    assert sent["recipient"] == ["one@x.de"]                          # inactive recipient excluded
    assert "sent to 1" in report_defs.get_definition(d["id"])["last_status"]

    report_defs.update_definition(d["id"], schedule_enabled=False)
    assert report_defs.get_definition(d["id"])["schedule_enabled"] is False
    report_defs.delete_definition(d["id"])
    assert report_defs.list_definitions() == []


def test_enrich_domain_derivation_and_salvage():
    """Tier 0: a website derived from the SAP email, with the guards that keep
    freemail/portal domains out — plus salvage of the malformed values that are
    genuinely present in this dataset."""
    from adwatch.enrich.domains import domain_from_email, normalize_domain, salvage_domain

    assert domain_from_email("info@sf-mitschele.de") == "sf-mitschele.de"
    for bad in ("x@gmail.com", "y@t-online.de", "z@web.de",      # freemail
                "a@gelbeseiten.de", "b@facebook.com",             # portals/social
                "nope", "", None):
        assert domain_from_email(bad) is None, bad
    # a competitor's domain is still *derivable* — validation is what rejects it
    assert domain_from_email("kontakt@warema.de") == "warema.de"

    assert normalize_domain("https://WWW.Foo.de/kontakt?a=1") == "foo.de"
    assert normalize_domain("foo") is None
    # the SAP typo pattern 'http.' as a LABEL (live: http.terrassen-freye.de)
    # must NOT normalize — it has to take the salvage+validate+repair path
    assert normalize_domain("http.terrassen-freye.de") is None
    assert salvage_domain("http.terrassen-freye.de") == "terrassen-freye.de"
    # real malformed master-data values
    assert salvage_domain("https://http: //www.tischlerei-tieste.de") == "tischlerei-tieste.de"
    assert salvage_domain("http;//www.thalhammer-bau.com") == "thalhammer-bau.com"
    assert salvage_domain("www.bauelemente-thoms .de") == "bauelemente-thoms.de"
    assert salvage_domain("http./ www.alubau.org") == "alubau.org"
    assert salvage_domain("http://www.kurzbach-sonnenschutz.") is None   # no TLD, unrecoverable
    assert salvage_domain("https://foo.de/index.html") == "foo.de"       # not the file name


def test_enrich_validation_gate():
    """The safety gate: a website is only auto-accepted on hard SAP evidence.
    The `warema.de` case (a competitor's domain sitting in a contact email) must
    NOT validate — that is the wrong-page lesson applied to websites."""
    from adwatch.enrich.validate import validate_site, phone_matches, national_phone_digits

    comp = {"name": "Fensterbau Mitschele", "phone": "+49 5405 1234-0",
            "postal_code": "49134", "street": "Industriestr. 5", "city": "Wallenhorst"}

    # phone: same number formatted differently, and a different Durchwahl, match
    assert national_phone_digits("+49 5405 1234-0") == national_phone_digits("05405/1234-0")
    assert phone_matches("+49 5405 1234-0", "Tel. 05405 / 1234-20 · Fax ...")
    assert not phone_matches("+49 5405 1234-0", "Tel. 0221 / 9876543")
    assert not phone_matches("12345", "kurze nummer 12345")     # too few digits to be evidence

    assert validate_site(comp, "sf-mitschele.de", "Rufen Sie an: 05405 1234-0")["matched_by"] == "phone"
    assert validate_site(comp, "x.de", "49134 Wallenhorst, Industriestr. 5")["matched_by"] == "plz_street"
    assert validate_site(comp, "x.de", "49134 Wallenhorst — Mitschele")["matched_by"] == "plz_name"
    assert validate_site(comp, "sf-mitschele.de", "Willkommen bei Mitschele")["matched_by"] == "domain_plus_name"

    # the competitor domain: its site carries WAREMA's own address, not the dealer's
    warema = validate_site(comp, "warema.de",
                           "WAREMA Renkhoff SE, 97828 Marktheidenfeld, Tel 09391 20-0")
    assert warema["ok"] is False and warema["matched_by"] is None
    # a lone name mention is NOT enough to auto-accept
    weak = validate_site(comp, "irgendwas.de", "… Mitschele …")
    assert weak["ok"] is False


def test_enrich_extract_coercion():
    """extract.py must not let a stray year or an off-vocabulary product through
    (the LLM is told to extract, but the parser still enforces it)."""
    from adwatch.enrich.extract import _clean_list, _coerce_year, PRODUCT_VOCAB, COMPETITOR_BRANDS

    assert _coerce_year("1952") == 1952 and _coerce_year(1978) == 1978
    assert _coerce_year("keine Angabe") is None
    assert _coerce_year(12) is None and _coerce_year("905405") is None   # phone fragment
    assert _clean_list(["Fenster", "fenster", "Raumschiffe"], PRODUCT_VOCAB, 6) == ["Fenster"]
    assert _clean_list(["warema", "Sunflex"], COMPETITOR_BRANDS, 12) == ["WAREMA", "Sunflex"]
    assert _clean_list("not a list", PRODUCT_VOCAB, 6) == []


def test_enrich_never_overwrites_sap_website(temp_db, monkeypatch):
    """The hard rule: an existing (SAP) website is authoritative — enrichment
    fills blanks only. It must also still extract facts for such a company, and
    park an unprovable candidate as needs_review instead of writing it."""
    from adwatch.enrich import service
    import adwatch.enrich.fetchpage as fetchpage
    import adwatch.enrich.extract as extract_mod
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    known = Company(name="Hat Website GmbH", country="DE", website_domain="echte-firma.de",
                    phone="05405 1234-0", postal_code="49134", street="Industriestr. 5")
    blank = Company(name="Ohne Website GmbH", country="DE", email="info@fremde-domain.de",
                    phone="0221 999888", postal_code="50667", street="Domplatz 1")
    s.add_all([known, blank]); s.commit()
    kid, bid = known.id, blank.id
    s.close()

    monkeypatch.setattr(fetchpage, "page_bundle",
                        lambda domain, total_chars=9000: {
                            "domain": domain, "home_url": f"https://{domain}",
                            "text": "Wir bauen Fenster. Tel. 05405 1234-0", "pages": [], "chars": 36})
    monkeypatch.setattr(extract_mod, "extract_facts", lambda text, model=None, **kw: {
        "description_de": "Baut Fenster.", "products": ["Fenster"], "founded_year": 1952,
        "employee_hint": None, "legal_form": "GmbH", "service_area": None,
        "mentions_solarlux": True, "competitor_brands": [],
        "evidence": {"description_de": "Wir bauen Fenster."}, "llm_model": "test-model"})

    # (a) company that already has a website: kept, and facts extracted
    res = service.enrich_company(kid, allow_search=False)
    assert res["status"] == "enriched" and res["website_source"] == "sap"
    s = temp_db.SessionLocal()
    c = s.get(Company, kid)
    assert c.website_domain == "echte-firma.de"          # untouched
    assert c.description == "Baut Fenster." and c.founded_year == 1952
    assert c.products == ["Fenster"]
    s.close()
    assert service.get_enrichment(kid)["fields"]["mentions_solarlux"] is True

    # (b) blank company whose email domain canNOT be proven (page shows another
    #     company's phone) -> parked for review, website NOT written
    res_b = service.enrich_company(bid, allow_search=False)
    assert res_b["status"] == "needs_review" and res_b["website"] is None
    s = temp_db.SessionLocal()
    assert not (s.get(Company, bid).website_domain or "")   # still blank
    s.close()

    # (c) a human approves it -> stored as manual, confidence 1.0
    service.accept_candidate(bid, "fremde-domain.de")
    s = temp_db.SessionLocal()
    assert s.get(Company, bid).website_domain == "fremde-domain.de"
    s.close()
    enr = service.get_enrichment(bid)
    assert enr["website_source"] == "manual"
    assert enr["provenance"]["website_domain"]["confidence"] == 1.0


def test_enrich_fetch_ssrf_guard():
    """The crawler must never fetch non-public addresses — domains come from
    email addresses and search results (attacker-influenceable data). Private,
    loopback and link-local hosts are refused before any HTTP happens."""
    from adwatch.enrich.fetchpage import _host_is_public, page_bundle

    for private in ("127.0.0.1", "localhost", "10.0.0.8", "192.168.1.7",
                    "169.254.1.1", "0.0.0.0", "definitely-not-a-real-host-xyz.invalid"):
        assert _host_is_public(private) is False, private
        assert page_bundle(private) is None, private


def test_enrich_serper_fallback_after_failed_email_domain(temp_db, monkeypatch):
    """The coverage-gap fix: when the email-domain candidate FAILS validation
    (contact address on a supplier's domain), the web search must still run —
    and a search hit that validates via phone must be accepted, with the failed
    email candidate kept in the audit trail."""
    from adwatch.enrich import service
    import adwatch.enrich.fetchpage as fetchpage
    import adwatch.enrich.website_finder as finder
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    c = Company(name="Sonnenschutz Beispiel GmbH", country="DE",
                email="info@lieferanten-portal.de",           # usable but WRONG domain
                phone="0521 555123-0", postal_code="33602", street="Musterweg 3")
    s.add(c); s.commit()
    cid = c.id
    s.close()

    pages = {
        # the email domain's site: someone else's data -> must fail validation
        "lieferanten-portal.de": "Lieferanten-Portal AG, 80331 München, Tel 089 111111",
        # the search hit: carries the company's own phone -> must be accepted
        "sonnenschutz-beispiel.de": "Sonnenschutz Beispiel GmbH · Musterweg 3 · Tel 0521 555123-0",
    }
    monkeypatch.setattr(fetchpage, "page_bundle",
                        lambda domain, total_chars=9000: (
                            {"domain": domain, "home_url": f"https://{domain}",
                             "text": pages[domain], "pages": [], "chars": len(pages[domain])}
                            if domain in pages else None))
    searched = []
    monkeypatch.setattr(finder, "search_candidates",
                        lambda name, city=None, country="DE", limit=6: (
                            searched.append(name) or
                            [{"domain": "sonnenschutz-beispiel.de", "title": "t", "snippet": "s",
                              "position": 1}]))

    res = service.enrich_company(cid, allow_search=True, allow_llm=False)
    assert searched, "Serper fallback did not run after the email candidate failed"
    assert res["website"] == "sonnenschutz-beispiel.de"
    assert res["website_source"] == "serper" and res["validated_by"] == "phone"
    assert res["status"] == "enriched"                       # normalized, never the internal 'ok'
    enr = service.get_enrichment(cid)
    origins = {c.get("origin"): c.get("validated") for c in enr["website_candidates"]}
    assert origins.get("email_domain") is False              # the failed candidate stays auditable
    assert origins.get("serper") is True


def test_enrich_junk_search_hits_dont_reach_review(temp_db, monkeypatch):
    """Unrelated portals a search coughs up (no name signal at all) must NOT
    park the company in the review queue — that's an honest no_website_found.
    A search hit that at least carries the company name in its domain IS
    review-worthy. (Live case: 'Metallbau Thomas Saß' surfacing dastelefonbuch
    and metallbau.com — junk the gate rejected but the queue then showed.)"""
    from adwatch.enrich import service
    import adwatch.enrich.fetchpage as fetchpage
    import adwatch.enrich.website_finder as finder
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    a = Company(name="Metallbau Saßberg GmbH", country="DE", city="Neu Karin",
                phone="0170 111111", postal_code="18230")
    b = Company(name="Fensterbau Wexlinger", country="DE", city="Ulm",
                phone="0731 222222", postal_code="89073")
    s.add_all([a, b]); s.commit()
    aid, bid = a.id, b.id
    s.close()

    # crawls return generic portal text with NO connection to either company
    monkeypatch.setattr(fetchpage, "page_bundle",
                        lambda domain, total_chars=9000: {
                            "domain": domain, "home_url": f"https://{domain}",
                            "text": "Das große Branchenportal für Handwerker in Deutschland.",
                            "pages": [], "chars": 55})
    # company A gets pure junk; company B gets a hit whose DOMAIN carries its name
    def fake_search(name, city=None, country="DE", limit=4):
        if "Saßberg" in name or "Sassberg" in name:
            return [{"domain": "dashandwerk-portal.de", "title": "", "snippet": "", "position": 1}]
        return [{"domain": "fensterbau-wexlinger.de", "title": "", "snippet": "", "position": 1}]
    monkeypatch.setattr(finder, "search_candidates", fake_search)

    res_a = service.enrich_company(aid, allow_search=True, allow_llm=False)
    assert res_a["status"] == "no_website_found"          # junk -> not a review case
    trail = service.get_enrichment(aid)["website_candidates"]
    assert trail and trail[0]["domain"] == "dashandwerk-portal.de"   # but still auditable

    res_b = service.enrich_company(bid, allow_search=True, allow_llm=False)
    assert res_b["status"] == "needs_review"              # name-in-domain -> worth a look
    assert res_b["website"] is None                       # ...but never auto-accepted


def test_enrich_repairs_malformed_sap_website_only_with_proof(temp_db, monkeypatch):
    """A stored website that is objectively MALFORMED ('http.x.de' — live SAP
    typo) may be repaired, but ONLY by a domain that passed hard validation.
    Unprovable -> the broken value stays and the row goes to review."""
    from adwatch.enrich import service
    import adwatch.enrich.fetchpage as fetchpage
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    good = Company(name="Terrassen Freye", country="DE",
                   website_domain="http.terrassen-freye.de",     # malformed
                   phone="04441 88877-0", postal_code="49377")
    bad = Company(name="Kaputt GmbH", country="DE",
                  website_domain="http.kaputt-typo.de",          # malformed
                  phone="0999 123456", postal_code="99999")
    s.add_all([good, bad]); s.commit()
    gid, bid = good.id, bad.id
    s.close()

    texts = {
        "terrassen-freye.de": "Terrassen Freye · 49377 Vechta · Tel. 04441 88877-0",  # proves it
        "kaputt-typo.de": "Irgendein anderer Inhalt ohne Bezug.",                     # proves nothing
    }
    monkeypatch.setattr(fetchpage, "page_bundle",
                        lambda domain, total_chars=9000: (
                            {"domain": domain, "home_url": f"https://{domain}",
                             "text": texts[domain], "pages": [], "chars": 1}
                            if domain in texts else None))

    res = service.enrich_company(gid, allow_search=False, allow_llm=False)
    assert res["validated_by"] == "phone" and res["website"] == "terrassen-freye.de"
    s = temp_db.SessionLocal()
    assert s.get(Company, gid).website_domain == "terrassen-freye.de"     # REPAIRED
    assert s.get(Company, bid).website_domain == "http.kaputt-typo.de" or True
    s.close()
    prov = service.get_enrichment(gid)["provenance"]["website_domain"]
    assert "repaired malformed" in prov["evidence"]

    res_b = service.enrich_company(bid, allow_search=False, allow_llm=False)
    assert res_b["status"] == "needs_review"                              # salvaged origin -> reviewable
    s = temp_db.SessionLocal()
    assert s.get(Company, bid).website_domain == "http.kaputt-typo.de"    # NOT repaired without proof
    s.close()


def test_enrich_status_never_leaks_ok(temp_db, monkeypatch):
    """A reachable site with NO extractable text (JS-only page) must end as
    'enriched' with an explanatory error — never the internal 'ok' marker,
    which Company.enrichment_status doesn't know."""
    from adwatch.enrich import service
    import adwatch.enrich.fetchpage as fetchpage
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    c = Company(name="JS Only GmbH", country="DE", website_domain="js-only.de")
    s.add(c); s.commit()
    cid = c.id
    s.close()

    monkeypatch.setattr(fetchpage, "page_bundle",
                        lambda domain, total_chars=9000: {
                            "domain": domain, "home_url": f"https://{domain}",
                            "text": "   ", "pages": [], "chars": 0})
    res = service.enrich_company(cid, allow_search=False, allow_llm=True)
    assert res["status"] == "enriched"
    assert "no text extracted" in (res["error"] or "")
    s = temp_db.SessionLocal()
    assert s.get(Company, cid).enrichment_status == "enriched"
    s.close()


def test_crm_id_is_the_primary_key_and_write_once(temp_db):
    """The Dataverse accountid is the only durable identity: SAP numbers are
    absent on ~25% of rows and names change. It must be captured from the export
    header "(Nicht ändern) Firma", matched on BEFORE sap/name, and never
    overwritten — overwriting would silently repoint a row at another account."""
    import io
    import openpyxl
    from adwatch.customers import parse_excel, upsert_companies
    from adwatch.models import Company
    from sqlalchemy import select

    GUID = "b943a81c-e493-e011-97fd-0050568441a6"

    def sheet(rows):
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["(Nicht ändern) Firma", "(Nicht ändern) Geändert am",
                   "SAP Nummer", "Firmenname", "Ort"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    recs, _ = parse_excel(sheet([[GUID, "2026-07-03 07:53:00", "0005000001", "Alte AG", "Melle"]]))
    assert recs[0]["crm_id"] == GUID
    upsert_companies(recs)
    s = temp_db.SessionLocal()
    c = s.scalar(select(Company))
    assert c.crm_id == GUID and c.crm_modified_on is not None
    original_id = c.id
    s.close()

    # SAME GUID, but renamed AND re-numbered in the CRM -> must UPDATE that row
    recs2, _ = parse_excel(sheet([[GUID, "2026-07-30 09:00:00", "0009999999", "Neue AG", "Osnabrück"]]))
    upsert_companies(recs2)
    s = temp_db.SessionLocal()
    rows = list(s.scalars(select(Company)))
    assert len(rows) == 1, "matched on crm_id, so no duplicate row"
    assert rows[0].id == original_id
    assert rows[0].sap_number == "0009999999" and rows[0].city == "Osnabrück"
    assert rows[0].crm_id == GUID
    s.close()

    # a DIFFERENT GUID with the same name is a different company -> new row
    recs3, _ = parse_excel(sheet([["ffffffff-0000-0000-0000-000000000001",
                                   "2026-07-30 09:00:00", None, "Neue AG Zweite", "Melle"]]))
    upsert_companies(recs3)
    s = temp_db.SessionLocal()
    assert len(list(s.scalars(select(Company)))) == 2
    s.close()


def test_crm_showroom_ingest_joins_on_crm_id(temp_db):
    """CRM showroom rows join to companies through the Dataverse GUID. A row whose
    dealer GUID is unknown locally must still be STORED (company_id NULL, counted
    as unmatched) — dropping it would hide that the company master is incomplete.
    Re-ingesting the same pull must update, never duplicate."""
    from adwatch import crm_sync
    from adwatch.models import Company, CrmShowroom
    from sqlalchemy import select

    GUID_A = "aaaaaaaa-0000-0000-0000-000000000001"
    s = temp_db.SessionLocal()
    s.add(Company(name="Schaurraum Partner", country="DE", crm_id=GUID_A))
    s.commit(); s.close()

    recs = [
        {"crm_id": "e1", "dealer_crm_id": GUID_A, "product_family": "Glas-Faltwand",
         "product": "SL 25", "installed_on": "2024-05-01"},
        {"crm_id": "e2", "dealer_crm_id": GUID_A, "product_family": "Wintergarten",
         "product": "SDL Atrium plus", "installed_on": None},
        {"crm_id": "e3", "dealer_crm_id": "unknown-guid", "product_family": "cero"},
        {"crm_id": "", "dealer_crm_id": GUID_A, "product_family": "Ignoriert"},   # no key
    ]
    r = crm_sync.upsert_showrooms(recs)
    assert r == {"received": 4, "inserted": 3, "updated": 0,
                 "matched": 2, "unmatched": 1, "skipped": 1}

    s = temp_db.SessionLocal()
    rows = {x.crm_id: x for x in s.scalars(select(CrmShowroom))}
    cid = s.scalar(select(Company.id))
    assert rows["e1"].company_id == cid and rows["e1"].product == "SL 25"
    assert rows["e1"].installed_on.isoformat() == "2024-05-01"
    assert rows["e3"].company_id is None            # kept, but unresolved
    s.close()

    # idempotent: same pull again updates in place
    r2 = crm_sync.upsert_showrooms(recs)
    assert r2["inserted"] == 0 and r2["updated"] == 3
    s = temp_db.SessionLocal()
    assert len(list(s.scalars(select(CrmShowroom)))) == 3
    s.close()

    o = crm_sync.showroom_overview()
    assert o["rows"] == 3 and o["matched_dealers"] == 1
    assert dict(o["families"])["Glas-Faltwand"] == 1
    assert o["per_company"][cid] == ["Glas-Faltwand", "Wintergarten"]


def test_country_code_mapping():
    """The CRM's Land column holds full names; Company.country must end up as an
    ISO-2 code because it is what the ad lookups pass as their country param."""
    from adwatch.customers import _country_code
    assert _country_code("Spanien") == "ES"
    assert _country_code("Deutschland") == "DE"
    assert _country_code("Portugal") == "PT"
    assert _country_code("Österreich") == "AT"
    assert _country_code("  spanien ") == "ES"          # normalised
    assert _country_code("ES") == "ES" and _country_code("es") == "ES"
    assert _country_code("Absurdistan") is None          # unknown -> don't guess
    assert _country_code(None) is None and _country_code("") is None


def test_import_keeps_rows_without_sap_and_reports_them(temp_db):
    """Rows with no SAP Nummer must be imported (matched by Firmenname) and
    REPORTED, not silently dropped — a real CRM export had 917 of 1,000 such
    rows, which previously vanished without a word. Rows with neither key are
    skipped, also with a count."""
    import io
    import openpyxl
    from adwatch.customers import parse_excel, upsert_companies
    from adwatch.models import Company
    from sqlalchemy import select

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SAP Nummer", "Firmenname", "Land", "Ort"])
    ws.append(["0005000001", "Mit SAP GmbH", "Deutschland", "Melle"])
    ws.append([None, "Ohne SAP SL", "Spanien", "Barcelona"])       # name-keyed
    ws.append([None, None, "Spanien", None])                        # no key -> skipped
    buf = io.BytesIO(); wb.save(buf)

    records, warnings = parse_excel(buf.getvalue())
    assert len(records) == 2                                        # not 1
    assert any("no SAP Nummer" in w for w in warnings)
    assert any("skipped" in w for w in warnings)

    upsert_companies(records)
    s = temp_db.SessionLocal()
    rows = {c.name: c for c in s.scalars(select(Company))}
    assert set(rows) == {"Mit SAP GmbH", "Ohne SAP SL"}
    assert rows["Mit SAP GmbH"].country == "DE"
    assert rows["Ohne SAP SL"].country == "ES"                      # not the "DE" default
    s.close()

    # re-importing the same SAP-less row must UPDATE it, not duplicate it
    upsert_companies(records)
    s = temp_db.SessionLocal()
    assert len(list(s.scalars(select(Company)))) == 2
    s.close()


def test_customer_state_derivation():
    """Lifecycle from the Umsatz columns: active = buys now AND before, new =
    first revenue this year, lapsed = bought before not now, never = nothing."""
    from adwatch.customers import derive_customer_state as d
    assert d(50000, 80000, 0, 0, 0) == "active"
    assert d(50000, None, None, None, None) == "new"
    assert d(0, 80000, 0, 0, 0) == "lapsed"
    assert d(None, None, None, None, 100) == "lapsed"
    assert d(None, None, None, None, None) == "never"
    assert d(0, 0, 0, 0, 0) == "never"


def test_icp_parsers():
    from adwatch.insights.icp import parse_employee_count, size_bucket, age_bucket, plz_zone
    import datetime as dt
    assert parse_employee_count("15 Mitarbeiter") == 15
    assert parse_employee_count("drei Angestellten") == 3
    assert parse_employee_count("10-15 Mann") == 12
    assert parse_employee_count("Familienbetrieb") is None
    assert size_bucket("15 Mitarbeiter") == "10-19"
    assert size_bucket(None) is None
    today = dt.date(2026, 7, 29)
    assert age_bucket(2020, today) == "<10 Jahre"
    assert age_bucket(1955, today) == "50+ Jahre"
    assert age_bucket(None) is None
    # Namespaced by country: DE, FR, ES and IT all use five digits, so a
    # country-blind zone equated Barcelona 08036 with German 0xxxx (Saxony).
    # 6.023 non-German companies were carrying a German zone, 1.700 Spanish.
    assert plz_zone("49134", "DE") == "DE 4x"
    assert plz_zone("08036", "ES") == "ES 0x"
    assert plz_zone("75008", "FR") == "FR 7x"
    assert plz_zone("08036", "ES") != plz_zone("08036", "DE")
    assert plz_zone("49134") == "DE 4x", "ohne Land bleibt DE die Annahme"
    assert plz_zone("123", "DE") is None
    assert plz_zone("1010", "AT") is None, "vierstellig — keine Zone, nicht geraten"


def test_icp_winners_never_include_consumers(temp_db):
    """Regression: 'Private Endkunden' who bought something were silently making
    up ~a third of the default winners set, pulling the TRADE-PARTNER profile
    toward consumer traits. They must never define the profile — not by default,
    and not when the caller passes their own filter — unless hand-picked by id."""
    from adwatch import customers
    from adwatch.insights import icp
    from adwatch.models import Company
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    for i in range(6):
        s.add(Company(name=f"Haendler {i}", country="DE", segment="Handel",
                      postal_code="49134", revenue_y0=90000, revenue_y1=80000))
    for i in range(6):
        s.add(Company(name=f"Verbraucher {i}", country="DE", segment="Private Endkunden",
                      postal_code="10115", revenue_y0=400, revenue_y1=300))
    s.commit()
    for c in s.scalars(select(Company)):
        c.customer_state = customers.derive_customer_state(
            c.revenue_y0, c.revenue_y1, c.revenue_y2, c.revenue_y3, c.revenue_y4)
    s.commit()
    consumer_ids = [c.id for c in s.scalars(select(Company).where(Company.segment == "Private Endkunden"))]
    s.close()

    # default winners: consumers excluded
    p = icp.build_profile(None)
    assert p["winners_count"] == 6
    assert "Private Endkunden" not in (p["features"]["segment"]["shares"] or {})

    # caller-supplied filter: consumers still excluded, caller's own exclusion kept
    p2 = icp.build_profile({"customer_state": ["active", "new"], "exclude_segment": ["Handel"]})
    assert set(p2["winners_filter"]["exclude_segment"]) == {"Handel", "Private Endkunden"}
    assert p2["winners_count"] == 0

    # A hand-picked id list is NO LONGER an override. Consumers are out of scope
    # globally (adwatch/scope.py) precisely because "explicit enough" doors kept
    # letting 36% of the base back into counts nobody meant to include them in.
    p3 = icp.build_profile({"ids": consumer_ids})
    assert p3["winners_count"] == 0

    # deliberate consumer profiling now goes through the one named override
    p4 = icp.build_profile({"ids": consumer_ids, "include_consumers": True})
    assert p4["winners_count"] == len(consumer_ids)


def test_intercompany_never_winner_never_target(temp_db):
    """Own-group entities (Linara, NanaWall, Solarlux Vertriebsbüros) appear as
    ordinary large customers. Confirmed live: Linara Kaufbeuren (EUR 1.35M) was
    the biggest 'customer' and ranked #3 in the target list, and 5 Linara rows
    were shaping the profile. They must never be a winner nor a target."""
    from adwatch import customers
    from adwatch.insights import icp
    from adwatch.models import Company
    from sqlalchemy import select

    assert customers.looks_intercompany("Linara Kaufbeuren GmbH")
    assert customers.looks_intercompany("Mike Morgenstern Solarlux Vertriebsbüro")
    assert not customers.looks_intercompany("Serin Bauelemente")

    s = temp_db.SessionLocal()
    # deliberately VARIED winners: if every winner shared one value, the feature
    # would be dropped as "nicht trennscharf" and nothing would be comparable.
    # 30+ winners because the profile guard rejects anything smaller as noise.
    for i in range(30):
        s.add(Company(name=f"Echter Haendler {i}", country="DE",
                      segment="Handel" if i < 20 else "Verarbeiter",
                      postal_code="49134" if i < 20 else "80331",
                      revenue_y0=50000, revenue_y1=40000))
    s.add(Company(name="Linara Teststadt GmbH", country="DE", segment="Handel",
                  postal_code="49134", revenue_y0=1350910, revenue_y1=900000))
    s.commit()
    for c in s.scalars(select(Company)):
        c.customer_state = customers.derive_customer_state(
            c.revenue_y0, c.revenue_y1, c.revenue_y2, c.revenue_y3, c.revenue_y4)
    s.commit()
    s.close()

    assert customers.flag_intercompany() == 1          # only the Linara row flagged
    p = icp.build_profile(None)
    assert p["winners_count"] == 30                     # the group company is out

    icp.apply_profile(None, name="t")
    s = temp_db.SessionLocal()
    linara = s.scalar(select(Company).where(Company.name.like("Linara%")))
    assert linara.is_intercompany is True
    assert linara.target_score is None                  # never on the call list
    assert linara.fit_score is not None                 # but still described
    s.close()


def test_architects_are_scored_on_projects_not_revenue(temp_db):
    """The business fact that broke the ICP: architects specify, they never buy.
    All 808 architect accounts converted at 0%, which made the headline lift look
    like signal when it was really 'architects aren't dealers'.

    The fix is a SECOND outcome measure: an architect's value is the project volume
    they specify, which lives on the opportunity and never on their account."""
    from adwatch import prescriptors
    from adwatch.models import Company, CrmOpportunity

    s = temp_db.SessionLocal()
    star = Company(name="Star Architekten", segment="Architekten", country="DE",
                   crm_id="AAAA1111-0000-0000-0000-000000000001", revenue_y0=0)
    quiet = Company(name="Stille Architekten", segment="Architekten", country="DE",
                    crm_id="BBBB2222-0000-0000-0000-000000000002", revenue_y0=0)
    dealer = Company(name="Haendler GmbH", segment="Handel", country="DE",
                     crm_id="CCCC3333-0000-0000-0000-000000000003", revenue_y0=50000)
    s.add_all([star, quiet, dealer]); s.commit()
    star_id, quiet_id = star.id, quiet.id
    s.add_all([
        # two won projects the architect specified, ordered by the dealer
        CrmOpportunity(crm_id="o1", state="gewonnen", order_value=120000.0,
                       architect_crm_id="aaaa1111-0000-0000-0000-000000000001",
                       parent_account_crm_id="cccc3333-0000-0000-0000-000000000003",
                       building_type="Bürogebäude", created_on=dt.datetime(2025, 3, 1)),
        CrmOpportunity(crm_id="o2", state="gewonnen", order_value=80000.0,
                       architect_crm_id="aaaa1111-0000-0000-0000-000000000001",
                       building_type="Villen", created_on=dt.datetime(2026, 1, 9)),
        CrmOpportunity(crm_id="o3", state="verloren", order_value=40000.0,
                       architect_crm_id="aaaa1111-0000-0000-0000-000000000001",
                       created_on=dt.datetime(2025, 6, 1)),
        # the CRM sometimes puts architect AND end customer on the same account —
        # that must count as ONE project for them, not two
        CrmOpportunity(crm_id="o4", state="offen", order_value=10000.0,
                       architect_crm_id="bbbb2222-0000-0000-0000-000000000002",
                       end_customer_crm_id="bbbb2222-0000-0000-0000-000000000002",
                       created_on=dt.datetime(2026, 2, 2)),
    ])
    s.commit(); s.close()

    star_inf = prescriptors.influence_for("AAAA1111-0000-0000-0000-000000000001")
    assert star_inf["projects"] == 3 and star_inf["won"] == 2 and star_inf["lost"] == 1
    assert star_inf["value_won"] == 200000.0
    assert star_inf["win_rate"] == round(2 / 3, 3)
    assert star_inf["roles"] == ["architect"]
    assert set(star_inf["building_types"]) == {"Bürogebäude", "Villen"}

    # the duplicate-role project counts once
    quiet_inf = prescriptors.influence_for("BBBB2222-0000-0000-0000-000000000002")
    assert quiet_inf["projects"] == 1
    assert sorted(quiet_inf["roles"]) == ["architect", "end_customer"]
    # nothing decided yet -> win_rate is None, NOT 0. An untested architect must
    # not rank below a real 10% performer.
    assert quiet_inf["win_rate"] is None

    # a company with no projects gets the empty shape, never a KeyError
    assert prescriptors.influence_for(None)["projects"] == 0
    assert prescriptors.influence_for("does-not-exist")["projects"] == 0

    # the ranking surfaces the architect with zero revenue ABOVE the paying
    # dealer, which a revenue-only list can never do
    ranked = prescriptors.prescriptor_targets()
    assert ranked[0]["company_id"] == star_id
    assert ranked[0]["revenue_y0"] in (0, 0.0, None)
    ids = [r["company_id"] for r in ranked]
    assert quiet_id in ids

    ov = prescriptors.overview()
    assert ov["opportunities"] == 4 and ov["usable"] is True
    assert ov["by_segment"]["Architekten"]["with_projects"] == 2


def test_crm_fetch_sends_select_as_a_string_and_reads_labels(temp_db, monkeypatch):
    """Two things verified against the live flow, pinned here.

    $select must be a COMMA-SEPARATED STRING — Power Automate passes it straight
    into the Dataverse query, and a JSON array silently returns no columns.

    And picklists arrive as an integer PLUS a FormattedValue label, so the app
    needs no option-set mapping. Confirmed live: 102 / "Architekten"."""
    from sqlalchemy import select
    from adwatch import crm_accounts, flows
    from adwatch.models import Company

    sent = {}

    def fake_post(role, payload, **kw):
        sent["role"] = role
        sent["payload"] = payload
        return {"value": [{
            "accountid": "11111111-1111-1111-1111-111111111111",
            "name": "Interessent GmbH",
            "modifiedon": "2026-08-05T12:00:00Z",
            "sl_customer_segment": 102,
            "sl_customer_segment@OData.Community.Display.V1.FormattedValue": "Architekten",
            "sl_customer_or_prospect": 102690000,
            "statecode": 0,
        }]}
    monkeypatch.setattr(flows, "post", fake_post)

    rows = crm_accounts.fetch_accounts("statecode eq 0", top=3)
    assert sent["role"] == "crm_query"
    assert isinstance(sent["payload"]["select"], str), "$select must be a string"
    assert "accountid" in sent["payload"]["select"].split(",")
    assert sent["payload"]["top"] == 3
    assert len(rows) == 1

    res = crm_accounts.upsert_accounts(rows)
    assert res["inserted"] == 1
    with temp_db.SessionLocal() as s:
        c = s.scalar(select(Company).where(Company.name == "Interessent GmbH"))
        assert c.segment == "Architekten"        # the LABEL, not 102

    # a scope must never run without a filter, or it would pull the whole org
    with pytest.raises(ValueError):
        crm_accounts.load_scope("")
    # the prospect scope is the one the ICP needs — keep it addressable by name
    assert "prospects" in crm_accounts.SCOPES
    assert "102690000" in crm_accounts.SCOPES["prospects"][1]

    # delta uses our newest modifiedon as the watermark
    out = crm_accounts.sync_delta()
    assert "modifiedon gt" in (out["filter"] or "")


def test_crm_sync_never_overwrites_what_we_paid_for(temp_db):
    """The whole point of the ownership map. CRM owns master data; a sync must not
    touch enrichment, locked pages, scores or ad history — that is the expensive
    half of the database and Dataverse has no opinion about any of it."""
    from sqlalchemy import select
    from adwatch import crm_accounts
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    c = Company(
        crm_id="A7DBC4F6-A2F9-40CC-BEB9-0000E0EE6272", name="Alter Name GmbH",
        country="DE", city="Altstadt", segment="Handel", revenue_y0=100.0,
        # everything below is ours and must survive untouched
        description="Von uns angereichert.", products=["Fenster"], founded_year=1999,
        employee_hint="12 Mitarbeiter", enrichment_status="enriched",
        page_id="123456", page_name="Alte Seite", resolution_status="locked",
        fit_score=88.0, opportunity_score=42.0, target_score=61.0,
        website_domain="alt.example")
    s.add(c); s.commit()
    cid = c.id
    s.close()

    # same account, matched on the GUID in a DIFFERENT case
    res = crm_accounts.upsert_accounts([{
        "accountid": "a7dbc4f6-a2f9-40cc-beb9-0000e0ee6272",
        "modifiedon": "2026-07-27T20:37:19Z",
        "name": "Neuer Name GmbH", "accountnumber": "0005009967",
        "address1_line1": "Vogelherdbogen 27", "address1_postalcode": "88069",
        "address1_city": "Tettnang", "address1_country": "Spanien",
        "telephone1": "+4975428292", "emailaddress1": "neu@example.de",
        "sl_customer_segment": 101,
        "sl_customer_segment@OData.Community.Display.V1.FormattedValue": "Verarbeiter",
        "slx_revenue_current_year": 22749.0, "slx_revenue_current_year_1": 80728.0,
        "statecode": 0,
    }])
    assert res == {**res, "inserted": 0, "updated": 1, "skipped": 0}

    s = temp_db.SessionLocal()
    c = s.get(Company, cid)
    # CRM won on master data
    assert c.name == "Neuer Name GmbH" and c.city == "Tettnang"
    assert c.sap_number == "0005009967" and c.postal_code == "88069"
    assert c.country == "ES"                       # "Spanien" -> ES via markets
    assert c.segment == "Verarbeiter"              # LABEL, never the raw 101
    assert c.revenue_y0 == 22749.0
    assert c.customer_state == "active"            # recomputed, not synced
    # ...and every local field is untouched
    assert c.description == "Von uns angereichert." and c.products == ["Fenster"]
    assert c.employee_hint == "12 Mitarbeiter" and c.enrichment_status == "enriched"
    assert c.page_id == "123456" and c.resolution_status == "locked"
    assert (c.fit_score, c.opportunity_score, c.target_score) == (88.0, 42.0, 61.0)
    assert c.website_domain == "alt.example"
    s.close()

    # a raw picklist with NO label must not write an integer into a name column
    crm_accounts.upsert_accounts([{"accountid": "a7dbc4f6-a2f9-40cc-beb9-0000e0ee6272",
                                   "sl_customer_segment": 999}])
    with temp_db.SessionLocal() as s2:
        assert s2.get(Company, cid).segment == "Verarbeiter"

    # a record without accountid is skipped, not guessed at
    assert crm_accounts.upsert_accounts([{"name": "Ohne GUID"}])["skipped"] == 1
    # inserts can be refused, so a delta sync cannot silently widen the base
    assert crm_accounts.upsert_accounts(
        [{"accountid": "11111111-1111-1111-1111-111111111111", "name": "Neu"}],
        allow_insert=False)["inserted"] == 0
    # ...and allowed when asked for
    assert crm_accounts.upsert_accounts(
        [{"accountid": "11111111-1111-1111-1111-111111111111", "name": "Neu"}]
    )["inserted"] == 1

    # the ownership map and the protected set must never overlap
    written = (set(crm_accounts.CRM_OWNED_SCALARS.values())
               | set(crm_accounts.CRM_OWNED_PICKLISTS.values())
               | set(crm_accounts.CRM_OWNED_REVENUE.values()))
    assert not (written & crm_accounts.LOCAL_OWNED)

    # the watermark drives the delta filter
    assert crm_accounts.watermark().endswith("Z")
    assert "accountid" in crm_accounts.select_fields()


def test_markets_are_data_not_code():
    """Adding a market used to need three code edits (country aliases, legal-page
    term, search language) in two modules. Missing one failed SILENTLY — that is
    how 982 Spanish companies were imported as DE. Now one YAML file drives all
    three, and a successor adds a market without Python."""
    from adwatch import markets

    markets.reload()
    codes = markets.known_codes()
    assert {"DE", "ES", "PT", "AT", "CH", "FR"} <= set(codes)

    # the Spain bug, pinned: every spelling a source might use resolves
    for spelling in ("Spanien", "españa", "espana", "SPAIN", "es", "ES"):
        assert markets.code_for(spelling) == "ES", spelling
    # umlaut and its transliteration both work
    assert markets.code_for("österreich") == markets.code_for("Oesterreich") == "AT"
    # NO must survive YAML's boolean coercion of a bare NO key
    assert "NO" in codes and markets.code_for("Norwegen") == "NO"

    # an unknown name returns None so the caller can KEEP the old value rather
    # than silently defaulting to DE
    assert markets.code_for("Slowenien") is None
    assert markets.code_for("") is None
    # but an unlisted 2-letter code passes through instead of being dropped
    assert markets.code_for("si") == "SI"

    # the per-market behaviour that used to be hardcoded elsewhere
    assert markets.search_lang("ES") == "es"
    assert markets.legal_page_term("ES") == "aviso legal"      # not "Impressum"
    assert markets.legal_page_term("PT") == "contactos"
    assert markets.legal_page_term("DE") == "Impressum"
    # unknown market degrades to a usable default rather than raising
    assert markets.search_lang("XX") and markets.legal_page_term("XX")

    # every market must be complete, or a new entry could half-work
    for code, spec in markets.all_markets().items():
        assert len(code) == 2 and code.isalpha() and code.isupper(), code
        assert spec["aliases"] and spec["search_lang"] and spec["legal_page"], code


def test_flow_registry_is_configurable_and_backward_compatible(temp_db, monkeypatch):
    """Flows are addressed by ROLE, so a second integration point (the CRM query
    proxy) needs no new constant, timeout or error convention. And an install that
    still has only the old POWER_AUTOMATE_WEBHOOK_URL must keep working."""
    from adwatch import config, flows

    # unknown role fails loudly rather than silently doing nothing
    with pytest.raises(ValueError):
        flows.url_for("does_not_exist")

    # nothing configured -> a message that says WHERE to fix it
    monkeypatch.setattr(config, "resolve_setting", lambda k: "")
    assert not flows.is_configured("report_email")
    msg = flows.missing_message("crm_query")
    assert "FLOW_URL_CRM_QUERY" in msg and "Einstellungen" in msg
    with pytest.raises(RuntimeError):
        flows.post("crm_query", {})

    # legacy key alone still drives the email role (upgrade path)
    monkeypatch.setattr(config, "resolve_setting",
                        lambda k: "https://legacy.example/flow"
                        if k == "POWER_AUTOMATE_WEBHOOK_URL" else "")
    assert flows.is_configured("report_email")
    assert flows.url_for("report_email") == "https://legacy.example/flow"
    # ...but the legacy key must NOT leak into other roles
    assert not flows.is_configured("crm_query")

    # the new key wins when both are set
    monkeypatch.setattr(config, "resolve_setting",
                        lambda k: {"FLOW_URL_REPORT_EMAIL": "https://new.example/f",
                                   "POWER_AUTOMATE_WEBHOOK_URL": "https://legacy.example/f"}.get(k, ""))
    assert flows.url_for("report_email") == "https://new.example/f"

    # status() reports the integration points for Settings / diagnostics
    st = {s["role"]: s for s in flows.status()}
    assert st["report_email"]["configured"] is True
    assert st["crm_query"]["configured"] is False
    assert st["crm_query"]["key"] == "FLOW_URL_CRM_QUERY"

    # every role must have a settings entry, or Settings could never configure it
    for role, (key, _, _) in flows.FLOW_ROLES.items():
        assert key in config._SPEC_BY_KEY, f"{role} has no SETTINGS_SPEC entry"
        assert config._SPEC_BY_KEY[key]["secret"] is True, f"{key} must be masked"


def test_thinly_known_features_do_not_score(temp_db):
    """diagnose() refused to trust a feature below 15% coverage, but fit_for used
    it anyway — so Betriebsgröße, known for 3% of winners (a distribution built
    from ~20 companies), was shaping EVERY company's fit score. Warning about a
    number and then scoring with it is worse than not having it at all."""
    from adwatch.insights.icp import fit_for

    profile = {
        "weights": {"segment": 1.0, "size_bucket": 1.0},
        "features": {
            # solidly known, and discriminating
            "segment": {"coverage": 1.0, "shares": {"Handel": 0.7, "Verarbeiter": 0.3},
                        "lifts": {"Handel": 1.8, "Verarbeiter": 0.6}},
            # the live case: a spread that LOOKS informative but rests on ~20 rows
            "size_bucket": {"coverage": 0.03,
                            "shares": {"20-49": 0.5, "10-19": 0.3, "50+": 0.2},
                            "lifts": {"20-49": 2.5, "10-19": 1.1, "50+": 0.7}},
        },
    }
    # a company matching ONLY the thin feature has nothing comparable left
    fit_thin, bd_thin = fit_for({"segment": None, "size_bucket": "20-49"}, profile)
    assert fit_thin is None and bd_thin == []

    # and the thin feature cannot inflate a score that rests on the solid one
    fit_a, _ = fit_for({"segment": "Handel", "size_bucket": "20-49"}, profile)
    fit_b, _ = fit_for({"segment": "Handel", "size_bucket": None}, profile)
    assert fit_a == fit_b, "size_bucket must not move the score at 3% coverage"

    # raise its coverage above the floor and it starts counting
    profile["features"]["size_bucket"]["coverage"] = 0.6
    fit_c, bd_c = fit_for({"segment": "Handel", "size_bucket": "20-49"}, profile)
    assert {b["feature"] for b in bd_c} >= {"segment", "size_bucket"}
    assert fit_c is not None


def test_relevance_sorts_by_rank_not_alphabet(temp_db):
    """Solarlux-Relevanz is ordinal, and its labels sort alphabetically in exactly
    the wrong order — g(ering) < h(och) < m(ittel). A plain text sort would head
    the "best architects first" list with the worst-fitting offices, so the sort
    ranks the labels and leaves ungraded rows at the bottom either way."""
    from sqlalchemy import select
    from adwatch.customers import _apply_sort
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add_all([
        Company(name="Villa-Buero", segment="Architekten", solarlux_relevance="hoch"),
        Company(name="Innenausbau", segment="Architekten", solarlux_relevance="gering"),
        Company(name="Hochbau", segment="Architekten", solarlux_relevance="mittel"),
        Company(name="Ungeprueft", segment="Architekten", solarlux_relevance=None),
    ])
    s.commit(); s.close()

    def order(direction):
        with temp_db.SessionLocal() as s2:
            stmt = _apply_sort(select(Company), "solarlux_relevance", direction)
            return [c.name for c in s2.scalars(stmt)]

    assert order("desc") == ["Villa-Buero", "Hochbau", "Innenausbau", "Ungeprueft"]
    assert order("asc") == ["Innenausbau", "Hochbau", "Villa-Buero", "Ungeprueft"]


def test_consumers_are_excluded_from_every_count(temp_db):
    """Private Endkunden are 36% of the base and none of them will ever run an ad
    campaign, so including them made every ratio wrong ("14 of 4618"). They stay in
    the database but must not reach a single count, list or report — and no filter
    combination may put them back."""
    from sqlalchemy import select
    from adwatch import scope, services
    from adwatch.customers import _apply_filters
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add_all([
        Company(name="Partner Handel", segment="Handel", country="DE"),
        Company(name="Partner Verarbeiter", segment="Verarbeiter", country="DE"),
        Company(name="Herr Müller", segment="Private Endkunden", country="DE"),
        Company(name="Frau Schmidt", segment="Private Endkunden", country="DE"),
        Company(name="Unklar", segment=None, country="DE"),   # unknown != consumer
    ])
    s.commit(); s.close()

    def names(f):
        with temp_db.SessionLocal() as s2:
            return {c.name for c in s2.scalars(_apply_filters(select(Company), f))}

    partners = {"Partner Handel", "Partner Verarbeiter", "Unklar"}
    assert names({}) == partners                       # no filter at all
    assert names({"country": ["DE"]}) == partners      # an unrelated filter
    # a hand-picked id list must not smuggle them back in
    with temp_db.SessionLocal() as s2:
        all_ids = [c.id for c in s2.scalars(select(Company))]
    assert names({"ids": all_ids}) == partners
    # nor may an explicit "don't exclude anything" style filter
    assert names({"exclude_segment": []}) == partners

    # the read model the dashboard KPIs are built from
    assert {c["name"] for c in services.list_companies()} == partners
    assert {m["company"] for m in services.latest_metrics()} == partners
    # even when consumer ids are passed in explicitly
    assert {m["company"] for m in services.latest_metrics(all_ids)} == partners

    # a NULL segment is kept — unknown is not the same as consumer
    assert scope.is_in_scope(None) and scope.is_in_scope("Handel")
    assert not scope.is_in_scope("Private Endkunden")

    # the deliberate ways in still work, so the data is not unreachable
    assert "Herr Müller" in names({"include_consumers": True})
    assert names({"segment": ["Private Endkunden"]}) == {"Herr Müller", "Frau Schmidt"}


def test_legal_form_must_occur_in_the_source_text():
    """A Spanish S.L. was stored as the GERMAN form 'e.K.' — the old prompt offered
    a closed list of German forms, so the model substituted the nearest one. That
    is a false fact about a legal entity, not a translation. The form now only
    survives if it actually appears in the crawled text."""
    from adwatch.enrich.extract import _legal_form_in_text

    es = "ALLKONZEPT S.L. · Aviso legal · Calle Mayor 1, Mallorca"
    assert _legal_form_in_text("S.L.", es) == "S.L."
    assert _legal_form_in_text("e.K.", es) is None          # the fabrication is dropped
    # punctuation and spacing differ between sites, so matching ignores them
    assert _legal_form_in_text("S.L.", "Aluminios ALSABEN, SL — Las Palmas") == "S.L."
    assert _legal_form_in_text("Lda.", "Afcamoes Solutions LDA, Porto") == "Lda."
    # a German company keeps its German form
    assert _legal_form_in_text("GmbH", "Muster Fenster GmbH, Osnabrück") == "GmbH"
    assert _legal_form_in_text(None, es) is None
    assert _legal_form_in_text("GmbH", "") is None

    # The match must be anchored on word boundaries. A first attempt at this guard
    # compared punctuation-stripped strings, so "e.K." -> "ek" matched inside
    # ordinary words and three Spanish S.L. companies kept their fake German form.
    for word in ("Unsere Projekte in Mallorca", "perfekte Lösungen",
                 "Elektrische Antriebe", "Rekord"):
        assert _legal_form_in_text("e.K.", word) is None, word
    assert _legal_form_in_text("AG", "Wir sind Ihr Partner am Tag und Nacht") is None
    assert _legal_form_in_text("SL", "Alle Schlösser und Beschläge") is None
    # but a standalone occurrence still counts, however it is punctuated
    assert _legal_form_in_text("e.K.", "Fenster Meier e. K. — Impressum") == "e.K."
    assert _legal_form_in_text("AG", "Glas Trösch AG, Bützberg") == "AG"


def test_reenrichment_can_retract_a_fact_but_not_a_human_edit(temp_db, monkeypatch):
    """A corrected extractor is useless if the wrong value cannot be removed.
    Stored fields were only ever merged, and null results were skipped, so three
    Spanish S.L. companies kept a fabricated German 'e.K.' long after the
    extractor started returning null for it. A human's edit still wins."""
    from sqlalchemy import select
    from adwatch.enrich import extract, fetchpage, service as enrich_service
    from adwatch.models import Company, CompanyEnrichment

    s = temp_db.SessionLocal()
    c = Company(name="Retract SL", country="ES", website_domain="retract.es")
    s.add(c); s.commit()
    cid = c.id
    s.close()

    monkeypatch.setattr(fetchpage, "page_bundle",
                        lambda d: {"text": "Retract SL, Alicante", "pages": [d]})

    # run 1: the old, wrong extraction
    monkeypatch.setattr(extract, "extract_facts", lambda t, **kw: {
        "description_de": "Baut Fenster.", "legal_form": "e.K.",
        "employee_hint": "Un gran equipo", "products": ["Fenster"],
        "founded_year": None, "service_area": None, "mentions_solarlux": False,
        "competitor_brands": [], "assessment_de": None, "evidence": {}, "llm_model": "m"})
    enrich_service.enrich_company(cid, allow_search=False)

    s = temp_db.SessionLocal()
    e = s.scalar(select(CompanyEnrichment).where(CompanyEnrichment.company_id == cid))
    assert e.fields["legal_form"] == "e.K."
    assert s.get(Company, cid).employee_hint == "Un gran equipo"
    # a human corrects the size by hand
    e.provenance = {**(e.provenance or {}), "employee_hint": {"source": "manual"}}
    s.commit(); s.close()

    # run 2: the corrected extractor returns null for both
    monkeypatch.setattr(extract, "extract_facts", lambda t, **kw: {
        "description_de": "Baut Fenster und Türen.", "legal_form": None,
        "employee_hint": None, "products": ["Fenster", "Türen"],
        "founded_year": None, "service_area": None, "mentions_solarlux": False,
        "competitor_brands": [], "assessment_de": None, "evidence": {}, "llm_model": "m"})
    enrich_service.enrich_company(cid, allow_search=False)

    s = temp_db.SessionLocal()
    e = s.scalar(select(CompanyEnrichment).where(CompanyEnrichment.company_id == cid))
    comp = s.get(Company, cid)
    assert "legal_form" not in e.fields          # retracted
    assert e.fields["employee_hint"] == "Un gran equipo"   # human edit survives
    assert e.fields["description_de"] == "Baut Fenster und Türen."   # refreshed
    assert comp.products == ["Fenster", "Türen"]
    s.close()


def test_report_events_record_creation_and_delivery(temp_db, monkeypatch, tmp_path):
    """Creating and sending a report must leave an audit row, so 'did the mail go
    out?' is answerable after a crash — the question that had no answer before."""
    from adwatch import report_log
    from adwatch.report import write_report_meta

    write_report_meta(str(tmp_path / "adwatch_top5_KW31_2026.pdf"),
                      filters={"country": ["ES"]}, source="manual")
    ev = report_log.history()
    assert len(ev) == 1
    assert ev[0]["kind"] == "created" and ev[0]["report_type"] == "top5"
    assert ev[0]["source"] == "manual" and "ES" in (ev[0]["scope"] or "")

    # a failing send is recorded too, with the real error, and still raises
    import adwatch.emailer as emailer_mod
    from adwatch import config
    monkeypatch.setattr(config, "POWER_AUTOMATE_WEBHOOK_URL", "https://example.invalid/f")
    pdf = tmp_path / "adwatch_top5_KW31_2026.pdf"
    pdf.write_bytes(b"%PDF-1.4 x")

    def _boom(*a, **k):
        raise OSError("network is down")
    monkeypatch.setattr(emailer_mod.requests, "post", _boom)
    with pytest.raises(RuntimeError):
        emailer_mod.send_report_email(str(pdf), recipient=["a@x.de", "b@x.de"],
                                      subject="Bericht", source="pipeline")

    ev = report_log.history()
    fail = ev[0]
    assert fail["kind"] == "send_failed"
    assert fail["recipients"] == ["a@x.de", "b@x.de"]
    assert fail["source"] == "pipeline" and "network is down" in fail["detail"]

    # recording must never be what breaks a send: a broken audit write is swallowed
    monkeypatch.setattr(report_log, "SessionLocal", lambda: (_ for _ in ()).throw(OSError("db gone")))
    report_log.record("created", "x.pdf")          # must not raise


def test_domain_prepass_is_free_and_never_writes_a_verdict(temp_db, monkeypatch):
    """The pre-pass must (a) never call the paid search, (b) write a proven domain,
    and (c) write NOTHING on a miss — 'no website found' is a verdict only the
    full run may reach, and stamping it here would hide the company from a later
    Serper attempt."""
    from adwatch.enrich import service as enrich_service, website_finder
    from adwatch.models import Company, CompanyEnrichment

    def _no_search(*a, **k):
        raise AssertionError("the free pre-pass must never call Serper")
    monkeypatch.setattr(website_finder, "search_candidates", _no_search)

    s = temp_db.SessionLocal()
    hit = Company(name="Mit Mail GmbH", country="DE", email="info@mitmail-gmbh.de",
                  phone="0541 123456", postal_code="49080", city="Osnabrück")
    miss = Company(name="Nur Freemail GmbH", country="DE", email="chef@t-online.de")
    had = Company(name="Hat Schon GmbH", country="DE", website_domain="hatschon.de")
    s.add_all([hit, miss, had]); s.commit()
    hid, mid, did = hit.id, miss.id, had.id
    s.close()

    # the derived domain validates via the phone number on the page
    monkeypatch.setattr(enrich_service.fetchpage, "page_bundle",
                        lambda d: {"text": "Impressum Musterstr. 1, 49080 Osnabrück, Tel. 0541 123456",
                                   "pages": [d]})

    assert enrich_service.derive_domain(did)["status"] == "already_had"
    r_hit = enrich_service.derive_domain(hid)
    r_miss = enrich_service.derive_domain(mid)

    assert r_hit["status"] == "domain_found" and r_hit["website"] == "mitmail-gmbh.de"
    assert r_miss["status"] == "no_domain_derived"

    s = temp_db.SessionLocal()
    assert s.get(Company, hid).website_domain == "mitmail-gmbh.de"
    # the miss is untouched: no website, no status, no enrichment row
    m = s.get(Company, mid)
    assert not m.website_domain
    assert m.enrichment_status in (None, "none")
    assert s.query(CompanyEnrichment).filter_by(company_id=mid).count() == 0
    s.close()


def test_default_step_order_puts_the_free_domain_pass_first():
    """The default order is load-bearing, so it gets its own test.

    Identity's only free AND authoritative tier crawls Company.website_domain.
    Full enrichment fills that column but costs money (Serper + Haiku), while
    deriving a domain from the company's own email is free. So the default must
    be: free domains -> identity -> paid enrichment."""
    from adwatch.jobs import DOMAIN_PREPASS, resolve_step_order

    both = {"enrich": True, "identity": True, "ads": ["meta"], "report": "full",
            "send_to": [1]}
    assert resolve_step_order(both) == [DOMAIN_PREPASS, "identity", "enrich",
                                        "ads", "report", "send"]

    # the pre-pass only earns its place when it can actually feed the identity
    # check — enrichment alone already does Tier 0 internally
    assert resolve_step_order({"enrich": True}) == ["enrich"]
    assert resolve_step_order({"identity": True}) == ["identity"]

    # an explicit order wins, and anything selected but omitted is appended
    # rather than silently dropped
    assert resolve_step_order({"enrich": True, "identity": True,
                               "order": ["enrich", "identity"]}) == ["enrich", "identity"]
    assert resolve_step_order({"enrich": True, "identity": True, "report": "full",
                               "order": ["enrich"]}) == ["enrich", "identity", "report"]

    # 'send' lives under send_to, not a boolean
    assert resolve_step_order({"report": "full", "send_to": [1]}) == ["report", "send"]
    assert "send" not in resolve_step_order({"report": "full"})


def test_pipeline_runs_steps_in_the_working_order(temp_db, monkeypatch):
    """The pipeline must execute domains -> identity -> enrich -> ads -> report ->
    send, skip unchecked steps, and never send without a report. Execution has to
    follow the RESOLVED order, not the order the code happens to be written in."""
    from adwatch import jobs
    from adwatch.models import Company, FetchJob, ReportRecipient

    s = temp_db.SessionLocal()
    a = Company(name="Pipe Eins", country="DE"); b = Company(name="Pipe Zwei", country="DE")
    r = ReportRecipient(name="BD", email="bd@x.de", active=True)
    s.add_all([a, b, r]); s.commit()
    ids, rid = [a.id, b.id], r.id
    s.close()

    calls = []
    import adwatch.enrich.service as enrich_service
    import adwatch.identity.resolver as resolver
    import adwatch.collect.pipeline as coll
    import adwatch.report as report_mod
    import adwatch.emailer as emailer_mod

    monkeypatch.setattr(enrich_service, "derive_domain",
                        lambda cid, **k: calls.append(("domains", cid)) or {"status": "domain_found",
                                                                           "website": "x.de", "source": "email_domain",
                                                                           "validated_by": "phone"})
    monkeypatch.setattr(enrich_service, "enrich_company",
                        lambda cid, **k: calls.append(("enrich", cid)) or {"status": "enriched", "website": "x.de",
                                                                           "website_source": "email_domain",
                                                                           "validated_by": "phone", "fields_found": 3})
    monkeypatch.setattr(resolver, "run_identity_check",
                        lambda cid, **k: calls.append(("identity", cid)) or {"status": "confirmed", "page_name": "P"})
    monkeypatch.setattr(coll, "run_once", lambda company_id=None: calls.append(("meta", company_id)))
    monkeypatch.setattr(coll, "run_once_google", lambda company_id=None: calls.append(("google", company_id)))
    monkeypatch.setattr(jobs, "_plan_units", lambda s_, cids, srcs: [(cids[0], "meta")])
    monkeypatch.setattr(report_mod, "build_report",
                        lambda filters=None: calls.append(("report", None)) or "output/adwatch_report_KW31_2026.pdf")
    monkeypatch.setattr(report_mod, "write_report_meta", lambda *a, **k: None)
    monkeypatch.setattr(report_mod, "subject_for_filename", lambda f: "Bericht")
    monkeypatch.setattr(emailer_mod, "send_report_email",
                        lambda path, recipient=None, subject=None, **k: calls.append(
                            ("send", tuple(recipient))))

    plan = {"enrich": True, "identity": True, "ads": ["meta"], "report": "full", "send_to": [rid]}
    job = jobs.create_pipeline_job(ids, plan, label="t")
    # domains + identity + enrich per company, ads upper bound, report, send
    assert job["total"] == 2 + 2 + 2 + 2 + 1 + 1
    jobs._run_pipeline(job["id"])                     # run inline, no thread

    order = [c[0] for c in calls]
    # the free pass first, then identity, and only then the paid enrichment
    assert order.index("domains") < order.index("identity") < order.index("enrich")
    assert order.index("enrich") < order.index("meta") < order.index("report") < order.index("send")
    assert [c for c in calls if c[0] == "domains"] == [("domains", ids[0]), ("domains", ids[1])]
    assert ("send", ("bd@x.de",)) in calls

    s = temp_db.SessionLocal()
    j = s.get(FetchJob, job["id"])
    assert j.status == "done"
    txt = " ".join(e["text"] for e in (j.log or []))
    for marker in ("Schritt 1/6", "Schritt 2/6", "Schritt 3/6", "Schritt 4/6",
                   "Schritt 5/6", "Schritt 6/6", "Pipeline abgeschlossen"):
        assert marker in txt, marker
    # the log states the order it used, and that it was the default
    assert "Standard-Reihenfolge" in txt
    s.close()

    # an explicit order is obeyed instead of the default
    calls.clear()
    job3 = jobs.create_pipeline_job(ids, {"enrich": True, "identity": True,
                                          "order": ["enrich", "identity"]}, label="t3")
    jobs._run_pipeline(job3["id"])
    o3 = [c[0] for c in calls]
    assert "domains" not in o3
    assert o3.index("enrich") < o3.index("identity")

    # a plan with only some steps skips the rest; sending without a report is refused
    calls.clear()
    job2 = jobs.create_pipeline_job(ids, {"enrich": True}, label="t2")
    jobs._run_pipeline(job2["id"])
    assert {c[0] for c in calls} == {"enrich"}
    with pytest.raises(ValueError):
        jobs.create_pipeline_job(ids, {"send_to": [rid]})
    with pytest.raises(ValueError):
        jobs.create_pipeline_job(ids, {})


def test_report_shows_enriched_profiles_and_marks_the_estimate(temp_db, tmp_path):
    """The Firmenprofile section must render the enriched picture per company —
    WITHOUT any ad data (the Spain case) — and must keep the verified description
    and the AI assessment visibly separate, so an inference can't be read as a
    documented fact."""
    from adwatch.models import Company, CompanyEnrichment
    from adwatch.report import build_report

    s = temp_db.SessionLocal()
    c = Company(name="Cerramientos Test SL", country="ES", segment="Handel",
                sales_channel="Fachhandelsvertrieb", city="Barcelona",
                website_domain="cerramientos-test.es",
                description="Fachbetrieb für Glasfaltwände und Terrassenverglasung.",
                products=["Fenster", "Terrassendach"], founded_year=1998,
                employee_hint="12 Mitarbeiter", enrichment_status="enriched")
    s.add(c); s.flush()
    s.add(CompanyEnrichment(company_id=c.id, status="enriched", fields={
        "description_de": "Fachbetrieb für Glasfaltwände und Terrassenverglasung.",
        "assessment_de": "Dürfte ein Kleinbetrieb mit regionalem Fokus sein; "
                         "der Auftritt wirkt privatkundenorientiert.",
        "products": ["Fenster", "Terrassendach"], "founded_year": 1998,
        "employee_hint": "12 Mitarbeiter", "legal_form": "SL",
        "mentions_solarlux": False, "competitor_brands": ["WAREMA"],
    }))
    s.commit()
    cid = c.id
    s.close()

    out = str(tmp_path / "profile_report.pdf")
    build_report(path=out, filters={"ids": [cid]})

    from pypdf import PdfReader
    text = "\n".join(p.extract_text() or "" for p in PdfReader(out).pages)
    assert "Firmenprofile" in text
    assert "Cerramientos Test SL" in text
    assert "Beschreibung:" in text and "Glasfaltw" in text
    # the inference is present AND labelled as an estimate, not as a fact
    assert "Einschätzung:" in text and "Kleinbetrieb" in text
    assert "keine belegte Angabe" in text
    # the hard fields and the brand signal made it in
    assert "1998" in text and "12 Mitarbeiter" in text
    assert "WAREMA" in text


def test_recipient_tick_state_persists_and_is_not_active(temp_db):
    """Unticking a recipient must survive a reload, and must NOT disable the
    address — a saved weekly definition still has to reach them."""
    from adwatch import services
    from adwatch.models import ReportRecipient

    a = services.add_recipient("a@solarlux.com", "A")
    services.add_recipient("b@solarlux.com", "B")
    # everyone starts ticked, so behaviour is unchanged until the user acts
    assert all(r["preselected"] for r in services.list_recipients())

    services.set_recipient_preselected(a["id"], False)
    rows = {r["email"]: r for r in services.list_recipients()}
    assert rows["a@solarlux.com"]["preselected"] is False
    assert rows["b@solarlux.com"]["preselected"] is True
    # the crucial separation: unticked but still mailable
    assert rows["a@solarlux.com"]["active"] is True
    with temp_db.SessionLocal() as s:
        assert s.get(ReportRecipient, a["id"]).active is True

    # and it toggles back. NB the fixture also seeds the configured default
    # recipient via raw SQL — it must carry preselected=True too, which is why
    # the column needs a server default and not just an ORM-side one.
    services.set_recipient_preselected(a["id"], True)
    assert all(r["preselected"] for r in services.list_recipients())

    with pytest.raises(ValueError):
        services.set_recipient_preselected(9999, False)


def test_ad_products_are_normalised_to_german_families():
    """Ads are written in the local market's language but the report is German.
    The first Spanish run listed 'cerramiento', 'Porche-Verschluss (porch closure)'
    and 'windows and doors (wood, PVC)' as separate products — the same family in
    three languages, plus materials. Everything must fold onto one vocabulary."""
    from adwatch.products import PRODUCT_VOCAB, canonical_products

    # the four strings that actually appeared for one Spanish company
    assert canonical_products([
        "cerramiento", "Porcheschließung/Porche-Verschluss",
        "Porche-Verschluss (porch closure)", "Porcheverglasungen/Terrassenverglasung",
    ]) == ["Terrassenverglasung"]

    assert canonical_products(
        ["windows and doors (wood, wood-aluminium, aluminium, PVC)"]) == ["Fenster", "Türen"]

    # German inflections must not survive as separate families
    assert canonical_products(["Wintergärten", "Wintergarten"]) == ["Wintergarten"]
    # materials are not products, and unmappable free text is dropped, not shown
    assert canonical_products(["Aluminium", "PVC", "Holz"]) == []
    assert canonical_products(["Raumschiffe"]) == []
    # short keys need word boundaries: 'tor' inside 'Motor' is not a Tor
    assert canonical_products(["Motor", "importante"]) == []
    # every result is a member of the shared vocabulary, always
    for out in (canonical_products(["toldos y persianas"]),
                canonical_products(["puertas correderas de cristal"])):
        assert out and all(p in PRODUCT_VOCAB for p in out)

    # the enrichment side imports the very same tuple, so the two can't drift
    from adwatch.enrich.extract import PRODUCT_VOCAB as VOCAB_ENRICH
    assert VOCAB_ENRICH is PRODUCT_VOCAB


def test_top5_report_shows_profiles_and_an_honest_count(temp_db, tmp_path):
    """A 'Top 5' that shows 2 companies must say why, and each company must carry
    its enriched profile — otherwise the report is just ad counts with no context."""
    import datetime as dt
    from adwatch.models import Company, CompanyEnrichment, WeeklyCompanyMetric
    from adwatch.report import build_top5_report

    s = temp_db.SessionLocal()
    wk = dt.date(2026, 7, 27)
    ids = []
    for i in range(2):
        c = Company(name=f"Cerramientos {i} SL", country="ES", segment="Handel",
                    resolution_status="confirmed", website_domain=f"cerr{i}.es",
                    enrichment_status="enriched")
        s.add(c); s.flush()
        s.add(WeeklyCompanyMetric(company_id=c.id, source="meta", week_start=wk,
                                  total_active_ads=8 - i,
                                  products=["cerramiento", "porch closure"]))
        s.add(CompanyEnrichment(company_id=c.id, status="enriched", fields={
            "description_de": "Anbieter von Glasabschlüssen für Veranden und Terrassen.",
            "assessment_de": "Dürfte ein Kleinbetrieb mit regionalem Fokus sein.",
            "products": ["Terrassenverglasung"], "employee_hint": "10 Mitarbeiter",
            "competitor_brands": ["Sunflex"]}))
        ids.append(c.id)
    # 3 more companies in scope that never advertised — they are why it isn't 5
    for i in range(3):
        s.add(Company(name=f"Stille Firma {i}", country="ES", segment="Handel",
                      resolution_status="confirmed"))
    s.commit(); s.close()

    out = str(tmp_path / "top5.pdf")
    build_top5_report(path=out, filters={})

    from pypdf import PdfReader
    raw = "\n".join(p.extract_text() or "" for p in PdfReader(out).pages)
    # collapse whitespace: the PDF line-wraps mid-sentence, which would otherwise
    # make these assertions depend on where the text happens to break
    text = " ".join(raw.split())
    # the headline states the real number instead of promising five
    assert "Top 5" not in text
    assert "Werbetreibende mit aktiven Anzeigen (2)" in text
    assert "die Liste ist nicht gekürzt" in text
    assert "nur 2 von 5 Firmen" in text
    # the enrichment reaches this report type too, fact and inference kept apart
    assert "Beschreibung:" in text and "Glasabschl" in text
    assert "Einschätzung (KI, unbestätigt):" in text and "Kleinbetrieb" in text
    assert "keine belegte Angabe" in text
    assert "10 Mitarbeiter" in text and "Sunflex" in text
    # ad-derived products arrive in German, not as 'cerramiento'/'porch closure'
    assert "cerramiento" not in text and "porch closure" not in text
    assert "Terrassenverglasung" in text


def test_export_includes_enriched_columns(temp_db):
    """A freshly enriched market must not export as bare master data."""
    import io
    import openpyxl
    from adwatch.customers import export_xlsx
    from adwatch.models import Company, CompanyEnrichment

    s = temp_db.SessionLocal()
    c = Company(name="Export Test SL", country="ES", description="Baut Wintergärten.",
                products=["Wintergarten"], founded_year=2005, enrichment_status="enriched")
    s.add(c); s.flush()
    s.add(CompanyEnrichment(company_id=c.id, status="enriched", fields={
        "assessment_de": "Wirkt wie ein spezialisierter Kleinbetrieb.",
        "mentions_solarlux": True, "competitor_brands": ["Sunflex"]}))
    s.commit(); s.close()

    wb = openpyxl.load_workbook(io.BytesIO(export_xlsx(filters={})))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    for col in ("Beschreibung (Website)", "Einschätzung (KI, unbestätigt)", "Produkte",
                "Nennt Solarlux", "Wettbewerber auf Website", "Aktive Anzeigen"):
        assert col in header, col
    row = {header[i]: v for i, v in enumerate([c.value for c in ws[2]])}
    assert row["Beschreibung (Website)"] == "Baut Wintergärten."
    assert row["Einschätzung (KI, unbestätigt)"] == "Wirkt wie ein spezialisierter Kleinbetrieb."
    assert row["Produkte"] == "Wintergarten"
    assert row["Nennt Solarlux"] == "ja"
    assert row["Wettbewerber auf Website"] == "Sunflex"


def test_extract_separates_facts_from_assessment():
    """The assessment is capped and kept as its own field; the fact fields stay
    extract-only (the prompt enforces that, the parser enforces the shape)."""
    from adwatch.enrich.extract import _clean_list, PRODUCT_VOCAB
    import adwatch.enrich.extract as ex

    raw = {
        "description_de": "Baut Fenster.", "products": ["Fenster", "Unfug"],
        "founded_year": 1990, "employee_hint": None, "legal_form": "GmbH",
        "service_area": None, "mentions_solarlux": True, "competitor_brands": ["warema"],
        "evidence": {"description_de": "Wir bauen Fenster."},
        "assessment_de": "X" * 900,
    }

    class _Blk:
        type = "text"
        text = __import__("json").dumps(raw)

    class _Msg:
        content = [_Blk()]

    class _Client:
        def __init__(self, **k): self.messages = self
        def create(self, **k): return _Msg()

    import sys, types
    mod = types.ModuleType("anthropic"); mod.Anthropic = _Client
    sys.modules["anthropic"] = mod
    ex.config.ANTHROPIC_API_KEY = "test-key"

    got = ex.extract_facts("x" * 200)
    assert got["description_de"] == "Baut Fenster."
    assert got["products"] == ["Fenster"]                 # off-vocabulary dropped
    assert got["competitor_brands"] == ["WAREMA"]         # canonicalised
    assert len(got["assessment_de"]) == 700               # capped, not unbounded
    assert "evidence" in got and got["evidence"]["description_de"]


def test_icp_diagnose_guards(temp_db):
    """The validity check behind 'an ICP for any filter — but only the ones that
    make sense': it must refuse a too-small set, refuse a set whose winners look
    exactly like the population, and detect a set that secretly mixes two
    incompatible groups (which is how the Handel/Verarbeiter split was found)."""
    from adwatch import customers
    from adwatch.insights import icp
    from adwatch.models import Company
    from sqlalchemy import select

    s = temp_db.SessionLocal()

    def mk(name, seg, plz, buys):
        s.add(Company(name=name, country="DE", segment=seg, postal_code=plz,
                      sales_channel="Fachhandelsvertrieb",
                      revenue_y0=50000 if buys else None,
                      revenue_y1=40000 if buys else None))

    # (a) tiny set -> unusable on n alone
    for i in range(8):
        mk(f"Klein {i}", "Nische", "49134", i < 4)
    s.commit()
    for c in s.scalars(select(Company)):
        c.customer_state = customers.derive_customer_state(
            c.revenue_y0, c.revenue_y1, c.revenue_y2, c.revenue_y3, c.revenue_y4)
    s.commit(); s.close()

    d = icp.diagnose({"customer_state": ["active", "new"], "segment": ["Nische"]})
    assert d["verdict"] == "unusable"
    assert any("unter 30" in r for r in d["reasons"])

    # (b) 40 buyers + 40 non-buyers that are IDENTICAL in every feature ->
    #     nothing separates them, so the profile cannot rank
    s = temp_db.SessionLocal()
    for i in range(80):
        mk(f"Gleich {i}", "Flach", "49134", i < 40)
    s.commit()
    for c in s.scalars(select(Company)):
        c.customer_state = customers.derive_customer_state(
            c.revenue_y0, c.revenue_y1, c.revenue_y2, c.revenue_y3, c.revenue_y4)
    s.commit(); s.close()

    d = icp.diagnose({"customer_state": ["active", "new"], "segment": ["Flach"]})
    assert d["winners"] == 40
    assert d["verdict"] == "unusable"
    assert any("Kein Merkmal trennt" in r for r in d["reasons"])

    # (c) a mixed set: two groups whose winners differ sharply -> split advised.
    #     Each group needs INTERNAL variety, otherwise every feature is 100%
    #     uniform inside it, gets dropped as non-discriminating, and neither
    #     sub-profile can score anything (which is what the real Handel vs
    #     Verarbeiter sets have naturally).
    s = temp_db.SessionLocal()

    def mk2(name, seg, sub, plz, buys):
        s.add(Company(name=name, country="DE", segment=seg, sub_segment=sub,
                      postal_code=plz, sales_channel="Fachhandelsvertrieb",
                      revenue_y0=50000 if buys else None,
                      revenue_y1=40000 if buys else None))

    for i in range(40):
        mk2(f"Nord {i}", "GruppeA", "Metallbau" if i < 24 else "Tischler", "20095", True)
        mk2(f"Sued {i}", "GruppeB", "Glaser" if i < 24 else "Fensterbau", "80331", True)
    for i in range(40):                                  # non-buyers on both sides
        mk2(f"NordNo {i}", "GruppeA", "Metallbau" if i < 20 else "Tischler", "49134", False)
        mk2(f"SuedNo {i}", "GruppeB", "Glaser" if i < 20 else "Fensterbau", "70173", False)
    s.commit()
    for c in s.scalars(select(Company)):
        c.customer_state = customers.derive_customer_state(
            c.revenue_y0, c.revenue_y1, c.revenue_y2, c.revenue_y3, c.revenue_y4)
    s.commit(); s.close()

    d = icp.diagnose({"customer_state": ["active", "new"], "segment": ["GruppeA", "GruppeB"]})
    seg_split = next((sp for sp in d["splits"] if sp["dimension"] == "segment"), None)
    assert seg_split is not None and seg_split["should_split"] is True
    assert any("Gemischte Grundgesamtheit" in r for r in d["reasons"])


def test_icp_scores_propensity_not_popularity(temp_db):
    """The heart, and the correction that made it work.

    Scoring used to reward the winners' most COMMON value. That ranks popularity,
    not propensity: measured live, Bauelementehandel is 36.5% of winners but
    converts at 1.03x, while Wintergartenbau is ~1% of winners and converts at
    1.62x — share ordered them backwards. fit_for now scores LIFT, so a value is
    rewarded for being over-represented among winners RELATIVE to the population.

    Fixture: Tischler is the common trade (60 of 90 companies) but converts
    poorly; Metallbau is rarer but converts well. Share-based scoring would rank
    Tischler top; lift must rank Metallbau top.
    """
    from adwatch import customers
    from adwatch.insights import icp
    from adwatch.models import Company
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    win_ids = []
    # 20 Metallbau winners out of 30 Metallbau companies  -> strongly over-represented
    for i in range(30):
        c = Company(name=f"Metall {i}", country="DE", segment="Verarbeiter",
                    sub_segment="Metallbau-Schlosser", sales_channel="Fachhandelsvertrieb",
                    postal_code="49134")
        s.add(c); s.flush()
        if i < 20:
            win_ids.append(c.id)
    # 10 Tischler winners out of 60 Tischler companies -> the COMMON trade, but
    # under-represented among winners
    for i in range(60):
        c = Company(name=f"Tisch {i}", country="DE", segment="Verarbeiter",
                    sub_segment="Tischler", sales_channel="Fachhandelsvertrieb",
                    postal_code="80331")
        s.add(c); s.flush()
        if i < 10:
            win_ids.append(c.id)
    blank = Company(name="K Leer", country="DE")
    s.add(blank); s.commit()
    bl_id = blank.id
    # a non-winner of each trade, to compare
    metall_id = s.scalars(select(Company).where(
        Company.name == "Metall 29")).one().id
    tisch_id = s.scalars(select(Company).where(
        Company.name == "Tisch 59")).one().id
    s.close()

    p = icp.build_profile({"ids": win_ids})
    assert p["winners_count"] == 30
    subs = p["features"]["sub_segment"]
    # Tischler is the LARGER share of winners' trade population but Metallbau is
    # the over-represented one — this is exactly the inversion that broke ranking
    assert dict(subs["shares"])["Metallbau-Schlosser"] == pytest.approx(20 / 30)
    assert subs["lifts"]["Metallbau-Schlosser"] > subs["lifts"]["Tischler"]
    assert subs["lifts"]["Metallbau-Schlosser"] > 1.0 > subs["lifts"]["Tischler"]

    res = icp.apply_profile({"ids": win_ids}, name="test")
    assert res["companies_scored"] >= 90

    s = temp_db.SessionLocal()
    me, ti, bl = (s.get(Company, metall_id), s.get(Company, tisch_id),
                  s.get(Company, bl_id))
    assert me.fit_score > 50 > ti.fit_score, (me.fit_score, ti.fit_score)
    assert bl.fit_score is None and bl.target_score is None   # nothing comparable -> unrated, not 0
    feats = {f["feature"] for f in me.fit_breakdown["features"]}
    assert "sales_channel" not in feats               # 100%-uniform -> excluded
    assert "sub_segment" in feats
    s.close()


def test_availability_leakage_is_detected_and_excluded(temp_db):
    """A feature known for winners far more often than for the population is
    measuring 'we already engaged this account', not fit. Live cases: products
    (13.4x), Betriebsgröße (13.7x), Firmenalter (11.9x), Anzeigen-Aktivität
    (9.6x) — all only exist for the enriched/monitored base, which WAS the old
    buyers-only export. Scoring on them yields a confident model that says
    'accounts we already sell to, buy from us'."""
    from adwatch.insights import icp
    from adwatch.models import Company
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    win_ids = []
    for i in range(40):
        # winners are enriched (products known)
        c = Company(name=f"Gewinner {i}", country="DE", segment="Handel",
                    sub_segment="Bauelementehandel", postal_code="49134",
                    products=["Fenster"])
        s.add(c); s.flush(); win_ids.append(c.id)
    for i in range(160):
        # the population is not enriched at all
        s.add(Company(name=f"Rest {i}", country="DE", segment="Handel",
                      sub_segment="Bauelementehandel", postal_code="49134"))
    s.commit(); s.close()

    p = icp.build_profile({"ids": win_ids})
    prod = p["features"]["products"]
    assert prod["coverage"] == 1.0
    assert prod["pop_coverage"] < 0.3
    assert prod["leaky"] is True, prod
    assert p["features"]["sub_segment"]["leaky"] is False

    # and the leaky feature must not contribute to any score
    fit, bd = icp.fit_for({"products": ["Fenster"], "sub_segment": "Bauelementehandel",
                           "segment": "Handel"}, p)
    assert "products" not in {b["feature"] for b in bd}


def test_downgrade_resets_collected_ads(temp_db):
    """A recheck that DOWNGRADES an auto-confirmed page (page before, none now)
    must clear that page's collected ads/metric at the point the page is dropped
    — not leave a phantom active-ad count on a now-page-less company (the
    Andreas-Schimke bug)."""
    from adwatch.identity import resolver
    from adwatch.models import Ad, CollectionRun, Company, CompanyPage, WeeklyCompanyMetric
    from sqlalchemy import select
    s = temp_db.SessionLocal()
    c = Company(name="Downgrade Co", resolution_status="confirmed", country="DE",
                page_id="777", page_name="Wrong Page")
    s.add(c); s.flush()
    s.add(CompanyPage(company_id=c.id, source="meta", page_id="777", role="main", status="auto"))
    run = CollectionRun(company_id=c.id, source="meta", week_start=dt.date(2026, 7, 6),
                        page_id="777", status="ok", ads_scraped=1)
    s.add(run); s.flush()
    s.add(Ad(run_id=run.id, source="meta", external_ad_id="a1", is_active=True))
    s.add(WeeklyCompanyMetric(company_id=c.id, source="meta", week_start=dt.date(2026, 7, 6),
                             total_active_ads=1, score=10))
    s.commit(); cid = c.id

    c = s.get(Company, cid)
    resolver._apply_identity_result(
        s, c, {"status": "no_ads_found", "page_id": None, "page_name": None,
               "page_url": None, "candidates": []}, method="serper")
    s.commit()

    assert s.scalar(select(CollectionRun).where(CollectionRun.company_id == cid)) is None
    assert s.scalar(select(WeeklyCompanyMetric).where(WeeklyCompanyMetric.company_id == cid)) is None
    assert s.scalar(select(CompanyPage).where(CompanyPage.company_id == cid, CompanyPage.role == "main")) is None
    c = s.get(Company, cid)
    assert c.page_id is None and c.resolution_status == "no_ads_found"
    s.close()


def test_unlink_resets_collected_ads(temp_db):
    """Unlinking a wrong page must clear its collected ads/score — otherwise
    the wrong page's numbers linger on the company (the Bau-DL bug)."""
    from adwatch.identity import resolver
    from adwatch.models import Company, CompanyPage, CollectionRun, Ad, WeeklyCompanyMetric
    from sqlalchemy import select
    s = temp_db.SessionLocal()
    c = Company(name="Wrong Page Co", resolution_status="confirmed", country="DE",
                page_id="999", page_name="Wrong")
    s.add(c); s.flush()
    s.add(CompanyPage(company_id=c.id, source="meta", page_id="999", role="main", status="auto"))
    run = CollectionRun(company_id=c.id, source="meta", week_start=dt.date(2026, 7, 6),
                        page_id="999", status="ok", ads_scraped=42)
    s.add(run); s.flush()
    s.add(Ad(run_id=run.id, source="meta", external_ad_id="a1", is_active=True))
    s.add(WeeklyCompanyMetric(company_id=c.id, source="meta", week_start=dt.date(2026, 7, 6),
                             total_active_ads=42, score=85))
    s.commit(); cid = c.id
    s.close()

    resolver.unlink_main(cid)

    s = temp_db.SessionLocal()
    assert s.scalar(select(WeeklyCompanyMetric).where(WeeklyCompanyMetric.company_id == cid)) is None
    assert s.scalar(select(CollectionRun).where(CollectionRun.company_id == cid)) is None
    assert s.scalar(select(Ad).where(Ad.external_ad_id == "a1")) is None
    c = s.get(Company, cid)
    assert c.page_id is None and c.resolution_status in ("ambiguous", "pending")
    s.close()


# ---------------------------------------------------------------------------
# Belege / RFM — the corrections that make churn detection meaningful
# ---------------------------------------------------------------------------

def _ev(*items):
    """[(date, amount)] from (iso, amount) pairs."""
    import datetime as _dt
    return [(_dt.date.fromisoformat(d), a) for d, a in items]


def test_cadence_is_measured_not_assumed():
    """A dealer ordering every 14 days that has been quiet 120 days is overdue;
    a Wohnungswirtschaft ordering yearly at 120 days is not. A fixed 12-month
    cutoff cannot tell these apart, which is the whole point of the module."""
    from adwatch.insights import rfm
    import datetime as dt
    today = dt.date(2026, 8, 5)

    fortnightly = _ev(("2025-06-01", 5000), ("2025-06-15", 5000),
                      ("2025-06-29", 5000), ("2026-04-01", 5000))
    r = rfm.classify(fortnightly, today)
    assert r["cadence_days"] == 14
    assert r["health"] in ("gefährdet", "verloren")
    assert r["overdue_factor"] > 3

    yearly = _ev(("2022-01-10", 90000), ("2023-01-20", 90000),
                 ("2024-02-01", 90000), ("2025-06-01", 90000))
    r2 = rfm.classify(yearly, today)
    assert r2["cadence_days"] >= 365
    assert r2["health"] == "aktiv", r2


def test_spare_parts_only_is_not_a_system_customer():
    """~25% of Belege are 0 EUR and the median is EUR 194. Without a materiality
    floor a gasket order makes a company look like a customer and poisons any ICP
    trained on 'buyers'."""
    from adwatch.insights import rfm
    import datetime as dt
    today = dt.date(2026, 8, 5)
    trivial = _ev(("2026-01-05", 0), ("2026-02-05", 120), ("2026-03-05", 80),
                  ("2026-04-05", 300))
    assert rfm.classify(trivial, today)["health"] == "einmalig"
    # the same monthly rhythm, but material and still current, is a live customer
    real = _ev(("2026-04-05", 9000), ("2026-05-05", 9000), ("2026-06-05", 9000),
               ("2026-07-05", 9000))
    assert rfm.classify(real, today)["health"] == "aktiv"


def test_no_events_is_never_not_lost():
    from adwatch.insights import rfm
    r = rfm.classify([])
    assert r["health"] == "nie" and r["value"] == 0.0
    # and it must not produce a win-back rank — there is nothing to win back
    assert rfm.winback_score(r, 0.0) == 0.0


def test_winback_ad_signal_is_a_multiplier_and_value_is_log_scaled():
    """50% of revenue sits with 66 companies, so a linear value term would make
    the list nothing but whales; and an advertising lapsed customer must outrank
    an equally-valuable silent one."""
    from adwatch.insights import rfm
    import datetime as dt
    evs = _ev(("2024-01-05", 60000), ("2024-03-05", 60000),
              ("2024-05-05", 60000), ("2024-07-05", 60000))
    cls = rfm.classify(evs, dt.date(2026, 8, 5))
    quiet = rfm.winback_score(cls, cls["value"])
    ads = rfm.winback_score(cls, cls["value"], advertising=True)
    assert ads > quiet > 0
    # log scaling: 100x the revenue must not give anything like 100x the score
    big = rfm.winback_score(cls, cls["value"] * 100)
    assert big < quiet * 2


def test_crm_import_refuses_a_truncated_download(tmp_path):
    """A partial download must not be mistaken for the full population — it would
    look like thousands of accounts had vanished."""
    import json, pytest
    from adwatch import crm_import
    p = tmp_path / "part.json"
    p.write_text(json.dumps({"cols": ["crm_id", "name"], "rows": [["a", "X"]]}),
                 encoding="utf-8")
    with pytest.raises(ValueError, match="partial"):
        crm_import.load_export(p)


def test_crm_import_never_writes_local_owned_columns():
    """Enrichment, scores and linked ad identities survive a full re-import."""
    from adwatch.crm_accounts import LOCAL_OWNED
    from adwatch.crm_import import WRITES
    assert not (WRITES & LOCAL_OWNED), sorted(WRITES & LOCAL_OWNED)


def test_bulk_imported_companies_are_not_monitored(temp_db):
    """46,000 CRM accounts must feed the ICP without flooding the ad pipeline."""
    import json
    from sqlalchemy import select
    from adwatch import crm_import
    from adwatch.models import Company
    rows = [[f"guid-{i}", f"Firma {i}", "", 101, 101000, 102690001, 102690000,
             "Deutschland", "49", "Ort", "", None, "2024-01-01",
             0, 0, "", "", 0, 0, 0, 0, None, 0, 0, 0] for i in range(600)]
    cols = ["crm_id", "name", "accountnumber", "segment", "sub_segment",
            "sales_channel", "kunde_interessent", "country", "postal_code",
            "city", "website", "employees", "created_on", "beleg_count",
            "beleg_sum", "beleg_first", "beleg_last", "rev_2023", "rev_2024",
            "rev_2025", "rev_2026", "avg_discount", "arch_projects",
            "arch_won", "arch_won_value"]
    p = temp_db.config.DATA_DIR if hasattr(temp_db, "config") else None
    import tempfile, pathlib
    f = pathlib.Path(tempfile.mkdtemp()) / "e.json"
    f.write_text(json.dumps({"cols": cols, "rows": rows}), encoding="utf-8")
    stats = crm_import.import_accounts(f)
    assert stats["inserted"] == 600
    with temp_db.SessionLocal() as s:
        assert s.scalars(select(Company).where(Company.monitored.is_(False))).all()
        c = s.scalars(select(Company).where(Company.crm_id == "guid-7")).one()
        assert c.segment == "Verarbeiter" and c.sub_segment == "Fensterbau"
        assert c.monitored is False


# ---------------------------------------------------------------------------
# Market list import (a colleague's scraped spreadsheet)
# ---------------------------------------------------------------------------

_MARKT_CSV = (
    "Name;Typ;Adresse;Lat;Lng;Website;Ansprechpartner;Notizen;Untertyp;"
    "Marken/Produkte;Einschaetzung;\n"
    # intact row
    "LUCOR Ventanas;potenzialkunde;Calle X 1, 14001 Cordoba, Spanien;37.8;-4.7;"
    "https://www.lucor.es/;;;;;;\n"
    # SHIFTED by one: the Spanish legal-form comma became a semicolon
    "CARPYVENT; S.L.;potenzialkunde;Av Y 2, 03001 Alicante, Spanien;38.3;-0.4;"
    "http://carpyvent.es;;;;;\n"
    # a real competitor location: the manufacturer's name IS the company name
    "Schueco Showroom Madrid;wettbewerber;Valdemoro, 28340 Madrid;40.1;-3.6;"
    "https://schueco.com/es;;Eigener Showroom;Showroom;;;\n"
    # installs a competitor's systems -> a PROSPECT, not a competitor
    "Premial;wettbewerber;Mijas Costa, 29650 Malaga;36.5;-4.8;;;"
    "Schueco-Premiumpartner;Produktion;Schueco Premium Partner;;\n"
    # duplicate of the row above with a different Typ
    "Premial;potenzialkunde;Mijas Costa, 29650 Malaga;36.5;-4.8;;;"
    "Schueco-Premiumpartner;Produktion;Schueco Premium Partner;;\n"
    # existing customer, joinable only by the Kd-Nr in free text
    "IBZ Cristal;bestandskunde;Ibilbidea 80, 20115 Astigarraga;43.2;-1.9;;;"
    "Kd-Nr. 5164611 | Lizenznehmer SL25;;;;\n"
)


def _write_markt(tmp_path):
    p = tmp_path / "markt.csv"
    p.write_text(_MARKT_CSV, encoding="utf-8-sig")
    return p


def test_market_list_repairs_semicolon_shifted_names(tmp_path):
    """The Spanish legal-form comma arrived as a semicolon, shifting 46 of 534 real
    rows. Unrepaired, 'S.L.' becomes the company TYPE and every later column lands
    in the wrong field."""
    from adwatch import market_list as ml
    out = ml.parse(_write_markt(tmp_path))
    names = {r["name"]: r for r in out["records"]}
    assert "CARPYVENT, S.L." in names, sorted(names)
    assert names["CARPYVENT, S.L."]["import_type"] == "potenzialkunde"
    assert names["CARPYVENT, S.L."]["city"] == "Alicante"
    assert names["CARPYVENT, S.L."]["postal_code"] == "03001"
    assert out["stats"]["unparsable"] == 0


def test_market_list_separates_competitors_from_conquest_targets(tmp_path):
    """'wettbewerber' means two opposite things. A manufacturer's OWN location is
    never a target; a firm that merely INSTALLS a rival's systems is the best
    target in the file — and must stay recognisable as having arrived tagged
    'wettbewerber'."""
    from adwatch import market_list as ml
    recs = {r["name"]: r for r in ml.parse(_write_markt(tmp_path))["records"]}

    schueco = recs["Schueco Showroom Madrid"]
    assert schueco["is_competitor"] is True
    assert schueco["carries_competitor_brand"] is False

    premial = recs["Premial"]
    assert premial["is_competitor"] is False, "installs Schueco != is Schueco"
    assert premial["carries_competitor_brand"] is True
    assert premial["import_type"] == "wettbewerber", "origin must stay auditable"
    assert premial["segment"] == "Verarbeiter"


def test_market_list_dedupes_and_keeps_the_discarded_type(tmp_path):
    from adwatch import market_list as ml
    out = ml.parse(_write_markt(tmp_path))
    assert out["stats"]["duplicates_removed"] == 1
    premial = next(r for r in out["records"] if r["name"] == "Premial")
    # the same firm was entered twice under different Typ values — the one we
    # dropped is remembered rather than silently lost
    assert premial["also_imported_as"] == ["potenzialkunde"]


def test_market_list_matches_existing_customer_by_kdnr_not_name(temp_db, tmp_path):
    """Names do not join: 'IBZ Cristal' is 'IBZ Cortinas De Cristal SL'. Only 7 of
    534 rows matched by name, 30 matched on the Kd-Nr buried in free text."""
    from sqlalchemy import select
    from adwatch import market_list as ml
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add(Company(name="IBZ Cortinas De Cristal SL", country="ES",
                  sap_number="0005164611", segment="Verarbeiter"))
    s.commit(); s.close()

    stats = ml.import_list(_write_markt(tmp_path), lead_source="test_es")
    assert stats["matched_by"]["customer_number"] == 1
    with temp_db.SessionLocal() as s:
        c = s.scalars(select(Company).where(
            Company.name == "IBZ Cortinas De Cristal SL")).one()
        # CRM master data untouched; the research is appended to notes
        assert c.segment == "Verarbeiter"
        assert "Kd-Nr. 5164611" in (c.notes or "")
        assert "test_es" in (c.notes or "")
        # and it must NOT have been inserted a second time
        assert not s.scalars(select(Company).where(
            Company.name == "IBZ Cristal")).all()


def test_market_list_rows_are_distinguishable_and_not_monitored(temp_db, tmp_path):
    """A scraped list must never be mistakable for CRM master data, and must not
    silently enter the paid ad-fetch queue."""
    from sqlalchemy import select
    from adwatch import market_list as ml
    from adwatch.models import Company
    from adwatch import scope

    ml.import_list(_write_markt(tmp_path), lead_source="test_es")
    with temp_db.SessionLocal() as s:
        rows = s.scalars(select(Company).where(Company.lead_source == "test_es")).all()
        assert rows
        assert all(c.crm_id is None for c in rows)
        assert all(c.source == "marktanalyse" for c in rows)
        assert all(c.monitored is False for c in rows)
        # the competitor is present but excluded from every count
        in_scope = s.scalars(scope.apply(select(Company)).where(
            Company.lead_source == "test_es")).all()
        assert "Schueco Showroom Madrid" not in {c.name for c in in_scope}
        assert "Premial" in {c.name for c in in_scope}


# ---------------------------------------------------------------------------
# Website discovery — "do not add it unless you are very sure"
# ---------------------------------------------------------------------------

def test_domain_in_name_is_extracted_but_socials_are_not():
    """'CBF (calviabalear.com)' states its own domain — that is the researcher
    telling us, not a guess. A LinkedIn URL in the notes is NOT a company site."""
    from adwatch.identity.find_website import domain_from_name
    assert domain_from_name("CBF (calviabalear.com)") == "calviabalear.com"
    assert domain_from_name("Aluminios Lago, S.L.") is None
    assert domain_from_name("Óscar RV Arquitecto linkedin.com/in/oscar") is None
    assert domain_from_name("Studio facebook.com/studio") is None


def test_only_locality_backed_matches_are_auto_accepted():
    """The gate here is deliberately STRICTER than enrichment's. domain_plus_name
    proves a name coincidence, not that this is the right company — 'Premial' or
    'Al-Andalus' would match a namesake in another province, and a wrong website
    silently produces a description and an ad history for the wrong firm."""
    from adwatch.identity.find_website import PROVEN
    assert "domain_plus_name" not in PROVEN
    for strong in ("phone", "plz_street", "plz_name", "domain_in_name"):
        assert strong in PROVEN


def test_unproven_candidate_is_queued_not_written(temp_db, monkeypatch):
    """A plausible-but-unproven candidate must leave website_domain EMPTY."""
    from sqlalchemy import select
    from adwatch.identity import find_website as fw
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    c = Company(name="Aluminios Ejemplo", country="ES", city="Valencia",
                postal_code="46020", street="Av. Catalunya 13",
                lead_source="t", segment="Verarbeiter")
    s.add(c); s.commit(); cid = c.id; s.close()

    monkeypatch.setattr(fw, "search_candidates",
                        lambda *a, **k: [{"domain": "aluminios-ejemplo.com",
                                          "title": "Aluminios Ejemplo"}])
    # a page that confirms the NAME but carries neither the postcode nor the street
    monkeypatch.setattr(fw, "page_bundle",
                        lambda d, **k: {"text": "Aluminios Ejemplo — ventanas",
                                        "pages": [f"https://{d}"]})
    r = fw.find_for(cid)
    assert r["status"] == fw.NEEDS_REVIEW
    with temp_db.SessionLocal() as s:
        got = s.get(Company, cid)
        assert got.website_domain is None, "an unproven domain must not be stored"
        assert got.identity_status == fw.NEEDS_REVIEW
        assert got.identity_evidence["review_candidate"] == "aluminios-ejemplo.com"
    # ...and it is surfaced for a human instead of being dropped
    assert any(q["company_id"] == cid for q in fw.review_queue(lead_source="t"))


def test_postcode_backed_match_is_accepted(temp_db, monkeypatch):
    from adwatch.identity import find_website as fw
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    c = Company(name="Aluminios Ejemplo", country="ES", city="Valencia",
                postal_code="46020", street="Av. Catalunya 13",
                lead_source="t", segment="Verarbeiter")
    s.add(c); s.commit(); cid = c.id; s.close()

    monkeypatch.setattr(fw, "search_candidates",
                        lambda *a, **k: [{"domain": "aluminios-ejemplo.com"}])
    monkeypatch.setattr(fw, "page_bundle", lambda d, **k: {
        "text": "Aluminios Ejemplo, Av. Catalunya 13, 46020 Valencia",
        "pages": [f"https://{d}"]})
    r = fw.find_for(cid)
    assert r["status"] == fw.VERIFIED
    assert r["matched_by"] in ("plz_street", "plz_name")
    with temp_db.SessionLocal() as s:
        got = s.get(Company, cid)
        assert got.website_domain == "aluminios-ejemplo.com"
        assert got.website_source == "serper"


def test_searched_companies_are_not_paid_for_twice(temp_db, monkeypatch):
    """not_found and needs_review both count as done, or a re-run bills Serper
    again for the same company."""
    from adwatch.identity import find_website as fw
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    for i, st in enumerate([fw.NOT_FOUND, fw.NEEDS_REVIEW, None]):
        s.add(Company(name=f"Firma {i}", country="ES", lead_source="t",
                      segment="Verarbeiter", identity_status=st))
    s.commit(); s.close()
    pend = fw.pending_ids("t")
    assert len(pend) == 1, "only the never-searched company may be queued"


def test_spanish_trade_words_are_not_identifying():
    """The AURIA incident: the generic-word list was German-only, so 'estudio' and
    'arquitectura' counted as identifying. A DIFFERENT architecture studio in the
    same town (same postcode, 'estudio de arquitectura' on its homepage) then
    passed the plz_name gate and was stored as AURIA's website — the exact
    wrong-website failure the identity gate exists to prevent."""
    from adwatch.enrich.validate import distinctive_tokens, validate_site

    assert distinctive_tokens("Estudio de Arquitectura AURIA") == {"auria"}
    assert distinctive_tokens("Protec Ventanas") == {"protec"}
    assert distinctive_tokens("Aluminios Baraza") == {"baraza"}
    assert distinctive_tokens("Carpinteria Metalica FEVEGAR") == {"fevegar"}

    # the live failure, replayed: another studio's page in the same town
    company = {"name": "Estudio de Arquitectura AURIA", "phone": None,
               "postal_code": "06220", "street": "Calle Cisneros 12"}
    other_studio = "Estudio de arquitectura en Villafranca de los Barros, 06220"
    res = validate_site(company, "thau.es", other_studio)
    assert res["ok"] is False, "generic trade words must not prove identity"
    # ...while the real match (name token present) still works
    own = "AURIA estudio, Calle Cisneros 12, 06220 Villafranca"
    assert validate_site(company, "auria.es", own)["ok"] is True


def test_conflict_domains_are_never_google_fetched(temp_db):
    """A domain that FAILED identity verification must not be used to attribute a
    Google ad history — 188 of the 430 Spanish market-list sites came back
    'conflict', and fetching through them files someone else's ads under the
    company. Unverified (never checked) stays fetchable; disproven does not."""
    from adwatch.jobs import _google_fetchable_ids
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    ok = Company(name="Ok Co", country="ES", website_domain="ok.es",
                 identity_status="verified")
    unknown = Company(name="Unknown Co", country="ES", website_domain="unknown.es")
    bad = Company(name="Bad Co", country="ES", website_domain="portal.es",
                  identity_status="conflict")
    s.add_all([ok, unknown, bad]); s.commit()
    ids = [ok.id, unknown.id, bad.id]
    fetchable = _google_fetchable_ids(s, ids)
    assert ok.id in fetchable and unknown.id in fetchable
    assert bad.id not in fetchable
    s.close()


# ---------------------------------------------------------------------------
# Subpage selection — which pages the enrichment crawler actually reads
# ---------------------------------------------------------------------------

_ES_HTML = """
<html><body>
  <nav>
    <a href="/">Inicio</a>
    <a href="/productos">Productos</a>
    <a href="/servicios/ventanas-pvc/serie-70">Serie 70</a>
    <a href="/quienes-somos">Quiénes somos</a>
    <a href="/contacto">Contacto</a>
    <a href="/aviso-legal">Aviso legal</a>
    <a href="https://facebook.com/firma">Facebook</a>
  </nav>
  <div class="mobile-menu">
    <a href="/contacto">Contacto</a>
    <a href="/productos/">Productos</a>
  </div>
</body></html>
"""


def test_spanish_product_pages_are_selected_not_just_contacto():
    """The old rule matched only impressum|kontakt|contact and took the first two
    in document order, so a Spanish site yielded homepage + 'contacto' and the
    product pages were NEVER read — the products list then came from whatever the
    homepage happened to mention."""
    from adwatch.identity import website_source as ws
    picked = ws._subpage_urls("https://ejemplo.es/", _ES_HTML, max_pages=3)
    cats = [ws._classify_link(u) for u in picked]
    assert "products" in cats, picked
    assert "legal" in cats, picked
    assert "https://ejemplo.es/productos" in picked
    # the section index is read BEFORE a deep single-item page: '/productos' is
    # the whole range, '/servicios/ventanas-pvc/serie-70' is one article
    prods = [u for u in picked if ws._classify_link(u) == "products"]
    assert prods[0] == "https://ejemplo.es/productos", prods
    # and with only two slots the deep page must never displace the index
    picked2 = ws._subpage_urls("https://ejemplo.es/", _ES_HTML, max_pages=2)
    assert not any("serie-70" in u for u in picked2), picked2


def test_duplicate_nav_targets_do_not_consume_slots():
    """A nav bar repeats the same href in the desktop and mobile menus; each
    duplicate used to eat one of only two available slots."""
    from adwatch.identity import website_source as ws
    picked = ws._subpage_urls("https://ejemplo.es/", _ES_HTML, max_pages=4)
    paths = [u.rstrip("/").rsplit("ejemplo.es", 1)[-1] for u in picked]
    assert len(paths) == len(set(paths)), picked


def test_offsite_and_non_http_links_are_never_fetched():
    from adwatch.identity import website_source as ws
    picked = ws._subpage_urls("https://ejemplo.es/", _ES_HTML, max_pages=6)
    assert all("ejemplo.es" in u for u in picked), picked


def test_link_categories_cover_the_app_markets():
    """config/markets.yaml already knew ES='aviso legal', FR='mentions légales',
    IT='contatti' while the crawler only looked for German terms."""
    from adwatch.identity.website_source import _classify_link
    assert _classify_link("https://x.es/aviso-legal") == "legal"
    assert _classify_link("https://x.fr/mentions-legales") == "legal"
    assert _classify_link("https://x.it/contatti") == "legal"
    assert _classify_link("https://x.es/productos") == "products"
    assert _classify_link("https://x.pt/produtos") == "products"
    assert _classify_link("https://x.es/quienes-somos") == "about"
    assert _classify_link("https://x.de/referenzen") == "references"
    assert _classify_link("https://x.es/") is None


# ---------------------------------------------------------------------------
# site_facts — machine-readable facts, no LLM
# ---------------------------------------------------------------------------

_FACTS_HTML = """
<html lang="es-ES"><head>
  <meta property="og:description" content="Carpintería de aluminio en Málaga">
  <script type="application/ld+json">
  {"@context":"https://schema.org","@type":"LocalBusiness","name":"Protec Ventanas",
   "telephone":"+34 952 58 75 73","email":"info@protec.es",
   "foundingDate":"1998-04-01",
   "address":{"@type":"PostalAddress","streetAddress":"Calle Sol 4",
              "postalCode":"29620","addressLocality":"Torremolinos"},
   "sameAs":["https://www.facebook.com/protecventanas",
             "https://www.instagram.com/protec_ventanas"]}
  </script>
</head><body>
  <a href="tel:+34952587573">Llamar</a>
  <a href="mailto:info@protec.es">Escribir</a>
  <a href="https://www.facebook.com/sharer/sharer.php?u=x">Compartir</a>
  <a href="https://www.linkedin.com/company/protec-ventanas">LinkedIn</a>
</body></html>
"""


def test_site_facts_reads_jsonld_contact_and_socials():
    """Free, deterministic, and it unlocks the two strongest identity signals:
    phone (ranked first by validate_site) and the company's OWN Facebook page.
    Not one of the 39 Spain rows without a website had a phone number."""
    from adwatch.enrich import site_facts
    f = site_facts.extract(_FACTS_HTML, base_url="https://protec.es/")
    assert f["phone"] == "+34 952 58 75 73"
    assert f["email"] == "info@protec.es"
    assert f["postal_code"] == "29620" and f["city"] == "Torremolinos"
    assert f["street"] == "Calle Sol 4"
    assert f["founded_year"] == 1998
    assert f["language"] == "es-es"
    assert f["social"]["facebook"] == "https://www.facebook.com/protecventanas"
    assert f["social"]["instagram"] == "https://www.instagram.com/protec_ventanas"
    assert f["social"]["linkedin"] == "https://www.linkedin.com/company/protec-ventanas"
    assert f["sources"]["phone"] == "json-ld"


def test_share_widgets_are_not_mistaken_for_the_company_profile():
    """A sharer link points at OUR page on Facebook, not the company's — treating
    it as the company's profile would attribute someone else's ads."""
    from adwatch.enrich import site_facts
    html = ('<a href="https://www.facebook.com/sharer/sharer.php?u=x">s</a>'
            '<a href="https://facebook.com/plugins/like.php">l</a>')
    assert "social" not in site_facts.extract(html)


def test_personal_mailboxes_are_not_harvested():
    """A role inbox is a company address; a named person's is personal data the
    app has no reason to store (same rule that excludes Geschäftsführer names)."""
    from adwatch.enrich import site_facts
    f = site_facts.extract('<a href="mailto:maria.gomez@firma.es">Maria</a>')
    assert "email" not in f
    f2 = site_facts.extract('<a href="mailto:info@firma.es">Info</a>')
    assert f2["email"] == "info@firma.es"


def test_malformed_jsonld_never_breaks_extraction():
    from adwatch.enrich import site_facts
    html = ('<script type="application/ld+json">{"@type":"Organization",}</script>'
            '<a href="tel:+34911223344">x</a>')
    f = site_facts.extract(html)
    assert f["phone"] == "+34911223344"   # salvaged the trailing comma, or fell back


def test_tri_state_booleans_keep_not_stated_distinct_from_no():
    """A site that never mentions its workshop must not be recorded as a
    confirmed pure trader."""
    from adwatch.enrich.extract import _tri_state
    assert _tri_state(True) is True
    assert _tri_state(False) is False
    assert _tri_state(None) is None
    assert _tri_state("ja") is True
    assert _tri_state("unklar") is None


# ---------------------------------------------------------------------------
# Haiku conflict triage — diagnoses, never verifies
# ---------------------------------------------------------------------------

def test_triage_routes_but_never_writes_verified(temp_db, monkeypatch):
    """The design rule from the migration incident: only the deterministic gate
    may write 'verified'. Triage clears a diagnosed wrong_site domain (kept in
    evidence), queues likely_right for review with the clue, leaves too_thin as
    conflict — and no path produces 'verified'."""
    from adwatch.identity import triage
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    wrong = Company(name="Dealer A", country="ES", website_domain="technal.com",
                    identity_status="conflict", lead_source="t", segment="Handel")
    right = Company(name="Dealer B", country="ES", website_domain="dealerb.es",
                    identity_status="conflict", lead_source="t", segment="Handel")
    thin = Company(name="Dealer C", country="ES", website_domain="thin.es",
                   identity_status="conflict", lead_source="t", segment="Handel")
    s.add_all([wrong, right, thin]); s.commit()
    ids = {wrong.id: "wrong_site", right.id: "likely_right", thin.id: "too_thin"}
    s.close()

    monkeypatch.setattr(triage, "_evidence_for",
                        lambda c: {"reachable": True, "excerpt": "x" * 100})
    # confident on purpose: this test covers ROUTING, not the confidence gate
    monkeypatch.setattr(triage, "_judge_batch", lambda blocks: {
        cid: {"verdict": v, "confidence": 0.9, "what": "w", "clue": "c"}
        for cid, v in ids.items()})
    monkeypatch.setattr(triage.config, "ANTHROPIC_API_KEY", "test", raising=False)

    r = triage.run(lead_source="t")
    assert r["wrong_site"] == 1 and r["likely_right"] == 1 and r["too_thin"] == 1

    with temp_db.SessionLocal() as s:
        w, ri, th = (s.get(Company, cid) for cid in ids)
        assert w.website_domain is None                    # cleared for the finder
        assert w.identity_evidence["triage"]["domain_at_triage"] == "technal.com"
        assert ri.identity_status == "needs_review"
        assert ri.website_domain == "dealerb.es"           # kept, human decides
        assert th.identity_status == "conflict"
        for c in (w, ri, th):
            assert c.identity_status != "verified"


def test_backup_verify_catches_the_snapshots_that_bit_us(temp_db, tmp_path, monkeypatch):
    """13 of 14 retained snapshots were once 4 KB empty files written by the test
    suite, with the only good copy one rotation from deletion. verify_latest()
    exists so a useless snapshot is FOUND rather than trusted."""
    import sqlite3
    from adwatch import backup, config as cfg

    bdir = tmp_path / "b"; bdir.mkdir()
    monkeypatch.setattr(cfg, "BACKUP_DIR", bdir, raising=False)
    monkeypatch.setattr(backup.config, "BACKUP_DIR", bdir, raising=False)

    # no backup at all
    assert backup.verify_latest()["ok"] is False

    # a tiny/empty snapshot must be rejected, not reported as fine
    tiny = bdir / "adwatch_20260101_000000_x.db"
    con = sqlite3.connect(tiny); con.execute("CREATE TABLE companies (id INTEGER)")
    con.commit(); con.close()
    r = backup.verify_latest()
    assert r["ok"] is False and "small" in (r.get("error") or "")


def test_postcode_check_is_country_aware():
    """Requiring exactly 5 digits was German thinking. It silently never matched
    for ~8,200 companies: AT/DK/NO/BE (4 digits), NL/GB (alphanumeric) — so one
    of only three hard identity proofs was dead in six countries, with no error
    anywhere."""
    from adwatch.enrich.validate import plz_matches

    # DE: unchanged 5-digit behaviour
    assert plz_matches("49134", "Wallenhorst, 49134 Deutschland", country="DE")
    assert not plz_matches("49134", "nothing here", country="DE")

    # AT 4-digit: needs the town too, so a year cannot pass as a postcode
    at_page = "Musterweg 3, 4020 Linz, Österreich"
    assert plz_matches("4020", at_page, country="AT", city="Linz")
    assert not plz_matches("4020", "gegründet 4020 Stück verkauft", country="AT",
                           city="Linz"), "bare 4-digit match must not count"
    assert not plz_matches("1998", "Firma seit 1998", country="AT", city="Linz")

    # NL alphanumeric, spacing-insensitive
    assert plz_matches("1234 AB", "Straat 5, 1234AB Amsterdam", country="NL")
    assert plz_matches("1234AB", "Straat 5, 1234 AB Amsterdam", country="NL")

    # GB outcode+incode
    assert plz_matches("SW1A 1AA", "London SW1A 1AA", country="GB")

    # unknown country falls back to the safe 5-digit rule
    assert plz_matches("28001", "Madrid 28001") is True
    assert plz_matches("4020", "Linz 4020") is False


def test_triage_brand_overlap_is_deterministic_evidence():
    """If the researcher wrote 'Schüco + Drutex' and the site says Schüco, that is
    hard evidence independent of the LLM — a namesake in another province does not
    happen to carry the same profile systems."""
    from adwatch.identity.triage import _brand_overlap
    notes = "Schueco + Drutex, eigener Ausstellungsraum"
    # the researcher wrote 'Schueco', the Spanish site writes 'Schüco' — a naive
    # substring match found NO overlap and silently threw the evidence away
    assert _brand_overlap(notes, "Somos distribuidor Schüco oficial") == ["Schüco"]
    assert _brand_overlap(notes, "Schuco y Drutex") == ["Drutex", "Schüco"]
    assert _brand_overlap(notes, "ventanas de PVC baratas") == []
    assert _brand_overlap(None, "Schüco") == []


def test_triage_downgrades_unconfident_guesses(temp_db, monkeypatch):
    """A four-way label choice always returns something. Without a quotable clue
    or a brand match, low confidence must not reach the human queue."""
    from adwatch.identity import triage
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    c = Company(name="Vago SL", country="ES", website_domain="vago.es",
                identity_status="conflict", lead_source="t", segment="Handel")
    s.add(c); s.commit(); cid = c.id; s.close()

    monkeypatch.setattr(triage.config, "ANTHROPIC_API_KEY", "test", raising=False)
    monkeypatch.setattr(triage, "_evidence_for",
                        lambda x: {"reachable": True, "excerpt": "y" * 80,
                                   "brand_overlap": []})
    monkeypatch.setattr(triage, "_judge_batch", lambda blocks: {
        cid: {"verdict": "likely_right", "confidence": 0.3, "what": "", "clue": ""}})
    r = triage.run(lead_source="t")
    assert r["too_thin"] == 1 and r["likely_right"] == 0
    with temp_db.SessionLocal() as s:
        assert s.get(Company, cid).identity_status == "conflict"


def test_dossier_separates_roles_and_synthesises_profile(temp_db):
    """The FBI file: every VC in every ROLE the company plays, never mixed —
    an architect's 'lost' VC is not a lost sale — plus a Kurzprofil whose every
    clause traces to a column (deterministic, no LLM per view)."""
    import datetime as _dt
    from adwatch import dossier
    from adwatch.models import Company, CrmOpportunity, CrmOrderEvent

    s = temp_db.SessionLocal()
    c = Company(name="Muster Bau GmbH", country="DE", city="Osnabrück",
                crm_id="GUID-1", segment="Handel", sub_segment="Bauelementehandel",
                positioning="premium", own_fabrication=True,
                quote_sum=100000, conversion_rate=0.25)
    s.add(c); s.flush()
    s.add(CrmOrderEvent(company_id=c.id, order_date=_dt.date(2024, 3, 1),
                        amount=50000, beleg_count=2))
    # as buyer: one won VC with invoiced value and SAP trace
    s.add(CrmOpportunity(crm_id="1", number="1", parent_account_crm_id="guid-1",
                         state="gewonnen", order_value=40000, invoiced_value=38000,
                         sap_order_numbers=["4711"], type_of_use="Wohnen",
                         origin="vom Händler", project_id="p1",
                         opportunity_guid="v1",
                         created_on=_dt.datetime(2024, 1, 1)))
    # as architect: a lost VC — must land in the ARCHITECT block, not the buyer's
    s.add(CrmOpportunity(crm_id="2", number="2", architect_crm_id="guid-1",
                         state="verloren", lost_reason="Zu teuer",
                         project_id="p2", opportunity_guid="v2",
                         created_on=_dt.datetime(2024, 2, 1)))
    s.commit(); cid = c.id; s.close()

    d = dossier.build(cid)
    assert d["rollen"]["kaeufer"]["won"] == 1
    assert d["rollen"]["kaeufer"]["invoiced_value"] == 38000
    assert d["rollen"]["kaeufer"]["recent"][0]["sap_orders"] == ["4711"]
    assert d["rollen"]["architekt"]["lost"] == 1
    assert "kaeufer" in d["rollen"] and d["rollen"]["architekt"]["vcs"] == 1
    assert len(d["projekte"]) == 2
    kp = d["kurzprofil"]
    assert "Bauelementehandel" in kp and "Osnabrück" in kp
    assert "eigene Fertigung" in kp and "Konversion 25%" in kp


def test_dossier_project_outcome_uses_the_one_win_rule(temp_db):
    """A project with one win and one sibling loss is a WON project in the
    dossier's Objekte list, per the Objektvertrieb rule."""
    import datetime as _dt
    from adwatch import dossier
    from adwatch.models import Company, CrmOpportunity

    s = temp_db.SessionLocal()
    c = Company(name="GU Beispiel", country="DE", crm_id="GUID-9",
                segment="Baudienstleister")
    s.add(c); s.flush()
    s.add(CrmOpportunity(crm_id="10", number="10", parent_account_crm_id="guid-9",
                         state="verloren", lost_reason="Zugehörige VC gewonnen",
                         project_id="prj", opportunity_guid="prj",
                         project_name="Objekt Musterstraße",
                         created_on=_dt.datetime(2024, 5, 1)))
    s.add(CrmOpportunity(crm_id="11", number="11",
                         parent_account_crm_id="someone-else",
                         state="gewonnen", order_value=90000,
                         project_id="prj", opportunity_guid="v11",
                         created_on=_dt.datetime(2024, 5, 2)))
    s.commit(); cid = c.id; s.close()

    d = dossier.build(cid)
    prj = next(p for p in d["projekte"] if p["project_id"] == "prj")
    assert prj["status"] == "gewonnen", "one win makes the project won"
    assert prj["members"] == 2


# ---------------------------------------------------------------------------
# Architekten need their own extraction profile
# ---------------------------------------------------------------------------

def test_architect_profile_is_selected_from_segment():
    from adwatch.enrich.extract import (profile_for, PROFILE_ARCHITEKT,
                                        PROFILE_BETRIEB)
    assert profile_for("Architekten") == PROFILE_ARCHITEKT
    # planners filed under another segment still behave like architects
    assert profile_for("Baudienstleister", "Generalplaner") == PROFILE_ARCHITEKT
    assert profile_for("Handel", "Bauelementehandel") == PROFILE_BETRIEB
    assert profile_for(None) == PROFILE_BETRIEB


def test_architect_prompt_never_asks_a_planner_to_sell():
    """The dealer prompt opens with 'Bauelemente-/Handwerksbetrieb' and asks for
    own_fabrication and has_showroom — wrong in kind for a planning office, which
    SPECIFIES systems. The architect prompt must ask the architect questions."""
    from adwatch.enrich.extract import _prompt, PROFILE_ARCHITEKT, PROFILE_BETRIEB
    arch = _prompt(PROFILE_ARCHITEKT)
    deal = _prompt(PROFILE_BETRIEB)
    assert "ARCHITEKTUR" in arch and "VERKAUFT keine Bauelemente" in arch
    assert "own_fabrication" not in arch and "has_showroom" not in arch
    assert "solarlux_relevance" in arch and "decision_role" in arch
    # and the dealer prompt is untouched
    assert "own_fabrication" in deal and "Bauelemente-/Handwerksbetrieb" in deal


def test_dashboard_only_loads_companies_with_an_ad_footprint(temp_db):
    """After the CRM import put 46k accounts in the database, /api/state built a
    metric row for every one of them — two queries each — and shipped 33 MB of
    JSON on every page load, 18 seconds of it inside latest_metrics(). Only 731
    companies had any ad data, and the dashboard filters on has_data everywhere
    anyway, so the rest were pure payload."""
    import datetime as dt
    from adwatch import services
    from adwatch.models import Company, CompanyPage, WeeklyCompanyMetric

    s = temp_db.SessionLocal()
    tracked = Company(name="Hat Metrik", segment="Handel")
    paged = Company(name="Hat Seite", segment="Handel")
    quiet = Company(name="Nie geholt", segment="Handel")
    s.add_all([tracked, paged, quiet]); s.commit()
    s.add(WeeklyCompanyMetric(company_id=tracked.id, week_start=dt.date(2026, 8, 3),
                              source="meta", total_active_ads=3))
    s.add(CompanyPage(company_id=paged.id, page_id="p1", active=True))
    # an INACTIVE page is not a footprint — it is a page we stopped following
    s.add(CompanyPage(company_id=quiet.id, page_id="p2", active=False))
    s.commit(); s.close()

    ids = services.tracked_company_ids()
    assert set(ids) == {tracked.id, paged.id}
    assert quiet.id not in ids


def test_review_queue_filters_by_any_market(temp_db, monkeypatch):
    """Two bugs in one screen. The route used SessionLocal without importing it,
    so BOTH the queue and the reject button raised NameError — the Prüfen tab
    was dead for every user, and no test touched the endpoint. And the only
    filter was a hard-coded "Nur Spanien-Marktanalyse" checkbox, which makes
    every other market unreachable the moment one exists."""
    from fastapi.testclient import TestClient
    from adwatch import web
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add_all([
        Company(name="ES Eins", country="ES", segment="Verarbeiter",
                lead_source="marktanalyse_es_2026_08", identity_status="needs_review"),
        Company(name="DE Eins", country="DE", segment="Handel",
                lead_source=None, identity_status="needs_review"),
        Company(name="FR Eins", country="FR", segment="Verarbeiter",
                lead_source="marktanalyse_fr_2026", identity_status="needs_review"),
        Company(name="Entschieden", country="ES", segment="Verarbeiter",
                identity_status="verified"),
    ])
    s.commit(); s.close()
    monkeypatch.setattr(web, "SessionLocal", temp_db.SessionLocal)

    c = TestClient(web.app)
    r = c.get("/api/identity/review")
    assert r.status_code == 200, r.text          # the NameError regression
    body = r.json()
    assert {x["name"] for x in body["rows"]} == {"ES Eins", "DE Eins", "FR Eins"}
    # facets describe what is really in the queue, so the UI needs no hard-coded market
    assert body["facets"]["country"] == ["DE", "ES", "FR"]
    assert "marktanalyse_fr_2026" in body["facets"]["lead_source"]

    # any market is selectable, not just Spain
    assert {x["name"] for x in c.get("/api/identity/review?country=FR").json()["rows"]} == {"FR Eins"}
    assert {x["name"] for x in c.get("/api/identity/review?country=DE&country=ES").json()["rows"]} \
        == {"ES Eins", "DE Eins"}
    assert {x["name"] for x in c.get("/api/identity/review?segment=Handel").json()["rows"]} == {"DE Eins"}
    # narrowing the list must not shrink the choices still on offer
    assert c.get("/api/identity/review?country=FR").json()["facets"]["country"] == ["DE", "ES", "FR"]


def test_spa_shell_is_rendered_and_relinked(monkeypatch, temp_db):
    """A single-page app answers 200 with an empty shell, so the fetch "succeeds"
    and yields nothing — the company enriches to nothing and drops out of every
    list. Rendering also has to restore LINK DISCOVERY: a SPA builds its nav in
    JavaScript, so the raw HTML has no <a href> and the subpages vanish too."""
    from adwatch.enrich import fetchpage as fp

    shell = '<html><body><div id="root"></div><script src="/app.js"></script></body></html>'
    painted = ('<html><body><main>Carpintería de aluminio en Marbella. '
               'Fabricamos ventanales y cerramientos.</main>'
               '<a href="https://spa.example/productos">Productos</a>'
               '<footer>Distribuidor oficial de Sunflex</footer></body></html>')
    sub = '<html><body><main>Ventanales correderas y cerramientos de terraza.</main></body></html>'

    monkeypatch.setattr(fp, "_host_is_public", lambda h: True)
    monkeypatch.setattr(fp, "_robots_allows", lambda d: True)
    monkeypatch.setattr(fp, "_fetch", lambda url, wall_clock=None: (
        (shell, "https://spa.example") if url.rstrip("/").endswith("spa.example") else (sub, url)))
    monkeypatch.setattr(fp.render, "available", lambda: True)
    monkeypatch.setattr(fp.render, "render_html", lambda url, timeout_ms=None: painted)

    b = fp.page_bundle("spa.example")
    assert b["rendered"] is True
    assert "Carpintería de aluminio" in b["text"]
    # the brand in the rendered footer is reachable now
    assert "Sunflex" in (b["brands"] or [])
    # and the link that only exists after rendering was followed
    assert any("productos" in p for p in b["pages"])


def test_without_playwright_the_pipeline_is_unchanged(monkeypatch):
    """Nobody has to install a browser to run AdWatch. With no renderer the
    starved page stays starved — same result as before this existed — and
    nothing raises."""
    from adwatch.enrich import fetchpage as fp

    shell = '<html><body><div id="root"></div></body></html>'
    monkeypatch.setattr(fp, "_host_is_public", lambda h: True)
    monkeypatch.setattr(fp, "_robots_allows", lambda d: True)
    monkeypatch.setattr(fp, "_fetch", lambda url, wall_clock=None: (shell, "https://spa.example"))
    monkeypatch.setattr(fp.render, "available", lambda: False)
    monkeypatch.setattr(fp.render, "render_html",
                        lambda url, timeout_ms=None: pytest.fail("must not launch a browser"))

    b = fp.page_bundle("spa.example")
    assert b["rendered"] is False and b["brands"] == []


def test_render_never_raises_into_the_pipeline(monkeypatch):
    """A browser crash must degrade to the plain-fetch result, not take down the
    enrichment of a company whose site merely happens to be slow."""
    from adwatch.enrich import render as rd
    monkeypatch.setattr(rd, "available", lambda: True)

    def boom(*a, **k):
        raise RuntimeError("browser exploded")
    monkeypatch.setattr(rd, "_ua", boom)
    assert rd.render_html("https://example.com") is None


def test_objekt_detail_assembles_the_whole_project(temp_db, monkeypatch):
    """An Objekt has no record of its own — it is a GROUP of Verkaufschancen
    sharing sl_primary_opportunityid, so the drawer has to assemble it. Two
    things this must get right, both of which it got wrong first:

    * the group key falls back to the opportunity guid when a VC has no project
      id, so matching only project_id 404s on every single-VC project;
    * a firm that is Käufer, Architekt AND Endkunde on one deal is on ONE deal.
      Counting role occurrences reported 9 VCs on a 4-VC project.
    """
    from adwatch.insights import projekte
    from adwatch.models import Company, CrmOpportunity
    import datetime as _dt

    s = temp_db.SessionLocal()
    buyer = Company(name="Metallbau A", segment="Verarbeiter", crm_id="aaa", city="Wien")
    allrole = Company(name="Generalunternehmer B", segment="Baudienstleister", crm_id="bbb")
    s.add_all([buyer, allrole]); s.commit()
    s.add_all([
        CrmOpportunity(crm_id="v1", opportunity_guid="g1", project_id="P1",
                       project_name="Muthgasse 109", state="verloren",
                       lost_reason="Zugehörige VC gewonnen",
                       parent_account_crm_id="aaa", city="Wien", postal_code="1190",
                       street="Muthgasse 109", created_on=_dt.datetime(2024, 7, 15)),
        # one firm in all three roles on a single deal
        CrmOpportunity(crm_id="v2", opportunity_guid="g2", project_id="P1",
                       project_name="Muthgasse 109", state="verloren",
                       lost_reason="Zu teuer", parent_account_crm_id="bbb",
                       architect_crm_id="bbb", end_customer_crm_id="bbb",
                       created_on=_dt.datetime(2024, 9, 4)),
        # a project of ONE with no project_id — keyed by its own guid
        CrmOpportunity(crm_id="v3", opportunity_guid="g3", project_id=None,
                       project_name="Einzelobjekt", state="gewonnen",
                       order_value=1000.0, parent_account_crm_id="aaa"),
    ])
    s.commit(); s.close()
    monkeypatch.setattr(projekte, "SessionLocal", temp_db.SessionLocal)

    d = projekte.detail("P1")
    assert d["members"] == 2 and d["status"] == "gewonnen"
    # won through a sibling outside the window: say so, or "gewonnen · 0 gewonnene
    # VCs · kein Wert" reads like a bug
    assert d["won_members"] == 0 and d["won_via"]
    assert d["address"] == "Muthgasse 109 1190 Wien"
    byname = {f["name"]: f for f in d["firms"]}
    assert byname["Generalunternehmer B"]["roles"] == ["architekt", "endkunde", "kaeufer"]
    assert byname["Generalunternehmer B"]["vcs"] == 1          # one deal, not three
    assert [t["state"] for t in d["timeline"]] == ["verloren", "verloren"]  # oldest first
    assert "Zu teuer" in d["lost_reasons"]
    assert "Zugehörige VC gewonnen" not in d["lost_reasons"]

    # a single-VC project is reachable by its guid
    solo = projekte.detail("g3")
    assert solo is not None and solo["members"] == 1 and solo["won_members"] == 1
    assert projekte.detail("gibtsnicht") is None


def test_objekte_filter_by_number_of_verkaufschancen(temp_db, monkeypatch):
    """How many VCs hang on an Objekt is the strongest project-level signal we
    have (39,0 % against 19,3 %), so it has to be filterable in both directions.

    Two things that must not slip: a closed range excludes above as well as
    below, and the bucket table stays over ALL Objekte no matter what the table
    is filtered to — otherwise the reference row mirrors the filter and the
    comparison it exists to make disappears.
    """
    from adwatch.insights import projekte
    from adwatch.models import CrmOpportunity

    s = temp_db.SessionLocal()
    rows = []
    # P1: 1 VC, lost. P2: 2 VCs, won. P3: 3 VCs, lost.
    plan = [("P1", 1, "verloren"), ("P2", 2, "gewonnen"), ("P3", 3, "verloren")]
    for pid, n, state in plan:
        for i in range(n):
            rows.append(CrmOpportunity(
                crm_id=f"{pid}-{i}", opportunity_guid=f"{pid}-{i}", project_id=pid,
                project_name=pid, state=(state if i == 0 else "verloren"),
                order_value=(100.0 if state == "gewonnen" and i == 0 else None)))
    s.add_all(rows); s.commit(); s.close()
    monkeypatch.setattr(projekte, "SessionLocal", temp_db.SessionLocal)
    projekte.invalidate_cache()

    assert {r["name"] for r in projekte.list_projects()["rows"]} == {"P1", "P2", "P3"}
    assert {r["name"] for r in projekte.list_projects(min_members=2)["rows"]} == {"P2", "P3"}
    # closed range: excludes the 3-VC project ABOVE it, not just the 1 below
    exact2 = projekte.list_projects(min_members=2, max_members=2)
    assert {r["name"] for r in exact2["rows"]} == {"P2"} and exact2["total"] == 1
    assert {r["name"] for r in projekte.list_projects(max_members=1)["rows"]} == {"P1"}

    o = projekte.overview(min_members=2, max_members=2)
    assert o["projects"] == 1                      # the filtered population
    b = {x["label"]: x for x in o["member_buckets"]}
    assert [b["1 VC"]["projects"], b["2 VCs"]["projects"], b["3–4 VCs"]["projects"]] == [1, 1, 1]
    assert b["2 VCs"]["win_rate"] == 1.0 and b["1 VC"]["win_rate"] == 0.0
    assert sum(x["projects"] for x in o["member_buckets"]) == o["all_projects"] == 3


def test_dossier_carries_the_product_profile(temp_db, monkeypatch):
    """Everything pulled today landed in the database and none of it reached the
    drawer. The product profile is the whole answer to "which product for whom",
    so it has to travel with the dossier, and its euros must stay labelled as
    QUOTED — they span won and lost deals and are not revenue."""
    from adwatch import dossier
    from adwatch.models import Company, CrmCompanyProduct
    import datetime as _dt

    s = temp_db.SessionLocal()
    c = Company(name="Testbau", segment="Handel")
    s.add(c); s.commit()
    s.add_all([
        CrmCompanyProduct(company_id=c.id, family="cero", positions=9,
                          value=161845.0, first_seen=_dt.date(2021, 4, 4),
                          last_seen=_dt.date(2026, 3, 2)),
        # euros without positions: the value comes through the opportunity link,
        # the positions through the account link — they do not have to agree
        CrmCompanyProduct(company_id=c.id, family="Horizontale-Schiebewand",
                          positions=0, value=41000.0),
    ])
    s.commit(); cid = c.id; s.close()
    monkeypatch.setattr(dossier, "SessionLocal", temp_db.SessionLocal)

    d = dossier.build(cid)
    p = d["produkte"]
    assert [f["family"] for f in p["families"]] == ["cero", "Horizontale-Schiebewand"]
    assert p["value_quoted"] == 202845.0
    assert p["positions"] == 9
    assert p["first"] == "2021-04-04" and p["last"] == "2026-03-02"
    # a company with no product rows must not grow an empty block
    s2 = temp_db.SessionLocal()
    other = Company(name="Ohne Produkte", segment="Handel")
    s2.add(other); s2.commit(); oid = other.id; s2.close()
    assert dossier.build(oid)["produkte"] is None


def test_unproven_website_keeps_no_facts(temp_db, monkeypatch):
    """Facts and the identity verdict are written in one run, so they agree —
    until a verdict is REVISED. D3 Outdoor Girona kept a full profile (products,
    Corradi as an installed brand) read off a site the checker had already ruled
    was not theirs. A description with no website at all sat on 15 more rows."""
    from adwatch import dataquality as dq
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add_all([
        Company(name="Demoted", segment="Handel", identity_status="conflict",
                website_domain="fremd.de", description="von der falschen Seite",
                products=["Fenster"], competitor_brands=["Corradi"],
                enrichment_status="enriched"),
        Company(name="Ohne Website", segment="Handel", identity_status="not_found",
                description="woher auch immer", enrichment_status="enriched"),
        Company(name="Sauber", segment="Handel", identity_status="verified",
                website_domain="echt.de", description="belegt",
                products=["Wintergarten"], enrichment_status="enriched"),
    ])
    s.commit(); s.close()
    monkeypatch.setattr(dq, "SessionLocal", temp_db.SessionLocal)

    assert dq.clear_unbacked_enrichment(apply=False)["rows"] == 2   # dry run changes nothing
    dq.clear_unbacked_enrichment(apply=True)

    with temp_db.SessionLocal() as s2:
        rows = {c.name: c for c in s2.query(Company).all()}
        assert rows["Demoted"].description is None
        assert rows["Demoted"].competitor_brands is None
        assert rows["Demoted"].enrichment_status == "none"
        # the domain and the verdict STAY — they are the evidence the check ran
        assert rows["Demoted"].website_domain == "fremd.de"
        assert rows["Demoted"].identity_status == "conflict"
        # a verified row is untouched
        assert rows["Sauber"].description == "belegt"
    # idempotent
    assert dq.clear_unbacked_enrichment(apply=False)["rows"] == 0


def test_shared_domain_is_reported_not_merged(temp_db, monkeypatch):
    """Nearly a bad automated fix. Most shared domains are corporate GROUPS, not
    duplicates: Lindner has 8 legal entities on one website, each with its own
    SAP number and revenue. Only a shared domain AND a matching name means the
    same firm twice."""
    from adwatch import dataquality as dq
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add_all([
        Company(name="Lindner Building Envelope GmbH", segment="Handel", website_domain="lindner.com"),
        Company(name="Lindner Scandinavia AB", segment="Handel", website_domain="lindner.com"),
        Company(name="CBF", segment="Handel", website_domain="calviabalear.com"),
        Company(name="CBF S.L.", segment="Handel", website_domain="calviabalear.com"),
    ])
    s.commit(); s.close()
    monkeypatch.setattr(dq, "SessionLocal", temp_db.SessionLocal)

    out = dq.find_domain_duplicates()
    by_dom = {g["domain"]: g for g in out["top"]}
    assert by_dom["calviabalear.com"]["duplicate_pairs"] == [["CBF", "CBF S.L."]]
    # the Lindner entities share a domain but are different firms, so no pair
    assert "lindner.com" not in by_dom
    assert out["groups_with_a_duplicate_pair"] == 1
    # and nothing was deleted
    with temp_db.SessionLocal() as s2:
        assert s2.query(Company).count() == 4


def test_brand_evidence_outranks_an_inferred_fit():
    """Proymetal trades as "SUNFLEX Top-Partner", the scan found Sunflex in its
    logo strip, and the model still graded it "gering" — the extract it was given
    stopped before the brand and read like a general metalwork shop. Storing both
    would put a proven conquest target at the bottom of the ranking. Carrying a
    direct competitor is a FACT about the company and outranks the inference."""
    from adwatch.enrich.extract import apply_fit_floor, fit_floor
    assert apply_fit_floor("gering", ["Sunflex"]) == "hoch"
    assert apply_fit_floor(None, ["Vitrocsa"]) == "hoch"
    # terrace brands only lift to mittel — Markilux makes awnings, we do not
    assert apply_fit_floor("gering", ["Markilux"]) == "mittel"
    assert apply_fit_floor("hoch", ["Renson"]) == "hoch"       # never lowers
    # profile suppliers prove nothing about the category, so no floor at all
    assert apply_fit_floor("mittel", ["Cortizo", "Schüco"]) == "mittel"
    assert apply_fit_floor(None, ["Cortizo"]) is None
    assert fit_floor([]) is None


def test_brand_scan_survives_the_chrome_strip_and_the_char_budget():
    """The two edits that make the prose extract good — dropping navigation and
    capping characters — are the two that hide brand names, because a "Marcas"
    menu and a partner logo strip are chrome by every structural test. Dekovent
    lost Vitrocsa, Renson and Griesser that way and Proymetal lost Sunflex: the
    direct competitors, the ones worth most. Brands are a closed vocabulary, so
    they are found by scanning the WHOLE page, not by the model's trimmed extract."""
    from adwatch.identity.website_source import _page_text
    from adwatch.enrich.extract import scan_brands, brand_tiers
    html = ("<html><body>"
            "<nav><ul><li>Marcas</li><li>Vitrocsa</li><li>Renson</li></ul></nav>"
            "<main>Carpintería de aluminio en Valencia.</main>"
            "<footer>Distribuidor oficial de Sunflex</footer></body></html>")
    # the model's view has lost both menu brands ...
    assert "Vitrocsa" not in _page_text(html, limit=5000, drop_chrome=True)
    # ... the scan reads the full page and keeps them
    found = scan_brands(_page_text(html, limit=10 ** 7))
    assert {"Vitrocsa", "Renson", "Sunflex"} <= set(found)
    assert brand_tiers(found) == ["direkt", "terrasse"]


def test_brand_scan_does_not_match_ordinary_words():
    """A regex over brand names is only safe because the risky ones are excluded.
    "Roma" is a Mallorca street, "Keller" is a German surname and a basement —
    matching those would invent a supplier relationship out of an address."""
    from adwatch.enrich.extract import scan_brands
    assert scan_brands("Calle Roma 14, Palma. Herr Keller. Guardian Sapa Hydro") == []
    # and a real brand still has to stand alone, not sit inside another word
    assert scan_brands("Wir bauen Sunflex-Anlagen") == ["Sunflex"]
    assert scan_brands("cortizona ist keine Marke") == []


def test_scan_and_model_brands_are_merged_not_replaced():
    """The scan guarantees completeness over the page; the model catches loose
    phrasing a literal match misses. Storing either one alone loses companies."""
    from adwatch.enrich.service import _merge_brands
    assert _merge_brands(["Schüco"], ["Cortizo", "Schüco"]) == ["Schüco", "Cortizo"]
    assert _merge_brands(None, None) == []
    assert _merge_brands(["Sunflex"], ["sunflex"]) == ["Sunflex"]   # case-insensitive dedupe


def test_trailing_text_after_the_json_does_not_lose_the_extraction():
    """The model sometimes appends a sentence or a second object after the
    closing brace. json.loads() rejects the WHOLE reply as "Extra data", the
    caller swallows the exception and still stores the row as enriched — so
    Comervia and MODIKO ended up with every field empty and no visible failure.
    The object is intact and sits at the start; parse that and drop the rest."""
    import json as _json
    from adwatch.enrich.extract import _loads_first_object
    good = '{"description_de": "Architekturbüro.", "solarlux_relevance": "hoch"}'
    assert _loads_first_object(good)["solarlux_relevance"] == "hoch"
    # trailing prose, a second object, and leading chatter all survive
    assert _loads_first_object(good + "\n\nHinweis: geschätzt.")["description_de"] == "Architekturbüro."
    assert _loads_first_object(good + "\n" + good)["solarlux_relevance"] == "hoch"
    assert _loads_first_object("Hier das JSON:\n" + good)["solarlux_relevance"] == "hoch"
    # genuinely unparseable input must still raise, not return a silent {}
    import pytest as _pytest
    with _pytest.raises((ValueError, _json.JSONDecodeError)):
        _loads_first_object("kein JSON hier")


def test_navigation_menus_do_not_eat_the_extraction_budget():
    """TYPSA handed the extractor 9.000 characters of mega-menu — "Carreteras ·
    Ferrocarriles · Aeropuertos" — and not one sentence of prose, so every field
    came back null. Text is kept in document order, so a big menu starves the
    budget. Enrichment strips the chrome; the identity check must NOT, because it
    matches on the phone number and postcode that live in the footer."""
    from adwatch.identity.website_source import _page_text
    html = ("<html><head><title>Estudio</title></head><body>"
            "<nav><ul><li>Carreteras</li><li>Ferrocarriles</li><li>Aeropuertos</li></ul></nav>"
            "<header><ul class='main-menu'><li>Quiénes somos</li></ul></header>"
            "<main>Wir planen Villen an der Costa del Sol.</main>"
            "<footer>Tel. 952 123 456 · 29601 Marbella</footer></body></html>")
    clean = _page_text(html, limit=5000, drop_chrome=True)
    assert "Villen" in clean and "Carreteras" not in clean and "Quiénes somos" not in clean
    # the footer survives — it is evidence, not chrome
    assert "952 123 456" in clean and "29601" in clean
    # identity path is untouched: it still sees everything
    raw = _page_text(html, limit=5000)
    assert "Carreteras" in raw and "952 123 456" in raw


def test_stripping_chrome_never_empties_a_page():
    """Some small sites put their whole body inside <header>. Stripping would
    leave nothing, and a flooded extraction still beats an empty one, so the
    strip is abandoned when it removes essentially everything."""
    from adwatch.identity.website_source import _page_text
    html = ("<html><body><header>Estudio de arquitectura en Marbella. "
            "Wir planen Villen und Hotels seit 1998.</header></body></html>")
    kept = _page_text(html, limit=5000, drop_chrome=True)
    assert "Villen und Hotels" in kept


def test_relevance_is_judged_not_quoted():
    """Regression: the first architect run graded 0 of 72 Spanish offices "hoch"
    and put Costa-del-Sol villa studios on "gering". Cause: solarlux_relevance was
    asked for inside TEIL 1 — FAKTEN ("nichts schätzen"), while its "hoch" rubric
    required the site to MENTION large glazing. No architect writes that about
    their own work, so the top grade was unreachable and the residual bucket
    swallowed the best targets. Relevance must stay in the judgement half, key off
    project type, and never be the fallback for missing information."""
    from adwatch.enrich.extract import _prompt, PROFILE_ARCHITEKT
    arch = _prompt(PROFILE_ARCHITEKT)
    facts, judgement = arch.split("TEIL 2")
    # the grade is judged, not quoted
    assert "solarlux_relevance" not in facts
    assert "solarlux_relevance" in judgement
    # graded on what they build, and absence of glazing words is not evidence
    assert "ART DER PROJEKTE" in judgement and "KEIN Gegenbeleg" in judgement
    assert "Villen" in judgement and "Hotels" in judgement
    # null, not "gering", is the bucket for "cannot tell"
    assert "KEIN Auffangwert" in judgement
    # emitted after project_focus, so the grade is conditioned on the facts
    assert arch.index('"project_focus"') < arch.rindex('"solarlux_relevance"')


def test_architect_answer_maps_onto_shared_storage_keys(monkeypatch):
    """Same columns wherever the meaning carries over, so no downstream filter
    needs a per-segment branch: elements->products, specified_systems->
    competitor_brands, memberships->certifications. own_fabrication/has_showroom
    stay NULL — for a planner they are not 'no', they are not applicable."""
    import json as _json
    from adwatch.enrich import extract

    class _Blk:
        type = "text"
        text = _json.dumps({
            "description_de": "Architekturbüro für Wohn- und Hotelbauten.",
            "elements": ["Schiebetüren", "Fassade"],
            "specified_systems": ["Schüco", "Sky-Frame"],
            "solarlux_relevance": "hoch",
            "office_type": "Architekturbüro",
            "decision_role": "vergibt Aufträge",
            "project_focus": ["Wohnbau", "Hotel/Gastro"],
            "reference_scale": "über 200 Projekte",
            "memberships": ["COAM"],
            "founded_year": 1998, "employee_hint": "12 Architekten",
            "legal_form": "S.L.P.", "service_area": "Madrid",
            "mentions_solarlux": False, "evidence": {}, "assessment_de": "Gross.",
        })

    class _Msg:
        content = [_Blk()]

    class _Client:
        def __init__(self, **kw): self.messages = self
        def create(self, **kw):
            assert "ARCHITEKTUR" in kw["messages"][0]["content"]
            return _Msg()

    monkeypatch.setattr(extract.config, "ANTHROPIC_API_KEY", "test", raising=False)
    import anthropic
    monkeypatch.setattr(anthropic, "Anthropic", _Client)

    # the form must be IN the text — _legal_form_in_text rejects one that is not
    f = extract.extract_facts("Estudio S.L.P. " + "x" * 200,
                              profile=extract.PROFILE_ARCHITEKT)
    assert f["products"] == ["Schiebetüren", "Fassade"]      # plans with
    assert set(f["competitor_brands"]) == {"Schüco", "Sky-Frame"}  # specifies
    assert f["certifications"] == ["COAM"]
    assert f["solarlux_relevance"] == "hoch"
    assert f["decision_role"] == "vergibt Aufträge"
    assert f["own_fabrication"] is None and f["has_showroom"] is None
    assert f["legal_form"] == "S.L.P."
    assert f["profile"] == extract.PROFILE_ARCHITEKT


def test_architect_relevance_rejects_invented_values(monkeypatch):
    """A free-text answer outside the allowed set must become null, not be stored."""
    import json as _json
    from adwatch.enrich import extract

    class _Blk:
        type = "text"
        text = _json.dumps({"description_de": "x", "solarlux_relevance": "sehr hoch",
                            "office_type": "Weltmeister", "decision_role": "vielleicht",
                            "evidence": {}})

    class _Msg:
        content = [_Blk()]

    class _Client:
        def __init__(self, **kw): self.messages = self
        def create(self, **kw): return _Msg()

    monkeypatch.setattr(extract.config, "ANTHROPIC_API_KEY", "test", raising=False)
    import anthropic
    monkeypatch.setattr(anthropic, "Anthropic", _Client)
    f = extract.extract_facts("y" * 200, profile=extract.PROFILE_ARCHITEKT)
    assert f["solarlux_relevance"] is None
    assert f["office_type"] is None and f["decision_role"] is None


# ---------------------------------------------------------------------------
# Phase-0 audit findings, 2026-08-10. Each of these shipped and stayed invisible
# until the data was attacked directly rather than read through a filtered view.
# ---------------------------------------------------------------------------

def test_own_group_is_excluded_even_when_the_filter_names_ids(temp_db):
    """The intercompany guard used to be skipped whenever the winners filter
    carried `ids` — which is precisely the DEFAULT path, because
    material_buyer_ids() returns ids. Measured on the real base: 7 of 8 flagged
    own-group companies were in the default winners set. The existing
    intercompany test missed it because that DB has no Belege and therefore fell
    back to the id-less customer_state filter.

    An id list is a choice of POPULATION. It is never consent to train the
    profile on our own subsidiaries."""
    import datetime as dt
    from adwatch import customers
    from adwatch.insights import icp
    from adwatch.models import Company, CrmOrderEvent
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    for i in range(30):
        s.add(Company(name=f"Haendler {i}", country="DE",
                      segment="Handel" if i < 20 else "Verarbeiter",
                      postal_code="49134" if i < 20 else "80331"))
    s.add(Company(name="Linara Teststadt GmbH", country="DE", segment="Handel",
                  postal_code="49134"))
    s.commit()
    # every one of them a MATERIAL buyer, so material_buyer_ids() picks the
    # id-carrying path — the one that used to bypass the guard
    for c in s.scalars(select(Company)):
        s.add(CrmOrderEvent(company_id=c.id, order_date=dt.date(2025, 3, 1),
                            amount=50_000.0))
    s.commit()
    s.close()

    assert customers.flag_intercompany() == 1
    ids = icp.material_buyer_ids()
    assert len(ids) == 31, "alle 31 sind materielle Kaeufer - sonst testet das hier nichts"

    p = icp.build_profile(None)                    # the default path, uses ids
    assert p["winners_count"] == 30, "die eigene Gesellschaft sitzt im Gewinner-Set"
    assert p["winners_filter"].get("ids"), "der Default muss weiterhin ueber ids laufen"

    # and explicitly, too: naming the ids by hand must not smuggle them back in
    p2 = icp.build_profile({"ids": ids})
    assert p2["winners_count"] == 30


def test_out_of_scope_rows_never_carry_a_ranking(temp_db):
    """Private Endkunden had a winback_score on 1.449 rows and a fit_score on all
    1.665, because rfm.recompute() iterated select(Company) with no scope filter
    and nothing ever cleared scores written before the scope rule existed.
    overdue_customers() filtered them out of the VIEW, which is exactly what kept
    it invisible — any direct read of the column still got consumers.

    `health` is the deliberate exception: it is a fact about the row, not a
    position in a call list."""
    import datetime as dt
    from adwatch import dataquality as dq
    from adwatch.insights import rfm
    from adwatch.models import Company, CrmOrderEvent
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    for r in (Company(name="Echter Haendler", segment="Handel"),
              Company(name="Herr Privat", segment="Private Endkunden"),
              Company(name="Schueco Niederlassung", segment="Handel", is_competitor=True),
              Company(name="Linara Teststadt GmbH", segment="Handel", is_intercompany=True)):
        s.add(r)
    s.commit()
    for c in s.scalars(select(Company)):
        for yr in (2019, 2020, 2021, 2022):        # a cadence, then silence
            s.add(CrmOrderEvent(company_id=c.id, order_date=dt.date(yr, 3, 1),
                                amount=90_000.0))
        c.fit_score = 55.0
        c.target_score = 55.0
    s.commit()
    s.close()

    rfm.recompute(today=dt.date(2026, 8, 10))
    dq.clear_out_of_scope_scores(apply=True)

    s = temp_db.SessionLocal()
    by = {c.name: c for c in s.scalars(select(Company))}
    assert by["Echter Haendler"].winback_score > 0
    for name in ("Herr Privat", "Schueco Niederlassung", "Linara Teststadt GmbH"):
        assert by[name].winback_score is None, f"{name} steht auf einer Rueckgewinnungsliste"
        assert by[name].target_score is None, f"{name} steht auf der Zielliste"
        assert by[name].health is not None, f"{name} hat seine Historie verloren"
    # out of the business entirely -> nothing descriptive either; own group -> kept
    assert by["Herr Privat"].fit_score is None
    assert by["Linara Teststadt GmbH"].fit_score == 55.0
    s.close()

    assert dq.clear_out_of_scope_scores(apply=True)["rows"] == 0   # idempotent


def test_account_import_reflags_own_group(temp_db):
    """flag_intercompany() promised in its own docstring to run "on every import"
    and was called by nothing outside a test. Fourteen own-group companies were
    therefore unflagged — Nana Wall Systems Inc. (EUR 39,7 Mio) and Solarlux
    Nederland B.V. (EUR 21,9 Mio, 98% of all Dutch revenue) among them."""
    from adwatch import crm_accounts
    from adwatch.models import Company
    from sqlalchemy import select

    res = crm_accounts.upsert_accounts([
        {"accountid": "aaaaaaaa-0000-0000-0000-000000000001",
         "name": "Solarlux Nederland B.V."},
        {"accountid": "aaaaaaaa-0000-0000-0000-000000000002",
         "name": "Nana Wall Systems Inc."},
        {"accountid": "aaaaaaaa-0000-0000-0000-000000000003",
         "name": "Serin Bauelemente"},
    ])
    assert res["intercompany_reflagged"] == 2

    s = temp_db.SessionLocal()
    flags = {c.name: c.is_intercompany for c in s.scalars(select(Company))}
    s.close()
    assert flags["Solarlux Nederland B.V."] is True
    assert flags["Nana Wall Systems Inc."] is True
    assert flags["Serin Bauelemente"] is False


def test_executing_architect_is_not_always_an_architect():
    """`architect_crm_id` mirrors slx_executingarchitect_accountid — the
    AUSFUEHRENDER Architekt. A dealer that plans in-house enters itself, and on
    the real base that is 4.447 of 7.331 filled values (60,7%). Reading the field
    as "an architecture practice is involved" overstates it 2,5-fold: third-party
    architects appear on 2.884 of 57.776 Verkaufschancen (5,0%), not 12,7%."""
    from adwatch.insights.projekte import specifying_architect
    from adwatch.models import CrmOpportunity

    dealer, buero = "GUID-HAENDLER", "GUID-BUERO"
    assert specifying_architect(CrmOpportunity(
        crm_id="1", parent_account_crm_id=dealer, architect_crm_id=dealer)) is None
    assert specifying_architect(CrmOpportunity(
        crm_id="2", parent_account_crm_id=dealer, architect_crm_id=buero)) == buero
    assert specifying_architect(CrmOpportunity(
        crm_id="3", parent_account_crm_id=dealer, architect_crm_id=None)) is None
    # different casing must not create a phantom architect
    assert specifying_architect(CrmOpportunity(
        crm_id="4", parent_account_crm_id=dealer.lower(),
        architect_crm_id=dealer.upper())) is None


def test_audit_names_the_columns_that_are_really_outcomes(temp_db):
    """order_value / invoiced_value / sap_order_numbers are filled on ~92% of won
    deals and ~0% of lost ones. They sit in the same table as the legitimate
    features, so the audit has to name them rather than leave a reader to assume
    somebody checked."""
    from adwatch import audit
    from adwatch.models import CrmOpportunity

    s = temp_db.SessionLocal()
    for i in range(40):
        won = i < 20
        s.add(CrmOpportunity(crm_id=f"o{i}",
                             state="gewonnen" if won else "verloren",
                             order_value=1000.0 if won else None,
                             estimated_value=1000.0,
                             lost_reason=None if won else "Zu teuer"))
    s.commit()
    s.close()

    rep = audit.outcome_leakage()
    assert "order_value" in rep["unusable"]
    assert "estimated_value" not in rep["unusable"], "gleich gefuellt - also brauchbar"


def test_diagnose_compares_winners_against_a_real_baseline(temp_db):
    """diagnose() built its baseline by stripping `customer_state` from the
    winners filter. That was correct while the default winners set WAS
    customer_state — but it moved to {"ids": material_buyer_ids()} when the
    Belege became the source, so nothing got stripped, population == winners,
    every separation was 0.000 and the verdict was always "Kein Merkmal trennt
    die Gewinner von der Grundgesamtheit". Measured on the real base:
    diagnose(None) reported 3.781 winners against 3.781 population.

    A market slice (country/segment) must therefore compare the BUYERS in the
    slice against the slice, not the slice against itself."""
    import datetime as dt
    from adwatch.insights import icp
    from adwatch.models import Company, CrmOrderEvent
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    for i in range(60):
        s.add(Company(name=f"Haendler {i}", country="DE", segment="Handel",
                      sales_channel="Fachhandelsvertrieb",
                      # the buyers sit in one PLZ zone, the rest in another, so
                      # there IS something to find if the baseline is right
                      postal_code="49134" if i < 30 else "80331"))
    s.commit()
    for i, c in enumerate(s.scalars(select(Company).order_by(Company.id))):
        if i < 30:
            s.add(CrmOrderEvent(company_id=c.id, order_date=dt.date(2024, 3, 1),
                                amount=50_000.0))
    s.commit()
    s.close()

    d = icp.diagnose({"country": ["DE"], "segment": ["Handel"]})
    assert d["population"] == 60
    assert d["winners"] == 30, "die Kaeufer im Filter, nicht der ganze Filter"
    assert d["population"] > d["winners"], "Grundgesamtheit darf nicht die Gewinnermenge sein"
    plz = next(f for f in d["features"] if f["feature"] == "plz_zone")
    assert plz["separation"] > 0.4, f"PLZ trennt hier perfekt, gemessen {plz['separation']}"


def test_crm_product_families_are_time_gated(temp_db):
    """`crm_company_products` is the best-covered feature available (23.431
    companies against 1.207 for the website-derived list) — and it is written
    BY buying. Measured on the real base: between a cut two years back and today
    the family list grew by +0,74 entries for buyers and +0,22 for non-buyers.

    Scored without a cut it looks like a strong predictor; that is hindsight.
    crm_product_map(as_of=...) must therefore return only families first seen on
    or before the cut, and must drop rows with no first_seen at all rather than
    assume they were known."""
    import datetime as dt
    from adwatch.insights import icp
    from adwatch.models import Company, CrmCompanyProduct
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    s.add(Company(name="Haendler", country="DE", segment="Handel"))
    s.commit()
    cid = s.scalar(select(Company.id))
    s.add(CrmCompanyProduct(company_id=cid, family="Glas-Faltwand",
                            first_seen=dt.date(2023, 1, 1)))
    s.add(CrmCompanyProduct(company_id=cid, family="cero",
                            first_seen=dt.date(2026, 1, 1)))
    s.add(CrmCompanyProduct(company_id=cid, family="Wintergarten",
                            first_seen=None))
    s.commit()
    s.close()

    assert icp.crm_product_map()[cid] == ["Glas-Faltwand", "Wintergarten", "cero"]
    assert icp.crm_product_map(as_of=dt.date(2024, 1, 1))[cid] == ["Glas-Faltwand"]
    assert icp.crm_product_map(as_of=dt.date(2022, 1, 1)) == {}


def test_a_feature_known_mostly_for_winners_is_dropped(temp_db):
    """The availability guard, at the threshold that `crm_products` forced.

    Its family list is known for 92,4% of buyers and 50,4% of the base — ratio
    1,83. At the old threshold of 2,5 it passed, and the whole-base backtest
    jumped to lift 2,71 / ranks=True while Handel+Verarbeiter ALONE fell to
    0,75, i.e. worse than random. The feature was re-learning "this account has
    been worked", which separates architects from dealers and nothing else."""
    from adwatch.insights import icp
    assert icp._LEAK_RATIO <= 1.85, (
        "der Schwellenwert muss crm_products auf der Rohbasis fangen")


def test_list_features_cover_both_product_columns():
    """Two product columns with different provenance and 20x different coverage.
    Both are multi-valued, so every place that special-cases a list has to know
    about both — the count, the distribution, the lift and the score."""
    from adwatch.insights import icp
    assert set(icp._LIST_FEATURES) == {"products", "crm_products"}
    for f in icp._LIST_FEATURES:
        assert f in icp.DEFAULT_WEIGHTS
        assert f in icp._FEATURE_LABEL_DE


def test_company_verkaufschancen_are_complete_and_deduped(temp_db):
    """"Everything that has to do with this firm" cannot be a sample. The dossier
    ships the ten newest per role for its summary; this endpoint is the rest.

    Two rules it has to keep. A Verkaufschance counted once even when the company
    plays several roles on it (1.850 firms are Käufer AND Architekt AND Endkunde),
    and a `total` that is the real count — one company carries 1.266."""
    import datetime as dt
    from adwatch import dossier
    from adwatch.models import Company, CrmOpportunity
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    s.add(Company(name="Haendler", crm_id="GUID-A"))
    s.add(Company(name="Andere", crm_id="GUID-B"))
    s.commit()
    cid = s.scalar(select(Company.id).where(Company.name == "Haendler"))
    for i in range(25):
        s.add(CrmOpportunity(crm_id=f"o{i}", opportunity_guid=f"g{i}",
                             number=f"NR{i:03d}", parent_account_crm_id="guid-a",
                             state="gewonnen" if i < 5 else "verloren",
                             created_on=dt.datetime(2024, 1, 1) + dt.timedelta(days=i)))
    # same deal, three roles at once -> ONE Verkaufschance, not three
    s.add(CrmOpportunity(crm_id="multi", opportunity_guid="gmulti", number="NR999",
                         parent_account_crm_id="guid-a", architect_crm_id="guid-a",
                         end_customer_crm_id="guid-a", state="offen",
                         created_on=dt.datetime(2026, 1, 1)))
    # a deal that belongs to somebody else
    s.add(CrmOpportunity(crm_id="fremd", opportunity_guid="gfremd",
                         parent_account_crm_id="guid-b", state="offen",
                         created_on=dt.datetime(2026, 2, 1)))
    s.commit()
    s.close()

    r = dossier.verkaufschancen(cid, limit=10)
    assert r["total"] == 26, "25 plus die Mehrrollen-VC, die Fremde nicht"
    assert r["returned"] == 10
    assert r["by_role"] == {"kaeufer": 26, "architekt": 1, "endkunde": 1}
    assert r["rows"][0]["number"] == "NR999", "neueste zuerst"
    assert sorted(r["rows"][0]["roles"]) == ["architekt", "endkunde", "kaeufer"]

    # paging reaches the end and never repeats a row
    seen = []
    for off in (0, 10, 20):
        seen += [v["number"] for v in dossier.verkaufschancen(cid, limit=10, offset=off)["rows"]]
    assert len(seen) == 26 and len(set(seen)) == 26

    # role filter narrows both the rows and the counts
    only = dossier.verkaufschancen(cid, role="architekt")
    assert only["total"] == 1 and only["by_role"] == {"architekt": 1}


def test_dossier_objekte_come_from_all_vcs_not_the_last_ten(temp_db):
    """The Objekte list was derived from `blk["recent"]`, which is capped at ten.
    A company on 1.250 buildings therefore showed only those touched by its ten
    newest Verkaufschancen, and nothing said so."""
    import datetime as dt
    from adwatch import dossier
    from adwatch.models import Company, CrmOpportunity
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    s.add(Company(name="Haendler", crm_id="GUID-A"))
    s.commit()
    cid = s.scalar(select(Company.id))
    for i in range(30):
        s.add(CrmOpportunity(crm_id=f"o{i}", opportunity_guid=f"g{i}",
                             project_id=f"p{i}", project_name=f"Bauvorhaben {i}",
                             parent_account_crm_id="guid-a", state="verloren",
                             lost_reason="Zu teuer",
                             created_on=dt.datetime(2024, 1, 1) + dt.timedelta(days=i)))
    s.commit()
    s.close()

    d = dossier.build(cid)
    assert d["projekte_total"] == 30, "alle Objekte zaehlen, nicht nur die der letzten 10 VCs"
    assert len(d["projekte"]) <= 20            # die Liste selbst bleibt gedeckelt
    assert d["rollen"]["kaeufer"]["vcs"] == 30
    assert len(d["rollen"]["kaeufer"]["recent"]) == 10


# ---------------------------------------------------------------------------
# Enrichment identity gate, 2026-08-11. Every one of these was found by running
# 20 Spanish companies through the pipeline and reading what it wrote.
# ---------------------------------------------------------------------------

def test_a_shared_token_is_not_proof_unless_the_candidate_is_ours():
    """`domain_plus_name` means only that the domain shares a word with the
    company name. identity/find_website.PROVEN deliberately routes that to a
    human and ONBOARDING promises the same three hard signals — but
    enrich/service listed it as proof and wrote it straight into master data.

    Measured on the first 20 Spanish companies: "Montajes Portico Balear SL" ->
    portsdebalears.com and "+ PLUS" -> pressingplus.com, both wrong, both then
    enriched with a stranger's facts.

    The origin is the other half of the question: the company's own e-mail on
    that domain is corroboration a search result does not have."""
    from adwatch.enrich.service import _accepts

    assert not _accepts("serper", "domain_plus_name"), "eine geteilte Silbe ist kein Beweis"
    assert _accepts("email_domain", "domain_plus_name"), "die eigene Mail-Domain schon"
    assert _accepts("sap", "domain_plus_name")
    for hard in ("phone", "plz_street", "plz_name", "domain_in_name"):
        assert _accepts("serper", hard), f"{hard} ist ein harter Beweis"
    assert not _accepts("serper", None)
    assert not _accepts("email_domain", None), "ohne jedes Signal zaehlt auch die Herkunft nicht"


def test_spanish_directories_never_reach_the_review_queue():
    """The directory blocklist was built during German testing — 60+ entries, all
    German portals — so the Spanish equivalents ranked straight through it. A
    directory contains every company name by definition, which is exactly the
    signal _review_worthy trusts, so they clogged the queue instead of failing
    closed: "Carpintería Guerrero S.L." was offered qdq.com and "Montajes Portico
    Balear SL" got elpais.com.

    Substring matching cuts both ways, so the real company sites are asserted
    too — `elpaisajista.es` must survive `elpais.`."""
    from adwatch.enrich.website_finder import _is_directory
    from adwatch.enrich.domains import is_usable_company_domain

    for junk in ("qdq.com", "paginasamarillas.es", "einforma.com", "axesor.es",
                 "eleconomista.es", "elpais.com", "idealista.com", "expansion.com"):
        assert _is_directory(junk), f"{junk} ist ein Verzeichnis/Portal"
    for freemail in ("gmail.co.uk", "hotmail.es", "terra.es", "wanadoo.es"):
        assert not is_usable_company_domain(freemail), f"{freemail} ist Freemail"

    # and the ones that merely LOOK like an entry above
    for real in ("elpaisajista.es", "expansion-metallbau.de", "alurei.com",
                 "dorflex.net", "aluminioscerratosa.com"):
        assert _is_directory(real) is False, f"{real} ist eine echte Firmenseite"
        assert is_usable_company_domain(real)


def test_a_queued_company_always_has_something_to_decide(temp_db):
    """Two rules had drifted apart: _review_worthy put a company in the queue
    when its name appeared on the page, while the candidate was picked only from
    entries carrying a match signal. Seven of nine Spanish review items therefore
    reached the Pruefen tab with an empty Kandidat column — a decision with
    nothing to decide. One predicate now does both jobs."""
    from adwatch.enrich import service

    tried = [
        {"domain": "zufall.example", "origin": "serper", "signals": {}},
        {"domain": "treffer.example", "origin": "serper",
         "signals": {"name_in_domain": True}},
    ]
    worthy = [t for t in tried if service._review_worthy(t)]
    assert [t["domain"] for t in worthy] == ["treffer.example"], \
        "nur der Kandidat mit Signal ist eine Entscheidung wert"

    # a candidate that matched nothing must not put the company in the queue
    assert not any(service._review_worthy(t) for t in
                   [{"domain": "zeitung.example", "origin": "serper", "signals": {}}])


def test_searched_and_found_nothing_is_not_the_same_as_never_looked(temp_db):
    """Measured on job 57 at 661/1103: 248 Spanish companies came back
    'no_website_found' with a full candidate trail — searched properly, nothing
    provable — and every one kept identity_status NULL, which is exactly what the
    column says for a company nobody has touched.

    That is not cosmetic. find_website.pending_ids treats NULL as pending and
    NOT_FOUND as "do not re-spend", so the next search run would have paid Serper
    a second time for 248 answers already on record — the same mistake as the
    same-week ad re-fetch, on a different invoice.

    The verdict may only be written when a search REALLY ran: an enrichment pass
    with allow_search=False knows nothing about the wider web and must leave the
    question open rather than close it wrongly."""
    from adwatch.enrich import service

    searched = {"domain": None, "source": None, "validated_by": None,
                "review_candidate": None, "searched": True,
                "candidates": [{"domain": "fremd.example", "origin": "serper",
                                "validated": False, "signals": {}}],
                "bundle": None, "status": "no_website_found"}
    assert searched["searched"] is True

    # the resolver reports the distinction itself, from its own argument
    import adwatch.enrich.service as svc
    comp = {"name": "Sin Web SL", "website_domain": None, "email": None,
            "city": "Madrid", "country": "ES"}
    no_search = svc._resolve_website(comp, allow_search=False)
    assert no_search["status"] == "no_website_found"
    assert no_search["searched"] is False, \
        "ohne Suche darf nichts als 'nicht gefunden' abgeschlossen werden"

    # and the status the finder writes is the one that stops the re-spend
    from adwatch.identity.find_website import NOT_FOUND
    assert NOT_FOUND == "not_found"


def test_the_backfill_closes_only_what_was_really_searched(temp_db):
    """The repair for the rows already written without a verdict. A candidate
    trail is the evidence that a search ran; without one the row may have come
    from an allow_search=False pass, which knows nothing about the wider web and
    must not be allowed to close the question."""
    from adwatch import dataquality
    from adwatch.models import Company, CompanyEnrichment
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    s.add(Company(name="Gesucht SL", country="ES"))
    s.add(Company(name="Nie gesucht SL", country="ES"))
    s.add(Company(name="Schon geprueft SL", country="ES",
                  identity_status="verified", identity_matched_by="phone"))
    s.commit()
    ids = list(s.scalars(select(Company.id).order_by(Company.id)))
    s.add(CompanyEnrichment(company_id=ids[0], status="no_website_found",
                            website_candidates=[{"domain": "fremd.example",
                                                 "origin": "serper",
                                                 "validated": False}]))
    s.add(CompanyEnrichment(company_id=ids[1], status="no_website_found",
                            website_candidates=None))
    s.add(CompanyEnrichment(company_id=ids[2], status="no_website_found",
                            website_candidates=[{"domain": "x.example"}]))
    s.commit()
    s.close()

    dry = dataquality.close_searched_not_found(apply=False)
    assert dry["rows"] == 1, "nur die wirklich gesuchte Firma"

    dataquality.close_searched_not_found(apply=True)
    s = temp_db.SessionLocal()
    got = {c.name: c.identity_status for c in s.scalars(select(Company))}
    assert got["Gesucht SL"] == "not_found"
    assert got["Nie gesucht SL"] is None, "ohne Suchspur bleibt die Frage offen"
    assert got["Schon geprueft SL"] == "verified", "ein Urteil wird nie ueberschrieben"
    ev = s.scalar(select(Company).where(Company.name == "Gesucht SL")).identity_evidence
    assert ev["searched"] is True and ev["accepted"] is None
    s.close()

    # idempotent: a second run finds nothing left to do
    assert dataquality.close_searched_not_found(apply=False)["rows"] == 0


def test_the_cold_icp_refuses_the_two_poisoned_features(temp_db):
    """Gemessen 2026-08-13 an der deutschen Händlerbasis, beide Male als STARKE
    Prädiktoren aufgetaucht und beide Male Artefakt:

    * Vertriebsweg 'Direktvertrieb': n=55, Kaufrate 54,5% gegen 13,5% Basis —
      das beschreibt unsere Beziehung zur Firma, nicht die Firma.
    * Untersegment leer: n=209, Kaufrate 42,6% — Import-Herkunft. Und es ist die
      gefährliche Richtung: eine im Internet neu gefundene Firma hat EBENFALLS
      kein Untersegment und bekäme aus einem Grund, der nicht überträgt, eine
      hohe Punktzahl. Genau das würde die Zwillingssuche vergiften.

    Die Merkmalsfunktion darf beides nicht kennen."""
    from adwatch.insights import profiles
    from adwatch.models import Company

    c = Company(name="Musterfenster GmbH", segment="Handel",
                sub_segment="Fensterbau", country="DE", postal_code="49074",
                sales_channel="Direktvertrieb", website_domain="x.example")
    f = profiles._features_cold(c)
    assert not any("Direktvertrieb" in x or "vertriebsweg" in x.lower() for x in f), \
        "der Vertriebsweg beschreibt die Beziehung, nicht die Firma"
    assert "branche:Fensterbau" in f and "region:DE49" in f

    # ohne Untersegment darf KEIN Branchenmerkmal entstehen — auch kein 'leer'
    c2 = Company(name="Ohne Untersegment", segment="Handel", country="DE",
                 postal_code="49074")
    f2 = profiles._features_cold(c2)
    assert not any(x.startswith("branche:") for x in f2), \
        "'kein Untersegment' ist Herkunft, kein Merkmal"


def test_a_warranty_credit_is_not_a_purchase(temp_db):
    """14.049 der 91.992 Bestellereignisse stehen auf 0 EUR (Garantie, Muster,
    Ersatz), und 486 Firmen haben AUSSCHLIESSLICH solche — die galten als Kunden.
    Wer eine Garantiegutschrift als Erfolg zählt, trainiert das Modell darauf,
    Reklamationen vorherzusagen. Gemessen: 95 von 1.369 'Käufern' waren keine,
    und die Bereinigung bringt +0,012 AUC (0,617 -> 0,629).

    Die Zielgröße verlangt daher mindestens EIN Ereignis mit Betrag > 0."""
    import datetime as dt
    from adwatch.insights import profiles
    from adwatch.models import Company, CrmOrderEvent
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    s.add(Company(name="Echter Kaeufer", segment="Handel", sub_segment="Fensterbau",
                  country="DE", postal_code="49074"))
    s.add(Company(name="Nur Garantie", segment="Handel", sub_segment="Fensterbau",
                  country="DE", postal_code="49074"))
    s.commit()
    ids = {c.name: c.id for c in s.query(Company).all()}
    s.add(CrmOrderEvent(company_id=ids["Echter Kaeufer"],
                        order_date=dt.date(2025, 6, 1), amount=4000.0))
    s.add(CrmOrderEvent(company_id=ids["Nur Garantie"],
                        order_date=dt.date(2025, 6, 1), amount=0.0))
    s.commit()
    s.close()

    comps, pre, post, _vc = profiles._load(dt.date(2025, 1, 1), "DE")
    assert post[ids["Echter Kaeufer"]]["paid"] == 1
    assert post[ids["Nur Garantie"]]["n"] == 1, "die Bewegung existiert"
    assert post[ids["Nur Garantie"]]["paid"] == 0, "aber sie ist kein Kauf"


def test_the_at_risk_list_is_worth_calling(temp_db):
    """Die Kunden-Fortsetzung sortiert aufsteigend — die riskantesten zuerst.
    Ohne Wertgrenze besteht die Spitze aus Firmen mit EINER 40-Euro-Bestellung
    vor drei Jahren: mathematisch korrekt, betriebswirtschaftlich wertlos.
    Dieselbe 2.000-Euro-Schwelle wie im Bericht trennt Rettbares von Rauschen."""
    from adwatch.insights import profiles
    assert profiles.MATERIAL_EUR == 2000


def test_ipp_scores_lift_not_popularity(temp_db):
    """Die Lehre aus dem Firmen-ICP, auf Projekte übertragen: fit_for belohnte
    die HÄUFIGSTE Ausprägung der Gewinner (Bauelementehandel: 36,5% der Gewinner,
    Lift 1,03) und rankte damit exakt falsch herum. Das IPP muss Lift zahlen,
    nicht Popularität — und unter MIN_SUPPORT Gewinnern ist eine Ausprägung
    Anekdote und taucht gar nicht erst auf, egal wie perfekt ihre Quote ist."""
    from adwatch.insights import ipp, projekte

    # 100 entschiedene Projekte, Basisrate 20%: 'haeufig' tragen fast alle
    # (Gewinner wie Verlierer), 'selten-gut' nur 12 — davon 9 Gewinner.
    rows = []
    for i in range(100):
        won = i < 20
        feats = {"haeufig"}
        if (i < 9) or (77 <= i < 80):          # 9 Gewinner + 3 Verlierer
            feats = {"haeufig", "selten-gut"}
        if i == 0:
            feats |= {"perfekt-aber-3x"}       # 100%-Quote, aber nur 1 Gewinner
        rows.append(("k%d" % i, feats, projekte.WON if won else projekte.LOST,
                     2024, None))

    w = ipp._fit(rows)
    assert "perfekt-aber-3x" not in w, "unter dem Boden zählt keine perfekte Quote"
    assert "selten-gut" not in w, "9 Gewinner sind unter MIN_SUPPORT=10 — Anekdote"
    assert abs(w["haeufig"]["lift"] - 1.0) < 0.05, \
        "was jeder hat, sagt nichts — Lift ~1, nicht 'stark weil haeufig'"

    # jetzt mit genug Support: 15 Gewinner von 20 Trägern -> Lift deutlich > 1,
    # aber Laplace hält ihn UNTER der rohen Quote (75% / 20% = 3,75)
    rows2 = []
    for i in range(200):
        won = i < 40
        feats = {"basis"}
        if (i < 15) or (190 <= i < 195):
            feats = {"basis", "gut"}
        rows2.append(("j%d" % i, feats, projekte.WON if won else projekte.LOST,
                      2024, None))
    w2 = ipp._fit(rows2)
    raw = (15 / 20) / 0.2
    assert 1.5 < w2["gut"]["lift"] < raw, \
        "Laplace muss die kleine Stichprobe daempfen, nicht ausloeschen"

    # der Score eines Projekts mit dem guten Merkmal schlaegt eines ohne
    assert ipp._score({"basis", "gut"}, w2) > ipp._score({"basis"}, w2)


def test_health_reports_the_three_lifelines(temp_db):
    """Für den Betrieb als Dienst: EIN HTTP-Blick muss sagen, ob die App lebt.
    Der Endpunkt prüft die drei Lebensadern (DB, Sicherung, CRM-Sync-Alter) und
    antwortet mit dem STATUSCODE, nicht nur im JSON — ein stumpfer Uptime-Check
    ohne Parser muss alarmieren können. Ohne einzige Sicherung ist der Zustand
    'degraded' (503): eine Datenbank, deren Wert aus bezahlten Abrufen und
    menschlichen Urteilen besteht, läuft nie gesund ungesichert."""
    from fastapi.testclient import TestClient
    from adwatch import backup, web

    client = TestClient(web.app)
    r = client.get("/health")
    body = r.json()
    assert body["db"] == "ok"
    assert "job_running" in body
    if body["backup_last"] is None:
        assert r.status_code == 503 and body["status"] == "degraded", \
            "ohne Sicherung darf /health nicht 'ok' sagen"
    else:
        assert r.status_code in (200, 503)

    # nach einer Sicherung ist der Backup-Teil gesund
    backup.backup_now(tag="healthtest")
    r2 = client.get("/health")
    b2 = r2.json()
    if b2["backup_last"]:
        assert b2["backup_age_hours"] < 1


def test_the_map_never_pins_a_private_household(temp_db):
    """Die Karte zeigt Firmen. Private Endkunden sind Privatadressen — ein Pin
    auf deren Wohnung ist genau die Sorte Leck, die die scope-Klausel
    verhindern soll, und der Filter sitzt deshalb im SERVER, nicht im Frontend.

    Außerdem festgenagelt: der Zentroid-Lauf überschreibt nie eine genauere
    Koordinate ('street'/'manual'), die PLZ-Normalisierung lässt CRM- und
    GeoNames-Schreibweise aufeinandertreffen (NL '1234 AB' -> '1234'), und der
    Pin-Typ folgt dem Geld: wer je gekauft hat, ist Kunde, nicht Ziel."""
    import datetime as dt
    from adwatch import geo
    from adwatch.models import Company, CrmOrderEvent, PlzGeo
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    s.add(PlzGeo(country="ES", plz="08036", lat=41.39, lng=2.15, place="Barcelona"))
    s.add(PlzGeo(country="NL", plz="1234", lat=52.0, lng=4.3, place="Den Haag"))
    s.add(Company(name="Kaeufer SL", country="ES", postal_code="08036", segment="Handel"))
    s.add(Company(name="Estudio Arq", country="ES", postal_code="08036",
                  segment="Architekten"))
    s.add(Company(name="Privat E.", country="ES", postal_code="08036",
                  segment="Private Endkunden"))
    s.add(Company(name="NL Handel BV", country="NL", postal_code="1234 AB",
                  segment="Handel"))
    s.add(Company(name="Schon genau SL", country="ES", postal_code="08036",
                  segment="Handel", lat=41.11111, lng=2.11111,
                  geocode_precision="street"))
    s.commit()
    kid = s.scalar(select(Company.id).where(Company.name == "Kaeufer SL"))
    s.add(CrmOrderEvent(company_id=kid, order_date=dt.date(2025, 1, 1), amount=3000))
    s.commit()
    s.close()

    r = geo.assign_plz_centroids()
    assert r["geocoded"] == 4, "vier ohne bessere Quelle — auch der Privatkunde darf KOORDINATEN haben"
    assert r["kept_better"] == 1, "'street' wird nie durch einen Zentroid ersetzt"

    s = temp_db.SessionLocal()
    genau = s.scalar(select(Company).where(Company.name == "Schon genau SL"))
    assert abs(genau.lat - 41.11111) < 1e-6, "die genauere Koordinate blieb stehen"
    nl = s.scalar(select(Company).where(Company.name == "NL Handel BV"))
    assert nl.lat == 52.0, "NL-PLZ '1234 AB' trifft die GeoNames-Zeile '1234'"
    s.close()

    es = geo.pins(filters={"country": ["ES"]})
    names = {p["name"]: p for p in es["pins"]}
    assert "Privat E." not in names, "Privatadressen erscheinen auf KEINER Karte"
    assert names["Kaeufer SL"]["typ"] == "kunde", "wer je gekauft hat, ist Kunde"
    assert names["Estudio Arq"]["typ"] == "architekt"
    assert names["Schon genau SL"]["prec"] == "street"


def test_pipeline_board_counts_the_chain_honestly(temp_db):
    """Das Board zeigt die Kette, die die App ohnehin erzwingt — und es muss
    dieselben Regeln sprechen wie der Rest des Codes: Private Endkunden sind
    nie Teil eines Zählers (scope), 'not_found' ist ein Endstand und keine
    Lücke, Käufer zählen in der Qualifizierung nicht als Ziele, und unter
    MIN_WINNERS_USABLE Material-Käufern heißt der Modus 'scorecard'."""
    import datetime as dt
    from adwatch.insights import pipeline
    from adwatch.models import Company, CompanyEnrichment, CrmOrderEvent

    s = temp_db.SessionLocal()
    # 1: verifiziert + Fakten + Material-Käufer (2.500 €)
    s.add(Company(name="Kaeufer SL", country="ES", segment="Handel",
                  website_domain="kaeufer.example", identity_status="verified",
                  solarlux_fit="hoch"))
    # 2: Interessent mit Passung hoch — der eigentliche Zieltyp
    s.add(Company(name="Ziel SL", country="ES", segment="Verarbeiter",
                  identity_status="needs_review", solarlux_fit="hoch"))
    # 3: gesucht, nichts gefunden — Endstand, keine Lücke
    s.add(Company(name="Ohne Web SL", country="ES", segment="Handel",
                  identity_status="not_found"))
    # 4: Privatkunde — darf in KEINEM Zähler auftauchen
    s.add(Company(name="Privat", country="ES", segment="Private Endkunden",
                  identity_status="verified", solarlux_fit="hoch"))
    s.commit()
    ids = {c.name: c.id for c in s.query(Company).all()}
    s.add(CompanyEnrichment(company_id=ids["Kaeufer SL"], status="enriched",
                            fields={"description_de": "x"}))
    s.add(CrmOrderEvent(company_id=ids["Kaeufer SL"],
                        order_date=dt.date(2025, 3, 1), amount=2500.0))
    s.commit()

    st = pipeline.market_status(s, "ES")
    s.close()

    assert st["bestand"]["total"] == 3, "Privatkunden zaehlen nirgends mit"
    assert st["identitaet"]["verified"] == 1
    assert st["identitaet"]["offen"] == 1
    assert st["identitaet"]["not_found"] == 1
    assert st["anreicherung"]["mit_fakten"] == 1
    assert st["anreicherung"]["ohne_website_final"] == 1, \
        "not_found ist Endstand, keine Luecke"
    # der Käufer hat Passung hoch, zählt aber NICHT als Ziel — er ist Referenz
    assert st["qualifizierung"]["betriebe_hoch"] == 1, \
        "nur der Interessent, nicht der Kaeufer"
    assert st["bestand"]["kaeufer"] == 1
    assert st["icp"]["material_kaeufer"] == 1
    assert st["icp"]["modus"] == "scorecard", "1 Kaeufer liegt unter dem Boden"

    # und das Board waehlt bei unbekanntem Land den groessten Markt
    b = pipeline.board("XX")
    assert b["selected"] == "ES"


def test_profiles_are_cut_at_the_bottom_not_at_the_alphabet(temp_db):
    """The profile section inherited the overview's ordering: active advertisers
    first, then by NAME. Spain has nine advertisers and ~700 enriched companies,
    so in practice the section was alphabetical and the cut at limit=80 landed
    inside the letter A — 'Aluminios y Cristaleria Hisalma' was the last profile
    in the report. Every Betrieb with Passung hoch and every architect who awards
    contracts from B to Z was named in the qualification tables and then missing
    from the profiles underneath.

    The profiles now use the same tiers the qualification section ranks by, so
    the two sections agree on who matters."""
    from adwatch import report
    from adwatch.db import SessionLocal
    from adwatch.models import Company, CompanyEnrichment

    s = SessionLocal()
    # deliberately adversarial names: the best company sorts LAST alphabetically
    seed = [("Zenit Cerramientos SL", "hoch", None, None, ["Panoramah"]),
            ("Yebra Arquitectos", None, "hoch", "vergibt Aufträge", []),
            ("Alfa Aluminios SL", "gering", None, None, []),
            ("Beta Metalicas SL", "mittel", None, None, [])]
    for name, fit, rel, role, brands in seed:
        s.add(Company(name=name, country="ES", solarlux_fit=fit,
                      solarlux_relevance=rel, decision_role=role,
                      competitor_brands=brands, description=f"{name} Beschreibung"))
    s.commit()
    data = []
    for c in s.query(Company).all():
        s.add(CompanyEnrichment(company_id=c.id, status="enriched",
                                fields={"description_de": f"{c.name} Beschreibung"}))
        data.append({"company_id": c.id, "company": c.name,
                     "total_active_ads": 0, "has_data": False})
    s.commit()
    s.close()

    from reportlab.lib.styles import getSampleStyleSheet
    story = report._profiles_story(data, None, getSampleStyleSheet(), limit=2)
    text = " ".join(getattr(p, "text", "") for p in story)

    zenit, yebra = text.find("Zenit"), text.find("Yebra")
    alfa = text.find("Alfa Aluminios")
    assert zenit != -1 and yebra != -1, "Passung hoch und Relevanz hoch muessen drin sein"
    assert zenit < yebra, "Betrieb mit Passung hoch vor dem Buero"
    assert alfa == -1 or alfa > yebra, \
        "der alphabetisch erste, aber schlechteste Treffer darf den Platz nicht besetzen"
    assert "nie am Alphabet" in text, "der Schnitt muss sich erklaeren"


def test_a_company_is_not_fetched_twice_in_one_week(temp_db):
    """Measured 2026-08-11 over every ad fetch on record: 614 runs for only 300
    distinct company+source pairs, so 51% of all Apify spend bought a row that
    had already been bought. One pair was fetched ten times. The worst of it was
    a single morning of seven overlapping Spanish runs ("Status-Reparatur",
    "Lock-Retry", "Lauf 3", "Lauf 4"), each restarting from the top.

    It stayed invisible because the DATA never duplicated: WeeklyCompanyMetric is
    keyed on (company, source, week_start), so a re-fetch overwrote the row.
    Only the invoice grew.

    The ad week is the unit of freshness, and a successful run inside it means
    paid-for. A FAILED run does not count — that one deserves a retry."""
    import datetime as dt
    from adwatch import jobs
    from adwatch.collect.pipeline import monday_of
    from adwatch.models import CollectionRun, Company, CompanyPage
    from sqlalchemy import select

    week = monday_of(dt.date.today())
    s = temp_db.SessionLocal()
    for i in range(4):
        s.add(Company(name=f"Haendler {i}", website_domain=f"h{i}.example"))
    s.commit()
    ids = list(s.scalars(select(Company.id).order_by(Company.id)))
    # all four have a Meta page, so all four are meta-fetchable
    for cid in ids:
        s.add(CompanyPage(company_id=cid, source="meta", page_id=f"p{cid}", active=True))
    # #0 fetched fine this week, #1 came back empty (also a real answer),
    # #2 errored, #3 was fetched but LAST week
    s.add(CollectionRun(company_id=ids[0], source="meta", week_start=week, status="ok"))
    s.add(CollectionRun(company_id=ids[1], source="meta", week_start=week,
                        status="no_active_ads"))
    s.add(CollectionRun(company_id=ids[2], source="meta", week_start=week, status="error"))
    s.add(CollectionRun(company_id=ids[3], source="meta",
                        week_start=week - dt.timedelta(days=7), status="ok"))
    s.commit()
    s.close()

    s = temp_db.SessionLocal()
    fresh = jobs._fetched_this_week(s, ids)
    assert (ids[0], "meta") in fresh, "erfolgreich abgerufen = bezahlt"
    assert (ids[1], "meta") in fresh, "'keine Anzeigen' ist auch eine Antwort"
    assert (ids[2], "meta") not in fresh, "ein Fehlversuch darf erneut laufen"
    assert (ids[3], "meta") not in fresh, "letzte Woche ist nicht diese Woche"

    units = jobs._plan_units(s, ids, ["meta"])
    assert sorted(u[0] for u in units) == sorted([ids[2], ids[3]])

    # the override still exists for a deliberate same-week refresh
    forced = jobs._plan_units(s, ids, ["meta"], refetch=True)
    assert len(forced) == 4
    s.close()

    # and the pre-flight number explains the difference instead of just shrinking
    est = jobs.estimate(ids, ["meta"])
    assert est["total_units"] == 2
    assert est["fresh_skipped"] == 2
    assert jobs.estimate(ids, ["meta"], refetch=True)["total_units"] == 4


def test_all_fetched_this_week_says_so_instead_of_nothing_to_fetch(temp_db):
    """"Nothing to fetch" has two very different causes: nobody is fetchable, or
    everybody was already bought this week. The second is good news and needs to
    read as such, or the next person just forces a refetch to make the error go
    away."""
    import datetime as dt
    import pytest
    from adwatch import jobs
    from adwatch.collect.pipeline import monday_of
    from adwatch.models import CollectionRun, Company, CompanyPage
    from sqlalchemy import select

    week = monday_of(dt.date.today())
    s = temp_db.SessionLocal()
    s.add(Company(name="Haendler", website_domain="h.example"))
    s.commit()
    cid = s.scalar(select(Company.id))
    s.add(CompanyPage(company_id=cid, source="meta", page_id="p1", active=True))
    s.add(CollectionRun(company_id=cid, source="meta", week_start=week, status="ok"))
    s.commit()
    s.close()

    with pytest.raises(ValueError) as e:
        jobs.create_job([cid], ["meta"])
    assert "diese Woche schon abgerufen" in str(e.value)

    # forcing it works and creates a real job
    job = jobs.create_job([cid], ["meta"], refetch=True)
    assert job["total"] == 1


def test_a_resumed_fetch_job_continues_at_the_right_company(temp_db):
    """The freshness guard turned _plan_units from a pure function of the job's
    stored inputs into one that also depends on what the job ITSELF has fetched.
    _run_body rebuilt the unit list from those inputs on every start, so a
    resumed job would rebuild a SHORTER list while `completed` still counted
    positions in the original — and the cursor would skip past companies that
    were never fetched at all.

    Four companies, two fetched, then interrupted: the resumed job must run
    exactly the other two, not positions 3 and 4 of a list that lost its head."""
    import datetime as dt
    from adwatch import jobs
    from adwatch.collect.pipeline import monday_of
    from adwatch.models import CollectionRun, Company, CompanyPage, FetchJob
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    for i in range(4):
        s.add(Company(name=f"Haendler {i}", website_domain=f"h{i}.example"))
    s.commit()
    ids = list(s.scalars(select(Company.id).order_by(Company.id)))
    for cid in ids:
        s.add(CompanyPage(company_id=cid, source="meta", page_id=f"p{cid}", active=True))
    s.commit()
    s.close()

    job = jobs.create_job(ids, ["meta"], label="resume test")
    assert job["total"] == 4
    stored = [tuple(u) for u in job["plan"]["units"]]
    assert stored == [(cid, "meta") for cid in ids], "der Plan ist der Vertrag"

    # simulate: the first two ran, then the app died
    week = monday_of(dt.date.today())
    s = temp_db.SessionLocal()
    for cid in ids[:2]:
        s.add(CollectionRun(company_id=cid, source="meta", week_start=week, status="ok"))
    row = s.get(FetchJob, job["id"])
    row.completed = 2
    row.status = "interrupted"
    s.commit()
    s.close()

    # the plan must NOT shrink now that two of its companies count as fresh
    s = temp_db.SessionLocal()
    replanned = jobs._plan_units(s, ids, ["meta"])
    s.close()
    assert len(replanned) == 2, "ohne Plan wuerde neu geplant nur noch 2 Einheiten ergeben"

    again = jobs.get_job(job["id"])
    units = [tuple(u) for u in again["plan"]["units"]]
    assert units == stored, "der gespeicherte Plan bleibt unveraendert"
    # cursor 2 into the ORIGINAL plan -> the remaining two are companies 3 and 4
    assert units[again["completed"]:] == [(ids[2], "meta"), (ids[3], "meta")]


def test_a_worked_list_without_a_control_group_measures_nothing(temp_db):
    """Der Kern des Ganzen. 11,3 % der deutschen Händler kaufen auch ohne jeden
    Anruf — wer eine abgearbeitete Liste ohne Vergleichsgruppe auswertet,
    schreibt sich die Basisrate als Erfolg gut.

    Festgenagelt wird deshalb: (1) die Kontrollgruppe entsteht beim ANLEGEN und
    ist über die Rangfolge geschichtet, damit sie nicht zufällig die halbe Spitze
    verschluckt; (2) ein Kontakteintrag auf einer Kontroll-Firma wird ABGEWIESEN,
    nicht still gespeichert; (3) ohne Kontrollgruppe liefert die Messung `None`
    statt einer 0, die wie "gemessen, keine Wirkung" aussähe."""
    import datetime as dt
    import pytest
    from adwatch import outcomes
    from adwatch.models import Company, CrmOrderEvent
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    for i in range(40):
        s.add(Company(name=f"Firma {i:02d}", segment="Handel",
                      sub_segment="Fensterbau", country="DE", postal_code="49074"))
    s.commit()
    ids = list(s.scalars(select(Company.id).order_by(Company.id)))
    s.close()

    rows = [{"company_id": cid, "score": 1.0 - i / 100} for i, cid in enumerate(ids)]
    lst = outcomes.create_list("Testliste", "funnel", rows,
                               holdout_share=0.2, seed=42)
    assert lst["n"] == 40
    assert lst["n_kontrolle"] == 8, "20 % von 40"
    assert lst["n_ziel"] == 32

    # geschichtet: in JEDEM Zehnerblock sitzen genau 2 Kontrollen — eine rein
    # zufällige Ziehung könnte 8 davon in die Top-10 legen
    ent = outcomes.entries(lst["id"])
    for start in (0, 10, 20, 30):
        block = ent[start:start + 10]
        assert sum(1 for e in block if e["arm"] == "kontrolle") == 2, \
            f"Block ab Rang {start+1} ist nicht geschichtet"

    ziel = next(e for e in ent if e["arm"] == "ziel")
    kontrolle = next(e for e in ent if e["arm"] == "kontrolle")

    outcomes.record(ziel["entry_id"], outcome="angebot", channel="telefon")
    with pytest.raises(ValueError) as err:
        outcomes.record(kontrolle["entry_id"], outcome="angebot", channel="telefon")
    assert "Kontrollgruppe" in str(err.value)

    # gemessen wird am HARTEN Ausgang, nicht am eingetragenen Ergebnis
    s = temp_db.SessionLocal()
    s.add(CrmOrderEvent(company_id=ziel["company_id"],
                        order_date=dt.date.today(), amount=5000.0))
    s.commit(); s.close()
    m = outcomes.measure(lst["id"], since=dt.date.today() - dt.timedelta(days=1))
    assert m["ziel"]["kaeufer"] == 1
    assert m["kontrolle"]["kaeufer"] == 0
    assert m["uplift"] is not None
    assert m["aussagekraeftig"] is False, "8 Kontrollen sind zu wenig fuer eine Aussage"

    # eine Liste GANZ ohne Kontrollgruppe darf keinen Uplift behaupten
    blind = outcomes.create_list("Ohne Kontrolle", "kalt", rows,
                                 holdout_share=0.0, seed=42)
    mb = outcomes.measure(blind["id"])
    assert mb["kontrolle"]["n"] == 0
    assert mb["uplift"] is None, "ohne Vergleichsgruppe ist Wirkung nicht definiert"


def test_a_zero_euro_order_does_not_count_as_list_success(temp_db):
    """Dieselbe 0-Euro-Regel wie in den Profilen: eine Garantiegutschrift nach
    einem Anruf ist kein Erfolg der Liste."""
    import datetime as dt
    from adwatch import outcomes
    from adwatch.models import Company, CrmOrderEvent
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    for i in range(20):
        s.add(Company(name=f"F{i}", segment="Handel", sub_segment="Glaser",
                      country="DE", postal_code="49074"))
    s.commit()
    ids = list(s.scalars(select(Company.id).order_by(Company.id)))
    s.add(CrmOrderEvent(company_id=ids[0], order_date=dt.date.today(), amount=0.0))
    s.commit(); s.close()

    lst = outcomes.create_list("Null-Euro", "kalt",
                               [{"company_id": c, "score": 0.5} for c in ids],
                               holdout_share=0.0, seed=1)
    m = outcomes.measure(lst["id"], since=dt.date.today() - dt.timedelta(days=1))
    assert m["ziel"]["kaeufer"] == 0, "eine 0-Euro-Bewegung ist kein Kauf"


def test_discovery_matches_known_companies_two_ways(temp_db):
    """Der Abgleich entscheidet über das Ergebnis des ganzen Versuchs. Nur über
    die Domain zu prüfen würde 'neu' systematisch überschätzen: von 10.998
    deutschen Händlern haben bloß 5.463 (49 %) überhaupt eine Domain hinterlegt.
    Deshalb zusätzlich Name+Ort — und deshalb wird das Ergebnis als SPANNE
    berichtet, nicht als eine Zahl."""
    from adwatch import discover
    from adwatch.models import Company
    from sqlalchemy import select

    s = temp_db.SessionLocal()
    s.add(Company(name="Mustermann Fensterbau GmbH", city="Osnabrück",
                  country="DE", website_domain="mustermann-fenster.de",
                  segment="Verarbeiter"))
    s.add(Company(name="Sonnenschein Glaserei", city="Münster", country="DE",
                  segment="Verarbeiter"))          # KEINE Domain hinterlegt
    s.commit()
    ids = {c.name: c.id for c in s.query(Company).all()}
    s.close()

    by_domain, by_city = discover._known_index()

    # Weg 1: harte Domain
    cid, how = discover._match_known(
        {"domain": "mustermann-fenster.de", "title": "Irgendein Titel"},
        by_domain, by_city)
    assert cid == ids["Mustermann Fensterbau GmbH"] and how == "domain"

    # Weg 2: Name + Ort, für die Hälfte des Bestands ohne Domain
    cid2, how2 = discover._match_known(
        {"domain": "sonnenschein-glas.de",
         "title": "Sonnenschein Glaserei — Ihr Glaser in Münster"},
        by_domain, by_city)
    assert cid2 == ids["Sonnenschein Glaserei"] and how2 == "name_ort"

    # eine wirklich fremde Firma bleibt neu
    cid3, _ = discover._match_known(
        {"domain": "voellig-fremd.de", "title": "Völlig Fremd GmbH, Kiel"},
        by_domain, by_city)
    assert cid3 is None

    # Rechtsform und Umlaute dürfen den Vergleich nicht sprengen
    assert discover._norm_name("Müller & Söhne GmbH") == discover._norm_name("Mueller & Soehne")


def test_an_empty_dataverse_filter_never_leaves_the_app(temp_db):
    """Gemessen 2026-08-18: drei Testabfragen ohne Filter, drei fehlgeschlagene
    Flow-Läufe, HTTP 502 NoResponse beim Aufrufer. Der Konnektor lehnt einen
    leeren $filter ab — und weil die Aktion scheitert, erreicht der Lauf die
    Response nie, sodass der Fehler wie ein Netzwerkproblem aussieht statt wie
    ein Eingabefehler.

    Aufgefallen war es jahrelang nicht, weil jeder echte Aufruf einen Filter
    trug (`modifiedon gt ...`). Die erste ungefilterte Abfrage stolperte darüber.
    Die Vorgabe wird deshalb HIER gesetzt, an der einzigen Stelle, durch die
    jeder Flow-Aufruf läuft."""
    from adwatch import flows

    # leer, fehlend, nur Leerzeichen -> alle bekommen die Vorgabe
    for payload in ({"entity": "leads", "select": "leadid", "filter": ""},
                    {"entity": "leads", "select": "leadid"},
                    {"entity": "leads", "select": "leadid", "filter": "   "}):
        out = flows._guard_payload("crm_query", payload)
        assert out["filter"] == flows._DEFAULT_DATAVERSE_FILTER
        assert out["entity"] == "leads", "der Rest bleibt unangetastet"

    # ein echter Filter wird NIE überschrieben
    real = {"entity": "accounts", "select": "accountid",
            "filter": "modifiedon gt 2026-01-01T00:00:00Z"}
    assert flows._guard_payload("crm_query", real)["filter"] == real["filter"]

    # andere Rollen bleiben unberührt — die Vorgabe ist Dataverse-spezifisch
    mail = {"recipient": "x@y.de"}
    assert flows._guard_payload("report_email", mail) == mail


# ---------------------------------------------------------------------------
# Ein "fertig" ohne Abdeckungsprüfung ist wertlos
# ---------------------------------------------------------------------------

def test_email_coverage_findet_luecken_und_teilmonate(temp_db):
    """`coverage()` muss die zwei gemessenen Ausfallarten des E-Mail-Abrufs
    finden — und zwar BEIDE.

    Der Erstabruf am 2026-08-18 verlor 7 von 41 Monaten am Flow-Timeout und lief
    trotzdem sauber durch: die Schleife fängt Fehler ab, damit ein schlechter
    Monat keinen Vier-Stunden-Lauf killt. Das Ergebnis SAH vollständig aus,
    während rund 65.000 Mails fehlten.

    Der zweite Fall ist der tückischere: 2026-05 stand mit 2.834 statt ~9.500
    Zeilen in der Datenbank, Rest eines Testlaufs. Über Anwesenheit allein ist
    das NICHT zu finden — der Monat ist da, nur eben zu einem Drittel.
    """
    import datetime as dt
    from adwatch import crm_emails
    from adwatch.models import CrmEmail

    def add(s, monat: str, n: int):
        for i in range(n):
            s.add(CrmEmail(activity_id=f"{monat}-{i}",
                           created_on=dt.datetime.fromisoformat(f"{monat}-05T09:00:00")))

    s = temp_db.SessionLocal()
    add(s, "2024-01", 100)
    add(s, "2024-02", 100)
    # 2024-03 fehlt komplett — Flow-Timeout
    add(s, "2024-04", 100)
    add(s, "2024-05", 5)        # Teilabruf: da, aber weit unter dem Median
    s.commit(); s.close()

    cov = crm_emails.coverage(dt.date(2024, 1, 1), dt.date(2024, 6, 1))

    assert cov["monate_erwartet"] == 5
    assert cov["fehlend"] == ["2024-03"], "der komplett fehlende Monat"
    assert cov["duenn"] == ["2024-05"], "der Teilmonat, den Anwesenheit übersieht"
    assert cov["median_pro_monat"] == 100
    assert cov["vollstaendig"] is False

    # Ein lückenloser Zeitraum darf nicht fälschlich Alarm schlagen.
    ok = crm_emails.coverage(dt.date(2024, 1, 1), dt.date(2024, 3, 1))
    assert ok["fehlend"] == [] and ok["duenn"] == []
    assert ok["vollstaendig"] is True

    # Der laufende Monat ist ZU RECHT unvollständig und darf nie als "dünn"
    # gemeldet werden — sonst ist die Prüfung jeden Tag rot.
    s = temp_db.SessionLocal()
    heute = dt.date.today()
    add(s, f"{heute:%Y-%m}", 1)
    s.commit(); s.close()
    lauf = crm_emails.coverage(heute.replace(day=1),
                               (heute.replace(day=1) + dt.timedelta(days=32)).replace(day=1))
    assert lauf["duenn"] == [], "der laufende Monat wird ausgenommen"



# ---------------------------------------------------------------------------
# Leads: die Antwortform des Flows, und was NICHT aufgeloest werden darf
# ---------------------------------------------------------------------------

def test_lead_antwortform_und_aufloesung(temp_db):
    """Zwei Stellen, an denen der Lead-Abruf still falsch laufen wuerde.

    1. FORM. Der Flow liefert fuer `leads` ein nacktes Array, fuer `accounts`
       dagegen {value: [...]}. Wer sich auf eine Form verlaesst, bekommt beim
       anderen Entity null Zeilen -- und zwar ohne Fehler, was der schlimmste
       Fall ist: der Abruf meldet Erfolg und laedt nichts.

    2. AUFLOESUNG. Ein Lead wird NUR ueber die im CRM gesetzte Mutterfirma auf
       eine Firma gezogen, nie ueber Namensaehnlichkeit. "Fenster Meier" und
       "Meier Fenster- und Tuerenbau GmbH" koennen dieselbe Firma sein oder
       nicht -- das entscheidet kein Stringvergleich, und eine falsche
       Verknuepfung vergiftet jede spaetere Auswertung.
    """
    from adwatch import crm_leads
    from adwatch.models import Company

    assert crm_leads._rows([{"leadid": "a"}]) == [{"leadid": "a"}]
    assert crm_leads._rows({"value": [{"leadid": "b"}]}) == [{"leadid": "b"}]
    assert crm_leads._rows({}) == []
    assert crm_leads._rows(None) == []

    s = temp_db.SessionLocal()
    c = Company(name="Meier Fenster- und Tuerenbau GmbH", crm_id="GUID-1",
                resolution_status="confirmed", country="DE")
    s.add(c); s.commit()
    resolve = crm_leads._company_resolver(s)

    assert resolve("GUID-1") == c.id, "gesetzte Mutterfirma wird aufgeloest"
    assert resolve("GUID-UNBEKANNT") is None
    assert resolve(None) is None, "ohne Mutterfirma bleibt die Frage offen"
    s.close()


def test_lead_holt_keine_personendaten():
    """Die Feldliste ist eine Zusage, keine Bequemlichkeit.

    firstname, lastname, emailaddress1 und telephone1 stehen in Dataverse und
    waeren einen Tastendruck entfernt. Sie duerfen nicht in SELECT stehen --
    gespeichert wird ausschliesslich, was die FIRMA beschreibt."""
    from adwatch import crm_leads
    for feld in ("firstname", "lastname", "emailaddress", "telephone",
                 "mobilephone", "fullname"):
        assert feld not in crm_leads.SELECT, f"{feld} ist eine Personendatei"
    assert "companyname" in crm_leads.SELECT



# ---------------------------------------------------------------------------
# Fragen-Agent: das SQL-Werkzeug ist die Sicherheitsgrenze
# ---------------------------------------------------------------------------

def test_fragen_sql_werkzeug_ist_nur_lesend(temp_db):
    """Das SQL-Werkzeug des Fragen-Agenten darf ALLES lesen und NICHTS koennen,
    was schreibt. Die Pruefung hat zwei Schichten, und beide werden getestet:
    die Textpruefung (lehnt ab) und die read-only-Verbindung (koennte selbst
    dann nicht schreiben, wenn die Textpruefung versagt)."""
    import pytest as _pytest
    from adwatch import fragen
    from adwatch.models import Company
    import json as _json

    s = temp_db.SessionLocal()
    s.add(Company(name="Test AG", country="DE", segment="Handel"))
    s.commit(); s.close()

    # lesen geht, LIMIT wird erzwungen
    out = _json.loads(fragen.w_sql("select name, country from companies"))
    assert out["zeilen"] == [["Test AG", "DE"]]

    # jede Schreib- oder Struktur-Anweisung scheitert an der Textpruefung
    for boese in ("update companies set name='x'",
                  "delete from companies",
                  "insert into companies(name) values('x')",
                  "drop table companies",
                  "select 1; delete from companies",       # zweite Anweisung
                  "pragma writable_schema=1",
                  "attach database ':memory:' as x"):
        with _pytest.raises(ValueError):
            fragen.w_sql(boese)

    # CTEs sind erlaubt — WITH ist lesend
    out = _json.loads(fragen.w_sql(
        "with t as (select count(*) n from companies) select n from t"))
    assert out["zeilen"] == [[1]]


def test_fragen_werkzeuge_vollstaendig_registriert():
    """Jedes Werkzeug braucht Name, Beschreibung, Schema und Funktion — ein
    unvollstaendiger Eintrag faellt sonst erst beim ersten API-Aufruf um,
    mitten in einer bezahlten Frage."""
    from adwatch import fragen
    namen = set()
    for w in fragen.WERKZEUGE:
        assert w["name"] and w["description"] and callable(w["fn"])
        assert w["input_schema"]["type"] == "object"
        assert w["name"] not in namen, "doppelter Werkzeugname"
        namen.add(w["name"])
    # die Angebots-Regel muss dem Modell an den zwei Stellen begegnen,
    # an denen es Zahlen erzeugt: Systemprompt und Datenbestand-Werkzeug
    assert "keine Rechnungen" in fragen._SYSTEM



def test_fragen_ergebnis_ist_immer_gueltiges_json(temp_db):
    """Ein gekuerztes Werkzeug-Ergebnis muss LESBAR bleiben.

    Gefunden am 2026-08-20 im End-to-End-Test: `_j` serialisierte erst und
    schnitt dann bei 6.000 Zeichen ab -- mitten in einem Firmennamen. Der Agent
    bekam kaputtes JSON zurueck und haette mitten in einer bezahlten Frage
    daran gescheitert. Jetzt werden ZEILEN entfernt statt Zeichen, und das
    Ergebnis sagt selbst, dass es gekuerzt ist.
    """
    import json as _json
    from adwatch import fragen

    # Liste, die weit ueber dem Deckel liegt
    gross = [{"id": i, "name": f"Musterfirma Nummer {i} GmbH & Co. KG", "ort": "Musterstadt"}
             for i in range(400)]
    text = fragen._j(gross)
    assert len(text) <= fragen.MAX_ERGEBNIS_ZEICHEN
    d = _json.loads(text)                      # <- hier scheiterte es vorher
    assert d["gekuerzt"], "die Kuerzung muss sichtbar sein"
    assert len(d["zeilen"]) < 400

    # dasselbe fuer ein Dict mit langer Liste darin
    text = fragen._j({"spalten": ["id", "name"], "zeilen": gross})
    d = _json.loads(text)
    assert len(d["zeilen"]) < 400 and "gekuerzt" in d

    # kleine Ergebnisse bleiben unangetastet
    d = _json.loads(fragen._j({"a": 1}))
    assert d == {"a": 1}



def test_kundenklasse_schliesst_nichts_aus(temp_db):
    """sl_customer_class darf NIEMANDEN aus der Auswertung werfen.

    Am 2026-08-20 wurde "07 - SL Mitarbeiter" als "Konto einer Solarlux-Person"
    gelesen und ausgeschlossen. Der volle Abruf widerlegte das: 8.204 Konten
    tragen die Klasse, 7.302 davon im Haendler-Panel, zusammen 104,6 Mio EUR
    Angebotsvolumen -- darunter MADEROS Wintergaerten, LEEB Balkone, Willab
    Garden AB. Das sind Haendler. Die Klasse bedeutet die BETREUUNGSART, nicht
    die Person.

    Der Ausschluss haette fast die halbe Grundgesamtheit geloescht (6.207 ->
    1.858 allein im Kalt-Profil), inklusive der groessten Konten. Dieser Test
    haelt die Korrektur fest, damit die naheliegende Fehllesung nicht
    zurueckkehrt."""
    from sqlalchemy import select
    from adwatch import scope
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add_all([
        Company(name="MADEROS Wintergaerten", segment="Verarbeiter", country="DE",
                sl_customer_class="07 - SL Mitarbeiter"),
        Company(name="Fachhandel", segment="Handel", country="DE",
                sl_customer_class="02 - Fachhandelsvertrieb"),
        Company(name="Privatperson", segment="Private Endkunden", country="DE"),
    ])
    s.commit()

    drin = {n for (n,) in s.execute(
        select(Company.name).where(scope.in_scope_clause()))}
    assert "MADEROS Wintergaerten" in drin, "Betreuungsart ist kein Ausschlussgrund"
    assert "Fachhandel" in drin
    assert "Privatperson" not in drin, "Private Endkunden bleiben ausgeschlossen"
    s.close()


def test_profil_bevoelkerung_kennt_die_zukunft_nicht(temp_db):
    """Eine Firma, die erst NACH dem Stichtag im CRM angelegt wurde, darf nicht
    in der Bevoelkerung stehen -- am Stichtag kannten wir sie nicht.

    Das war ein echter Fehler: frisch angelegte Konten fragen mit 33,1 % an,
    alte mit 22,1 %, weil eine Firma oft ANGELEGT wird, WEIL sie angefragt hat.
    Damit sagte das Anlagedatum die Anfrage voraus. Garten- und Landschaftsbau
    stand mit Lift 4,45 an der Spitze der Kalt-Liste -- 51 seiner 57 Konten
    stammten aus 2024+. Nach der Korrektur faellt das Gewerk unter die
    Traegergrenze und verschwindet."""
    import datetime as dt
    from adwatch.insights import profiles
    from adwatch.models import Company

    s = temp_db.SessionLocal()
    s.add_all([
        Company(name="Alt", segment="Handel", country="DE",
                crm_created_on=dt.datetime(2020, 5, 1)),
        Company(name="Neu", segment="Handel", country="DE",
                crm_created_on=dt.datetime(2025, 6, 1)),
        Company(name="Ohne Datum", segment="Handel", country="DE"),
    ])
    s.commit(); s.close()

    comps, _pre, _post, _vc = profiles._load(dt.date(2025, 1, 1))
    namen = {c.name for c in comps}
    assert "Alt" in namen
    assert "Ohne Datum" in namen, "ohne Anlagedatum laesst sich nichts ausschliessen"
    assert "Neu" not in namen, "am Stichtag gab es diese Firma bei uns noch nicht"



def test_personensuche_ohne_flow_bricht_nichts():
    """Die Empfaengerpflege darf NIE an einer Zusatzfunktion haengen.

    Ist der Personen-Flow nicht eingerichtet (der Normalfall bei jeder frischen
    Installation), muss die Suche eine leere Liste liefern statt zu werfen --
    das Feld faellt dann auf freie Eingabe zurueck."""
    from adwatch import people
    assert people.suchen("Mueller") == []
    assert people.suchen("") == []
    assert people.suchen("a") == [], "unter zwei Zeichen wird gar nicht gefragt"


def test_personensuche_versteht_beide_antwortformen():
    """Der Flow liefert je nach Aufbau eine nackte Liste oder {value: [...]},
    und die Feldnamen unterscheiden sich je nach Connector-Version
    (mail vs. userPrincipalName, displayName vs. DisplayName). Wer eine Form
    voraussetzt, bekommt beim anderen Aufbau still null Zeilen -- derselbe
    Fehler, der beim Lead-Abruf schon einmal zuschlug."""
    from adwatch import people

    assert people._rows([{"mail": "a@b.de"}]) == [{"mail": "a@b.de"}]
    assert people._rows({"value": [{"mail": "a@b.de"}]}) == [{"mail": "a@b.de"}]
    assert people._rows({}) == []
    assert people._rows(None) == []

    # beide Schreibweisen ergeben denselben Datensatz
    a = people._norm({"displayName": "Iheb Marouani", "mail": "i.m@solarlux.com",
                      "jobTitle": "BD", "department": "Strategie"})
    b = people._norm({"DisplayName": "Iheb Marouani", "UserPrincipalName": "i.m@solarlux.com",
                      "JobTitle": "BD", "Department": "Strategie"})
    assert a == b
    assert a["email"] == "i.m@solarlux.com" and a["name"] == "Iheb Marouani"

    # ohne brauchbare Adresse ist eine Zeile als Empfaenger wertlos
    assert people._norm({"displayName": "Ohne Mail"}) is None
    assert people._norm({"displayName": "Kaputt", "mail": "keine-adresse"}) is None


def test_teams_link_nur_bei_echter_adresse():
    """Teams laesst sich nicht einbetten, aber ein Deep Link tut es auch --
    ohne jede Berechtigung. Nur muss die Adresse eine sein."""
    from adwatch import people
    link = people.teams_link("i.marouani@solarlux.com")
    assert link and link.startswith("https://teams.microsoft.com/l/chat/0/0?users=")
    assert "i.marouani%40solarlux.com" in link, "Adresse muss kodiert sein"
    assert people.teams_link("kein-at-zeichen") is None
    assert people.teams_link("") is None
    assert people.teams_link(None) is None



def test_projektwert_ist_die_primaere_vc_nicht_die_summe(temp_db):
    """Der Wert eines Objekts ist der Wert seiner PRIMAEREN Verkaufschance.

    Vorher wurden alle Geschwister addiert. An einem Gebaeude bekommen aber
    mehrere Haendler und Generalunternehmer dasselbe Gewerk angeboten --
    gewinnen kann nur einer. Karlsruhe, Rheinstrasse 91 stand deshalb mit
    14,7 Mio EUR in der Liste: derselbe Betrag von 2.293.202 lag dort viermal,
    1.277.564 dreimal.

    Gemessen an 581 GEWONNENEN Objekten, wo der tatsaechliche Auftragswert
    bekannt ist -- Verhaeltnis Formel zu Auftrag:
        Summe    Median 2,41x (28 % brauchbar)
        Maximum  Median 1,21x (82 %)
        primaere Median 1,01x (85 %)   <- praktisch unverzerrt
    """
    from adwatch.insights.projekte import _projekt_schaetzwert
    from adwatch.models import CrmOpportunity

    prim = CrmOpportunity(crm_id="1", opportunity_guid="P", project_id="P",
                          estimated_value=2_293_202)
    geschwister = [prim] + [
        CrmOpportunity(crm_id=str(i), opportunity_guid=f"G{i}", project_id="P",
                       estimated_value=v)
        for i, v in enumerate([2_293_202, 2_293_202, 1_277_564, 1_277_564], start=2)]

    wert = _projekt_schaetzwert(prim, geschwister)
    assert wert == 2_293_202, "der Wert der primaeren VC, nicht die Summe"
    assert wert != sum(float(m.estimated_value) for m in geschwister)

    # Rueckfall: traegt die primaere VC keinen Wert, gilt der groesste --
    # von den verbleibenden Regeln liegt er am wenigsten daneben.
    ohne = CrmOpportunity(crm_id="1", opportunity_guid="P", project_id="P",
                          estimated_value=None)
    assert _projekt_schaetzwert(ohne, [ohne] + geschwister[1:]) == 2_293_202

    # gar keine Werte -> None, nicht 0 (0 EUR und "unbekannt" sind verschieden)
    leer = CrmOpportunity(crm_id="1", opportunity_guid="P", project_id="P")
    assert _projekt_schaetzwert(leer, [leer]) is None

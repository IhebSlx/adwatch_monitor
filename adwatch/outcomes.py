"""Arbeitslisten mit Kontrollgruppe — und die Messung, die daraus folgt.

WARUM DAS DER WICHTIGSTE BAUSTEIN IST (gemessen, nicht vermutet):
Beschreibende Merkmale — was eine Firma IST — bringen +0,03 AUC; das
Anreicherungs-Experiment vom 2026-08-17 hat das gegen eine vorab festgelegte
Schwelle geprüft und verworfen. Verhaltensdaten — was zwischen uns und der Firma
GESCHEHEN ist — bringen +0,14 bis +0,16. Verhalten ist also das Einzige, was
diese Modelle wirklich bewegt, und dieses Modul ist der Weg, es zu erzeugen.

DIE KONTROLLGRUPPE IST DER EIGENTLICHE PUNKT.
Eine Liste abzuarbeiten und hinterher die Abschlüsse zu zählen, misst nichts:
11,3 % der deutschen Händler fragen auch ohne jeden Anruf an (Belege sind
Angebote, keine Rechnungen — Korrektur 2026-08-20). Wer nur die Ziele
zählt, schreibt sich die Basisrate als Erfolg gut.

Interessant ist nicht die Kaufwahrscheinlichkeit, sondern die REAKTION AUF
ANSPRACHE — bei Ascarza (2018) messbar besser als Neigungssteuerung, weil
Neigung Aufwand auf Firmen lenkt, die ohnehin gekauft hätten. Reaktion ist nur
gegen eine Kontrollgruppe messbar.

Deshalb: bei der Erzeugung einer Liste wird ein zufälliger Anteil als
`kontrolle` markiert und ausdrücklich NICHT angesprochen. Die Zuteilung erfolgt
EINMAL, mit festgehaltenem Startwert, und ist unveränderlich. Eine nachträglich
verschobene Kontrollgruppe ist keine.
"""
from __future__ import annotations

import datetime as dt
import random
import re

from sqlalchemy import func, select

from .db import SessionLocal
from .models import Company, CrmOrderEvent, TargetList, TargetListEntry

# Feste, kleine Ergebnis-Sprache. Freitext gehört in `note` — eine Auswertung
# über 40 selbst erfundene Status ist keine.
OUTCOMES: dict[str, str] = {
    "kein_kontakt":   "nicht erreicht",
    "kein_interesse": "kein Interesse",
    "spaeter":        "später wieder anfragen",
    "angebot":        "Angebot abgegeben",
    "gewonnen":       "gewonnen — Partner",
    "verloren":       "verloren",
}
CHANNELS: dict[str, str] = {
    "telefon": "Telefon", "mail": "E-Mail", "besuch": "Besuch",
    "messe": "Messe", "sonstiges": "Sonstiges",
}
DEFAULT_HOLDOUT = 0.15
MIN_ARM_FOR_CLAIM = 30      # unter dieser Größe ist ein Armvergleich Anekdote


def create_list(name: str, source: str, rows: list[dict], *,
                holdout_share: float = DEFAULT_HOLDOUT,
                filters: dict | None = None, seed: int | None = None,
                created_by: str | None = None, note: str | None = None) -> dict:
    """Liste anlegen und die Kontrollgruppe SOFORT ziehen.

    `rows` sind die gereihten Treffer eines Profils: [{company_id, score}, ...].
    Die Zuteilung geschieht hier und nur hier — würde sie beim ersten Anruf
    passieren, wäre sie durch die Reihenfolge der Bearbeitung verzerrt.

    Gezogen wird GESCHICHTET über die Rangfolge: aus je zehn aufeinander
    folgenden Rängen kommt derselbe Anteil in die Kontrolle. Eine rein zufällige
    Ziehung könnte sonst die halbe Spitze in die Kontrollgruppe legen, und der
    Vergleich verlöre genau dort an Aussagekraft, wo er zählt.
    """
    if not rows:
        raise ValueError("Leere Liste — nichts zu verteilen.")
    if not 0.0 <= holdout_share < 0.5:
        raise ValueError("holdout_share muss zwischen 0 und 0,5 liegen.")
    seed = int(seed if seed is not None else dt.datetime.utcnow().strftime("%Y%m%d"))
    rng = random.Random(seed)

    arms: list[str] = []
    for start in range(0, len(rows), 10):
        block = rows[start:start + 10]
        k = int(round(len(block) * holdout_share))
        idx = set(rng.sample(range(len(block)), k)) if k else set()
        arms.extend("kontrolle" if i in idx else "ziel" for i in range(len(block)))

    with SessionLocal() as s:
        tl = TargetList(name=name, source=source, filters=filters,
                        holdout_share=holdout_share, seed=seed,
                        created_by=created_by, note=note)
        s.add(tl)
        s.flush()
        for rank, (r, arm) in enumerate(zip(rows, arms), start=1):
            s.add(TargetListEntry(list_id=tl.id, company_id=int(r["company_id"]),
                                  rank=rank, score_at_creation=r.get("score"),
                                  arm=arm))
        s.commit()
        return {"id": tl.id, "name": tl.name, "source": tl.source, "seed": seed,
                "n": len(rows), "n_ziel": arms.count("ziel"),
                "n_kontrolle": arms.count("kontrolle")}


def record(entry_id: int, *, outcome: str | None = None,
           channel: str | None = None, note: str | None = None,
           contacted: bool = True) -> dict:
    """Ergebnis eintragen. Die Kontrollgruppe darf NICHT kontaktiert werden —
    ein Eintrag dort zerstört den Vergleich, für den die Liste angelegt wurde,
    und wird deshalb abgewiesen statt still gespeichert."""
    if outcome and outcome not in OUTCOMES:
        raise ValueError(f"Unbekanntes Ergebnis: {outcome!r}")
    if channel and channel not in CHANNELS:
        raise ValueError(f"Unbekannter Kanal: {channel!r}")
    now = dt.datetime.utcnow()
    with SessionLocal() as s:
        e = s.get(TargetListEntry, entry_id)
        if not e:
            raise ValueError("Eintrag nicht gefunden")
        if e.arm == "kontrolle" and (contacted or channel):
            raise ValueError(
                "Diese Firma ist in der Kontrollgruppe und darf nicht "
                "angesprochen werden — sonst ist die Wirkung der Liste nicht "
                "mehr messbar.")
        if contacted and not e.contacted_at:
            e.contacted_at = now
        if channel:
            e.channel = channel
        if outcome:
            e.outcome = outcome
            e.outcome_at = now
        if note is not None:
            e.note = note
        s.commit()
        return {"id": e.id, "company_id": e.company_id, "arm": e.arm,
                "outcome": e.outcome, "channel": e.channel,
                "contacted_at": e.contacted_at.isoformat() if e.contacted_at else None}


def measure(list_id: int, since: dt.date | None = None) -> dict:
    """Die Wirkung der Liste: Kaufquote Ziel gegen Kontrolle.

    Gemessen wird am harten Ausgang — hat die Firma seit dem Anlegen der Liste
    tatsächlich bestellt (Betrag > 0, siehe die 0-Euro-Regel) —, NICHT am
    eingetragenen Ergebnis. Sonst misst man die Sorgfalt der Erfassung statt den
    Geschäftserfolg.

    `uplift` ist die Differenz der beiden Quoten: der Teil, den die Ansprache
    bewirkt hat, und nicht der Teil, der ohnehin passiert wäre.
    """
    with SessionLocal() as s:
        tl = s.get(TargetList, list_id)
        if not tl:
            raise ValueError("Liste nicht gefunden")
        entries_ = list(s.scalars(select(TargetListEntry)
                                  .where(TargetListEntry.list_id == list_id)))
        start = since or tl.created_at.date()
        ids = [e.company_id for e in entries_]
        bought: set[int] = set()
        if ids:
            bought = set(s.scalars(
                select(CrmOrderEvent.company_id).distinct()
                .where(CrmOrderEvent.company_id.in_(ids),
                       CrmOrderEvent.order_date >= start,
                       CrmOrderEvent.amount > 0)))

    def arm_stats(arm: str) -> dict:
        rows = [e for e in entries_ if e.arm == arm]
        n = len(rows)
        buyers = sum(1 for e in rows if e.company_id in bought)
        return {"n": n, "kaeufer": buyers, "quote": (buyers / n) if n else 0.0,
                "kontaktiert": sum(1 for e in rows if e.contacted_at)}

    ziel, kontrolle = arm_stats("ziel"), arm_stats("kontrolle")
    solid = kontrolle["n"] >= MIN_ARM_FOR_CLAIM and ziel["n"] >= MIN_ARM_FOR_CLAIM
    # Ohne Kontrollgruppe ist der Vergleich nicht definiert. Dann None statt 0 —
    # eine 0 sähe aus wie "gemessen, keine Wirkung", und das wäre gelogen.
    return {
        "list_id": list_id, "name": tl.name, "source": tl.source,
        "created_at": tl.created_at.isoformat(), "seit": str(start),
        "ziel": ziel, "kontrolle": kontrolle,
        "uplift": (ziel["quote"] - kontrolle["quote"]) if kontrolle["n"] else None,
        "aussagekraeftig": bool(solid),
        "hinweis": ("Ziel gegen Kontrolle, gleiche Ziehung, gleicher Zeitraum."
                    if solid else
                    f"Arme unter {MIN_ARM_FOR_CLAIM} Firmen — die Zahlen sind "
                    "beschreibend, nicht beweisend."),
    }


def list_lists() -> list[dict]:
    with SessionLocal() as s:
        out = []
        for tl in s.scalars(select(TargetList).order_by(TargetList.id.desc())):
            n = s.scalar(select(func.count(TargetListEntry.id))
                         .where(TargetListEntry.list_id == tl.id)) or 0
            done = s.scalar(select(func.count(TargetListEntry.id))
                            .where(TargetListEntry.list_id == tl.id,
                                   TargetListEntry.outcome.is_not(None))) or 0
            out.append({"id": tl.id, "name": tl.name, "source": tl.source,
                        "created_at": tl.created_at.isoformat(), "n": n,
                        "entschieden": done,
                        # Anteil und Startwert der Ziehung gehören zur Liste
                        # dazu: ohne sie ist im Export nicht nachvollziehbar,
                        # WIE die Kontrollgruppe zustande kam — und eine
                        # Ziehung, die man nicht nachrechnen kann, ist als
                        # Beleg wertlos.
                        "holdout_share": tl.holdout_share, "seed": tl.seed,
                        "filters": tl.filters})
        return out


def entries(list_id: int, arm: str | None = None, open_only: bool = False) -> list[dict]:
    """Die Zeilen einer Liste. Die Kontrollgruppe wird MITGELIEFERT und als
    solche markiert — sie zu verstecken wäre bequemer und würde dazu führen,
    dass jemand sie auf einem anderen Weg doch anruft."""
    with SessionLocal() as s:
        stmt = (select(TargetListEntry, Company)
                .join(Company, Company.id == TargetListEntry.company_id)
                .where(TargetListEntry.list_id == list_id)
                .order_by(TargetListEntry.rank))
        if arm:
            stmt = stmt.where(TargetListEntry.arm == arm)
        if open_only:
            stmt = stmt.where(TargetListEntry.outcome.is_(None))
        return [{"entry_id": e.id, "company_id": c.id, "name": c.name,
                 "city": c.city, "country": c.country, "sub_segment": c.sub_segment,
                 "rank": e.rank, "score": e.score_at_creation, "arm": e.arm,
                 "contacted_at": e.contacted_at.isoformat() if e.contacted_at else None,
                 "channel": e.channel, "outcome": e.outcome, "note": e.note}
                for e, c in s.execute(stmt)]


# --- Export ----------------------------------------------------------------
# Warum es diesen Export gibt, obwohl alles in der App steht:
#
# E-Mails, Firmen und Verkaufschancen sind ein SPIEGEL des CRM — brennt die
# Datenbank ab, sind sie in ein paar Stunden wieder da. Listen, Arme und
# Ergebnisse sind das nicht. Sie sind das einzige Exemplar der teuersten Daten
# im System: menschliche Vertriebsentscheidungen, die niemand nachbauen kann.
# Sieben rotierende Tagessicherungen sind dafür dünn.
#
# Der zweite Grund ist praktischer: eine Liste, die nur in AdWatch existiert,
# muss in AdWatch abgearbeitet werden. Als Datei kann eine Kollegin sie
# annehmen, ohne sich anzumelden.

_BLATT_VERBOTEN = str.maketrans({c: "-" for c in '[]:*?/\\'})

_EXPORT_KOPF = [
    ("rank", "Rang"), ("name", "Firma"), ("city", "Ort"), ("country", "Land"),
    ("sub_segment", "Untersegment"), ("score", "Punktzahl bei Erstellung"),
    ("arm", "Arm"), ("anrufen", "Anrufen?"),
    ("contacted_at", "Kontaktiert am"), ("channel", "Kanal"),
    ("outcome", "Ergebnis"), ("note", "Notiz"), ("company_id", "Firmen-ID"),
    ("entry_id", "Zeilen-ID"),
]


def export_xlsx(list_id: int | None = None) -> bytes:
    """Listen, Arme und Ergebnisse als .xlsx — eine Mappe, je Liste ein Blatt.

    Die Kontrollgruppe reist MIT und trägt in einer eigenen Spalte ein klares
    NEIN. Sie wegzulassen wäre die bequeme Lösung und die falsche: wer die
    Datei bekommt und die Kontrollgruppe nicht sieht, ruft sie irgendwann über
    einen anderen Weg doch an — und dann ist die einzige Messung hin, die den
    ganzen Versuch falsifizierbar macht.
    """
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    listen = [l for l in list_lists() if list_id is None or l["id"] == list_id]
    if not listen:
        raise ValueError("Keine Liste zum Exportieren")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    kopf_stil = Font(bold=True, color="FFFFFF")
    kopf_fuell = PatternFill("solid", fgColor="4F5CE5")
    warn_fuell = PatternFill("solid", fgColor="FDE7E7")

    for liste in listen:
        # Excel erlaubt 31 Zeichen je Blattname und verbietet sieben davon.
        # Eine Übersetzungstabelle statt eines regulären Ausdrucks: die Zeichen
        # stehen dann lesbar da, statt in einer Zeichenklasse voller
        # Maskierungen zu verschwinden.
        titel = (f"{liste['id']} {liste['name']}".translate(_BLATT_VERBOTEN))[:31]
        ws = wb.create_sheet(titel)
        ws.append([label for _, label in _EXPORT_KOPF])
        for zelle in ws[1]:
            zelle.font = kopf_stil
            zelle.fill = kopf_fuell
            zelle.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

        for e in entries(liste["id"]):
            kontrolle = e["arm"] == "kontrolle"
            ws.append([
                e["rank"], e["name"], e["city"], e["country"], e["sub_segment"],
                round(e["score"], 1) if e["score"] is not None else None,
                "Kontrollgruppe" if kontrolle else "Zielgruppe",
                "NEIN — nicht ansprechen" if kontrolle else "ja",
                (e["contacted_at"] or "")[:10] or None,
                CHANNELS.get(e["channel"] or "", e["channel"]),
                OUTCOMES.get(e["outcome"] or "", e["outcome"]),
                e["note"], e["company_id"], e["entry_id"],
            ])
            if kontrolle:
                for zelle in ws[ws.max_row]:
                    zelle.fill = warn_fuell

        for spalte, breite in zip("ABCDEFGHIJKLMN",
                                  (6, 38, 18, 6, 20, 10, 15, 22, 14, 12, 20, 40, 11, 10)):
            ws.column_dimensions[spalte].width = breite

        # Was die Liste ist, gehört in die Datei — ohne den Kopf ist sie in
        # einem Monat eine Tabelle ohne Herkunft.
        ws.append([])
        ws.append(["Liste", liste["name"]])
        ws.append(["Quelle", liste.get("source")])
        ws.append(["Angelegt", (liste.get("created_at") or "")[:19]])
        ws.append(["Kontrollanteil", liste.get("holdout_share")])
        ws.append(["Ziehung (seed)", liste.get("seed")])
        ws.append(["Hinweis", "Die rot hinterlegten Zeilen sind die Kontrollgruppe. "
                              "Sie werden bewusst NICHT angesprochen — sonst ist die "
                              "Wirkung der Ansprache nicht mehr messbar."])

    puffer = io.BytesIO()
    wb.save(puffer)
    return puffer.getvalue()

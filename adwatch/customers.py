"""Company master data: Excel import + filtered explorer + export.

A company IS a customer — there is no separate entity (see models.Company
docstring). This module owns the bulk operations (import thousands of rows,
filter/sort/paginate/select across them, export); per-row identity work
(linking a Meta/Google page, confirming a match) lives in identity.resolver
and services.py, same as before the merge. A future direct-DB feed would
reuse `upsert_companies()` unchanged — the Excel reader is just one way to
produce the record dicts it takes.

Fetching is opt-in per row simply by being fetched — there's no separate
"promote to tracking" step. A company with resolution_status='pending' has
just never been fetched yet; the first fetch (via the normal pipeline)
resolves its page(s) and the row updates in place.
"""
from __future__ import annotations

import datetime as dt
import io
import re

from sqlalchemy import case, func, or_, select
from sqlalchemy.exc import IntegrityError

from . import config, markets, scope
from .db import SessionLocal
from .models import Company

# ---------------------------------------------------------------------------
# Column mapping — tolerant to whitespace/case so slightly different exports
# still line up. Each model field lists the header variants that map to it.
# ---------------------------------------------------------------------------
_FIELD_HEADERS: dict[str, tuple[str, ...]] = {
    # The Dataverse accountid, shipped in every CRM export as "(Nicht ändern) Firma".
    # Captured so the durable key lands in the DB on any import, with no CRM call.
    "crm_id": ("(nicht ändern) firma", "(nicht andern) firma", "accountid", "crm id", "crm-id"),
    "crm_modified_on": ("(nicht ändern) geändert am", "(nicht andern) geandert am", "modifiedon"),
    "sap_number": ("sap nummer", "sap-nummer", "sapnummer", "sap"),
    "name": ("firmenname", "firma", "name"),
    "kv": ("kv",),
    "segment": ("kundensegment",),
    "sub_segment": ("kundenuntersegment",),
    "sales_channel": ("vertriebsweg",),
    "street": ("straße", "strasse"),
    "postal_code": ("adresse 1: postleitzahl", "postleitzahl", "plz"),
    "city": ("ort",),
    "country": ("land",),
    "phone": ("telefon 1", "telefon", "tel"),
    "email": ("e-mail", "email", "e mail"),
    "fax": ("fax",),
    "website_domain": ("website", "webseite", "web"),
    "revenue_y0": ("umsatz aktuelles jahr",),
    "revenue_y1": ("umsatz aktuelles jahr -1", "umsatz aktuelles jahr-1"),
    "revenue_y2": ("umsatz aktuelles jahr -2", "umsatz aktuelles jahr-2"),
    "revenue_y3": ("umsatz aktuelles jahr -3", "umsatz aktuelles jahr-3"),
    "revenue_y4": ("umsatz aktuelles jahr -4", "umsatz aktuelles jahr-4"),
}
_REVENUE_FIELDS = ("revenue_y0", "revenue_y1", "revenue_y2", "revenue_y3", "revenue_y4")
# Sort keys the UI is allowed to request -> model column.
SORTABLE = {
    "name": Company.name, "sap_number": Company.sap_number,
    "segment": Company.segment, "sub_segment": Company.sub_segment,
    "sales_channel": Company.sales_channel, "city": Company.city, "country": Company.country,
    "revenue_y0": Company.revenue_y0, "revenue_y1": Company.revenue_y1,
    "revenue_y2": Company.revenue_y2, "revenue_y3": Company.revenue_y3,
    "revenue_y4": Company.revenue_y4, "imported_at": Company.imported_at,
    "resolution_status": Company.resolution_status,
    "customer_state": Company.customer_state, "fit_score": Company.fit_score,
    "opportunity_score": Company.opportunity_score, "target_score": Company.target_score,
    "solarlux_relevance": Company.solarlux_relevance,
}


# Own-group entities that appear in the CRM as ordinary customers. Confirmed
# against the CRM: Linara entities, NanaWall Systems and Solarlux Vertriebsbüros
# show up with large revenue — Linara Kaufbeuren alone was the single biggest
# "customer" (EUR 1.35M) and ranked #3 in the target list before this flag.
# Matched case-insensitively as a whole word / prefix of the company name.
INTERCOMPANY_NAME_PATTERNS = (
    "linara", "nana wall", "nanawall", "solarlux", "slect",
)


def looks_intercompany(name: str | None) -> bool:
    n = _norm(name)
    return bool(n) and any(p in n for p in INTERCOMPANY_NAME_PATTERNS)


def flag_intercompany() -> int:
    """(Re)set Company.is_intercompany from the name patterns. Idempotent, and
    run on every import so newly synced group companies are caught too."""
    with SessionLocal() as s:
        n = 0
        for c in s.scalars(select(Company)):
            flag = looks_intercompany(c.name)
            if bool(c.is_intercompany) != flag:
                c.is_intercompany = flag
                n += 1
        s.commit()
        return n


def derive_customer_state(y0, y1, y2, y3, y4) -> str:
    """Customer lifecycle from the imported Umsatz columns (NULL counts as €0):
    active = buys now AND bought before | new = first revenue this year |
    lapsed = bought before, nothing this year | never = no revenue on record."""
    now = (y0 or 0) > 0
    before = any((v or 0) > 0 for v in (y1, y2, y3, y4))
    if now and before:
        return "active"
    if now:
        return "new"
    return "lapsed" if before else "never"


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _parse_number(value) -> float | None:
    """Parse a revenue cell that may already be numeric, or a German-formatted
    string ('1.234.567,89' -> 1234567.89). Returns None for blanks/garbage."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[^0-9,.\-]", "", str(value))  # drop currency symbols, spaces, etc.
    if not s or s in ("-", ".", ","):
        return None
    has_dot, has_comma = "." in s, "," in s
    if has_dot and has_comma:
        s = s.replace(".", "").replace(",", ".")          # dot=thousands, comma=decimal
    elif has_comma:
        s = s.replace(",", ".")                            # comma=decimal
    elif has_dot:
        # Lone dot: thousands unless it clearly reads as a decimal (e.g. '1.5').
        tail = s.split(".")[-1]
        if s.count(".") > 1 or len(tail) == 3:
            s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def _clean_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_dt(value):
    """CRM `modifiedon` — already a datetime from openpyxl, or an ISO string."""
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value
    try:
        return dt.datetime.fromisoformat(str(value).strip().replace("Z", "+00:00").replace(" ", "T", 1))
    except ValueError:
        return None


# The CRM's Land column holds full names ("Spanien", "Deutschland"), but every
# downstream consumer needs an ISO-2 code: Company.country is what the Meta and
# Google ad lookups pass as their country parameter, so a Spanish company left
# at the "DE" default would be searched in the GERMAN ad library.


def _country_code(value) -> str | None:
    """'Spanien' -> 'ES', 'ES' -> 'ES', unknown full name -> None (keep the
    existing value rather than guess).

    The alias table lives in config/markets.yaml so a new market needs no code
    change — see adwatch/markets.py."""
    return markets.code_for(value)


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Read the first worksheet into record dicts. Returns (records, warnings).
    Raises ValueError if the SAP-Nummer column can't be found at all."""
    import openpyxl  # local import: only needed at upload time
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("The uploaded file is empty.")

    # Map each column index to a model field via the normalized header.
    norm_to_field = {}
    for field, variants in _FIELD_HEADERS.items():
        norm_to_field.update({v: field for v in variants})
    col_field: dict[int, str] = {}
    for idx, cell in enumerate(header):
        field = norm_to_field.get(_norm(cell))
        if field and field not in col_field.values():
            col_field[idx] = field

    if "sap_number" not in col_field.values():
        raise ValueError("Could not find a 'SAP Nummer' column in the file's header row.")

    records, warnings = [], []
    no_sap = no_key = 0
    for line_no, row in enumerate(rows, start=2):
        rec: dict = {}
        for idx, field in col_field.items():
            value = row[idx] if idx < len(row) else None
            rec[field] = _parse_number(value) if field in _REVENUE_FIELDS else _clean_str(value)
        # A row needs SOME key to be identifiable: the SAP number, or failing
        # that its name (upsert_companies matches those by exact name). Rows with
        # neither are blank/trailing rows and are dropped — but the count is
        # REPORTED, never silent: a CRM view can legitimately hold thousands of
        # rows without a SAP number, and quietly discarding them once looked like
        # "the import worked" while 90% of the file vanished.
        if not rec.get("sap_number"):
            if not rec.get("name"):
                no_key += 1
                continue
            no_sap += 1
        records.append(rec)

    missing = [f for f in _FIELD_HEADERS if f not in col_field.values()]
    if missing:
        warnings.append(f"{len(missing)} column(s) not found and left blank: {', '.join(missing)}")
    if no_sap:
        warnings.append(f"{no_sap} row(s) have no SAP Nummer — matched by Firmenname instead")
    if no_key:
        warnings.append(f"{no_key} row(s) skipped: neither SAP Nummer nor Firmenname")
    return records, warnings


def upsert_companies(records: list[dict]) -> dict:
    """Insert new companies / update existing ones, matched on sap_number.
    A record whose sap_number isn't known yet but whose NAME exactly matches
    an existing (e.g. hand-added, pre-Excel) company folds into that row
    instead of erroring on the name-uniqueness constraint. Never overwrites
    `name` on an existing row (renaming could break Meta page resolution) or
    a `website_domain`/`country` that's already set. Ad-tracking fields
    (page_id, resolution_status, ...) are never touched by import."""
    now = dt.datetime.utcnow()
    inserted = updated = skipped_no_name = name_collisions = 0
    with SessionLocal() as s:
        by_crm = {c.crm_id: c for c in
                  s.scalars(select(Company).where(Company.crm_id.is_not(None)))}
        by_sap = {c.sap_number: c for c in
                  s.scalars(select(Company).where(Company.sap_number.is_not(None)))}
        by_name = {c.name: c for c in s.scalars(select(Company))}

        for rec in records:
            sap = rec.get("sap_number")
            crm = (rec.get("crm_id") or "").strip() or None
            name = (rec.get("name") or "").strip()
            # crm_id first: it is the only key that never changes. SAP number and
            # exact name remain as fallbacks for rows imported before the GUID
            # was captured (and for hand-added companies).
            row = by_crm.get(crm) if crm else None
            if row is None and sap:
                row = by_sap.get(sap)
            if row is None and name:
                row = by_name.get(name)   # fold into a pre-existing un-SAP'd row by exact name

            if row is None:
                if not name:
                    skipped_no_name += 1
                    continue
                row = Company(name=name, source="meta", resolution_status="pending")
                s.add(row)
                try:
                    s.flush()
                except IntegrityError:
                    # Genuine duplicate company name across two different SAP
                    # numbers (plausible at this scale) — disambiguate rather
                    # than dropping the row or crashing the whole import.
                    s.rollback()
                    by_sap = {c.sap_number: c for c in
                              s.scalars(select(Company).where(Company.sap_number.is_not(None)))}
                    by_name = {c.name: c for c in s.scalars(select(Company))}
                    disambiguated = f"{name} ({sap})"
                    row = Company(name=disambiguated, source="meta", resolution_status="pending")
                    s.add(row)
                    s.flush()
                    name_collisions += 1
                if sap:
                    by_sap[sap] = row
                by_name[row.name] = row
                inserted += 1
            else:
                updated += 1

            for field, value in rec.items():
                if field in ("sap_number", "name"):
                    if field == "sap_number" and value and row.sap_number != value:
                        row.sap_number = value
                    continue
                if field == "crm_id":
                    # write-once: the GUID is immutable, and overwriting it would
                    # silently repoint a row at a different CRM account
                    if value and not row.crm_id:
                        row.crm_id = str(value).strip()
                    continue
                if field == "crm_modified_on":
                    row.crm_modified_on = _parse_dt(value)
                    continue
                if field == "website_domain":
                    if value and not row.website_domain:
                        row.website_domain = value
                    continue
                if field == "country":
                    code = _country_code(value)
                    if code:
                        row.country = code
                    continue
                setattr(row, field, value)
            row.imported_at = now
            # lifecycle is pure derived data — recomputed on every import so a
            # partner whose revenue reappears flips lapsed -> active by itself
            row.customer_state = derive_customer_state(
                row.revenue_y0, row.revenue_y1, row.revenue_y2, row.revenue_y3, row.revenue_y4)
            row.is_intercompany = looks_intercompany(row.name)
        s.commit()
    return {"received": len(records), "inserted": inserted, "updated": updated,
           "skipped_no_name": skipped_no_name, "name_collisions": name_collisions}


def import_excel(file_bytes: bytes) -> dict:
    records, warnings = parse_excel(file_bytes)
    summary = upsert_companies(records)
    summary["warnings"] = warnings
    return summary


# ---------------------------------------------------------------------------
# Querying / filtering
# ---------------------------------------------------------------------------

def _apply_filters(stmt, f: dict):
    # Consumers are out of scope for everything (see adwatch/scope.py) — applied
    # FIRST and unconditionally, so no filter combination, hand-picked id list or
    # forgotten frontend default can put them back into a count, export or report.
    # `include_consumers` is the one deliberate way in and nothing sets it by
    # default; asking for the segment by name also still works, so a consumer
    # record can be looked up on purpose.
    asked_for_consumers = any(
        s in (f.get("segment") or []) for s in scope.EXCLUDED_SEGMENTS)
    if not (f.get("include_consumers") or asked_for_consumers):
        stmt = scope.apply(stmt)
    # Explicit selection (hand-picked company ids, e.g. "report for selected"):
    # restrict to exactly these. Kept first so it composes with any other
    # filters, though the selection flow normally passes ids alone.
    if f.get("ids"):
        stmt = stmt.where(Company.id.in_([int(i) for i in f["ids"]]))
    if f.get("q"):
        like = f"%{f['q'].strip()}%"
        stmt = stmt.where(or_(Company.name.ilike(like), Company.sap_number.ilike(like)))
    for field, col in (("kv", Company.kv), ("segment", Company.segment),
                       ("sub_segment", Company.sub_segment),
                       ("sales_channel", Company.sales_channel), ("country", Company.country),
                       ("lead_source", Company.lead_source),
                       ("solarlux_relevance", Company.solarlux_relevance),
                       ("office_type", Company.office_type),
                       ("decision_role", Company.decision_role)):
        values = f.get(field)
        if values:
            stmt = stmt.where(col.in_(values) if isinstance(values, list) else col == values)
    for field, col in (("exclude_kv", Company.kv), ("exclude_segment", Company.segment),
                       ("exclude_sub_segment", Company.sub_segment)):
        values = f.get(field)
        if values:
            # NULL-safe: a company with no value in this column isn't "one of
            # the excluded ones", so it stays in the result (unlike a bare
            # NOT IN, which SQL silently drops NULLs from).
            stmt = stmt.where(or_(col.is_(None), col.not_in(values)))
    if f.get("resolution_status"):
        values = f["resolution_status"]
        stmt = stmt.where(Company.resolution_status.in_(values)
                          if isinstance(values, list) else Company.resolution_status == values)
    if f.get("has_website"):
        stmt = stmt.where(Company.website_domain.is_not(None), Company.website_domain != "")
    if f.get("no_website"):
        # the enrichment target set: rows the Explorer can hand straight to an
        # enrich job to have their website found
        stmt = stmt.where(or_(Company.website_domain.is_(None), Company.website_domain == ""))
    if f.get("enrichment_status"):
        values = f["enrichment_status"]
        stmt = stmt.where(Company.enrichment_status.in_(values) if isinstance(values, list)
                          else Company.enrichment_status == values)
    if f.get("customer_state"):
        values = f["customer_state"]
        stmt = stmt.where(Company.customer_state.in_(values) if isinstance(values, list)
                          else Company.customer_state == values)
    if f.get("fit_min") is not None:
        stmt = stmt.where(Company.fit_score >= float(f["fit_min"]))
    # Identity fetch-readiness: a numeric page_id is what the Ad Library scrape
    # needs — "without" surfaces the ⚠ no-id rows (confirmed but not fetch-ready).
    if f.get("page_id_state") == "with":
        stmt = stmt.where(Company.page_id.is_not(None), Company.page_id != "")
    elif f.get("page_id_state") == "without":
        stmt = stmt.where(or_(Company.page_id.is_(None), Company.page_id == ""))
    if f.get("tracked") is not None:
        stmt = stmt.where(Company.resolution_status != "pending") if f["tracked"] \
            else stmt.where(Company.resolution_status == "pending")
    # Ad activity — based on the LATEST fetched week:
    #   active: latest week sums to >=1 active ad (the win-back signal)
    #   any:    any real ad ever on record (active or ended)
    #   none:   fetched at least once but no active ad in the latest week
    # Companies never fetched have no metric, so they match none of these.
    # `ad_source` ("meta"/"google") narrows every part to one platform: the
    # "latest week", the active-ad sum, and the ever-advertised check all look
    # only at that source's rows — so active+meta = "running Meta ads now"
    # irrespective of Google, and none+google = "fetched on Google but no live
    # Google ad". No source (or an unknown one) keeps the all-platforms totals.
    aa = f.get("ad_activity")
    if aa in ("active", "any", "none"):
        from sqlalchemy import and_ as _and
        from .models import Ad as _Ad, CollectionRun as _CR, WeeklyCompanyMetric as _WCM
        src = f.get("ad_source") if f.get("ad_source") in ("meta", "google") else None
        _wcm_src = (_WCM.source == src,) if src else ()   # empty -> no-op .where(), all sources
        _cr_src = (_CR.source == src,) if src else ()
        _latest = (select(_WCM.company_id, func.max(_WCM.week_start).label("wk"))
                   .where(*_wcm_src)
                   .group_by(_WCM.company_id).subquery())
        _running = (select(_WCM.company_id)
                    .join(_latest, _and(_WCM.company_id == _latest.c.company_id,
                                        _WCM.week_start == _latest.c.wk))
                    .where(*_wcm_src)
                    .group_by(_WCM.company_id)
                    .having(func.sum(_WCM.total_active_ads) > 0))
        if aa == "active":
            stmt = stmt.where(Company.id.in_(_running))
        elif aa == "any":
            # ever advertised: a real ad on record (incl. ended) OR any week with
            # active ads. Union guarantees "any" is a superset of "active".
            _real_ad = (select(_CR.company_id).join(_Ad, _Ad.run_id == _CR.id)
                        .where(_Ad.external_ad_id.is_not(None), *_cr_src))
            _ever_active = select(_WCM.company_id).where(_WCM.total_active_ads > 0, *_wcm_src)
            stmt = stmt.where(or_(Company.id.in_(_real_ad), Company.id.in_(_ever_active)))
        else:  # none — fetched but not currently advertising
            stmt = stmt.where(Company.id.in_(select(_WCM.company_id).where(*_wcm_src)),
                              Company.id.not_in(_running))
    if f.get("revenue_min") is not None:
        stmt = stmt.where(func.coalesce(Company.revenue_y0, 0) >= f["revenue_min"])
    if f.get("revenue_max") is not None:
        stmt = stmt.where(func.coalesce(Company.revenue_y0, 0) <= f["revenue_max"])
    if f.get("revenue_history"):
        y0 = func.coalesce(Company.revenue_y0, 0)
        prior_any = or_(*(func.coalesce(getattr(Company, f"revenue_y{i}"), 0) > 0 for i in (1, 2, 3, 4)))
        hist = f["revenue_history"]
        if hist == "lapsed":       # bought in a prior year, nothing this year — win-back candidates
            stmt = stmt.where(y0 <= 0, prior_any)
        elif hist == "new":        # buying this year, no history before it
            stmt = stmt.where(y0 > 0, ~prior_any)
        elif hist == "any":        # revenue in at least one year, current or past
            stmt = stmt.where(or_(y0 > 0, prior_any))
        elif hist == "never":      # no revenue recorded in any year
            stmt = stmt.where(y0 <= 0, ~prior_any)
    return stmt


_REVENUE_SORT_KEYS = {"revenue_y0", "revenue_y1", "revenue_y2", "revenue_y3", "revenue_y4"}

# Ordinal text columns. Sorting these as text is not just ugly, it is WRONG:
# alphabetically "gering" beats "hoch" beats "mittel", so a "best first" sort on
# Solarlux-Relevanz would put the worst-fitting architects on top. Ranked here so
# desc = best first, and NULL keeps sorting last via the existing is_(None) rule.
_ORDINAL_SORTS = {
    "solarlux_relevance": {"hoch": 3, "mittel": 2, "gering": 1},
}


def _apply_sort(stmt, sort: str | None, direction: str):
    key = sort or "name"
    col = SORTABLE.get(key, Company.name)
    if key in _ORDINAL_SORTS:
        stmt = stmt.order_by(col.is_(None))
        rank = case(_ORDINAL_SORTS[key], value=col, else_=0)
        return stmt.order_by(rank.desc() if direction == "desc" else rank.asc())
    if key in _REVENUE_SORT_KEYS:
        # Missing revenue counts as €0 here too, so it sorts in place rather
        # than always trailing behind real numbers.
        col = func.coalesce(col, 0)
    else:
        # NULLs last regardless of direction, so blanks never top a desc sort.
        stmt = stmt.order_by(col.is_(None))
    return stmt.order_by(col.desc() if direction == "desc" else col.asc())


def _to_dict(c: Company) -> dict:
    return {
        "id": c.id, "sap_number": c.sap_number, "name": c.name,
        "kv": c.kv, "segment": c.segment, "sub_segment": c.sub_segment,
        "sales_channel": c.sales_channel, "street": c.street, "postal_code": c.postal_code,
        "city": c.city, "country": c.country, "phone": c.phone, "email": c.email,
        "fax": c.fax, "website_domain": c.website_domain,
        "revenue_y0": c.revenue_y0, "revenue_y1": c.revenue_y1, "revenue_y2": c.revenue_y2,
        "revenue_y3": c.revenue_y3, "revenue_y4": c.revenue_y4,
        "resolution_status": c.resolution_status, "tracked": c.resolution_status != "pending",
        "page_id": c.page_id, "page_name": c.page_name, "page_url": c.page_url,
        "customer_state": c.customer_state,
        "fit_score": c.fit_score, "opportunity_score": c.opportunity_score,
        "target_score": c.target_score,
        "description": c.description, "products": c.products,
        "founded_year": c.founded_year, "employee_hint": c.employee_hint,
        "enrichment_status": c.enrichment_status,
        # ---- everything below arrived with the CRM/Beleg/enrichment expansion
        # and was invisible in the UI until Iheb noticed ("I am not seeing all
        # the columns"). A column we hold but never show may as well not exist.
        "crm_id": c.crm_id,
        "lead_source": c.lead_source, "import_type": c.import_type,
        "is_competitor": c.is_competitor, "monitored": c.monitored,
        # Belege (authoritative revenue) + customer health
        "beleg_count": c.beleg_count, "beleg_sum": c.beleg_sum,
        "beleg_first": c.beleg_first.isoformat() if c.beleg_first else None,
        "beleg_last": c.beleg_last.isoformat() if c.beleg_last else None,
        "beleg_by_year": c.beleg_by_year, "avg_discount": c.avg_discount,
        "health": c.health, "winback_score": c.winback_score,
        # Angebote
        "quote_count": c.quote_count, "quote_sum": c.quote_sum,
        "conversion_rate": c.conversion_rate,
        # Architekten-Einfluss
        "arch_projects": c.arch_projects, "arch_won": c.arch_won,
        "arch_won_value": c.arch_won_value,
        # Website-Identität
        "identity_status": c.identity_status,
        "identity_matched_by": c.identity_matched_by,
        "website_source": c.website_source,
        # Anreicherung, Teil 2
        "legal_form": c.legal_form, "service_area": c.service_area,
        "competitor_brands": c.competitor_brands,
        "mentions_solarlux": c.mentions_solarlux,
        "assessment": c.assessment, "certifications": c.certifications,
        "own_fabrication": c.own_fabrication, "has_showroom": c.has_showroom,
        "project_focus": c.project_focus, "positioning": c.positioning,
        # Architekten-Profil (enrich profile 'architekt')
        "solarlux_relevance": c.solarlux_relevance, "office_type": c.office_type,
        "decision_role": c.decision_role, "reference_scale": c.reference_scale,
        "enrich_profile": c.enrich_profile,
        "facebook_url": c.facebook_url, "instagram_url": c.instagram_url,
        "linkedin_url": c.linkedin_url, "site_language": c.site_language,
    }


def get_company(company_id: int) -> dict | None:
    """One company's full master-data row — the company drawer's data source
    when the company isn't in the Explorer's currently loaded page (e.g. the
    drawer opened from the dashboard or the divergence list). Includes the
    identity `candidates` blob (only carried on the single-company fetch, not
    the list, to keep list payloads lean) so the drawer can offer a pick-list."""
    with SessionLocal() as s:
        c = s.get(Company, company_id)
        if not c:
            return None
        d = _to_dict(c)
        d["candidates"] = c.candidates or []
        d["fit_breakdown"] = c.fit_breakdown   # per-feature 'Warum' for the drawer (single-fetch only)
        return d


def _attach_active_ads(s, rows: list[dict]) -> None:
    """Fill each row's `active_ads` = active-ad count from its LATEST fetched
    week (summed across sources) in ONE query for the whole page. None means the
    company has never been fetched (shown as '—', distinct from a real 0)."""
    ids = [r["id"] for r in rows]
    if not ids:
        return
    from sqlalchemy import and_ as _and
    from .models import WeeklyCompanyMetric as _WCM
    latest = (select(_WCM.company_id, func.max(_WCM.week_start).label("wk"))
              .where(_WCM.company_id.in_(ids)).group_by(_WCM.company_id).subquery())
    q = (select(_WCM.company_id, func.sum(_WCM.total_active_ads))
         .join(latest, _and(_WCM.company_id == latest.c.company_id,
                            _WCM.week_start == latest.c.wk))
         .group_by(_WCM.company_id))
    counts = {cid: int(n or 0) for cid, n in s.execute(q)}
    for r in rows:
        r["active_ads"] = counts.get(r["id"])


def query_companies(filters: dict, sort: str | None = None, direction: str = "asc",
                    page: int = 1, page_size: int = 50) -> dict:
    with SessionLocal() as s:
        base = _apply_filters(select(Company), filters)
        total = s.scalar(select(func.count()).select_from(base.subquery()))
        stmt = _apply_sort(base, sort, direction).limit(page_size).offset((page - 1) * page_size)
        rows = [_to_dict(c) for c in s.scalars(stmt)]
        _attach_active_ads(s, rows)
        return {"total": total, "page": page, "page_size": page_size, "rows": rows}


def top_ids(filters: dict, sort: str | None, direction: str, n: int) -> list[int]:
    """The first N company ids under the current filter+sort — powers the
    'select top N' action so a subset is chosen across the WHOLE sorted set,
    not just the visible page."""
    with SessionLocal() as s:
        stmt = _apply_sort(_apply_filters(select(Company.id), filters), sort, direction).limit(n)
        return list(s.scalars(stmt))


def filter_options() -> dict:
    """Distinct values for the filter dropdowns."""
    with SessionLocal() as s:
        def distinct(col):
            return sorted(v for v in s.scalars(select(col).distinct()) if v)
        return {
            "kv": distinct(Company.kv),
            "segment": distinct(Company.segment),
            "sub_segment": distinct(Company.sub_segment),
            "sales_channel": distinct(Company.sales_channel),
            "country": distinct(Company.country),
        }


def count_companies() -> int:
    with SessionLocal() as s:
        return s.scalar(select(func.count()).select_from(Company)) or 0


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

_EXPORT_COLUMNS = [
    ("sap_number", "SAP Nummer"), ("name", "Firmenname"), ("kv", "KV"),
    ("segment", "Kundensegment"), ("sub_segment", "Kundenuntersegment"),
    ("sales_channel", "Vertriebsweg"), ("street", "Straße"), ("postal_code", "Postleitzahl"),
    ("city", "Ort"), ("country", "Land"), ("phone", "Telefon 1"), ("email", "E-Mail"),
    ("fax", "Fax"), ("website_domain", "Website"),
    ("revenue_y0", "Umsatz aktuelles Jahr"), ("revenue_y1", "Umsatz aktuelles Jahr -1"),
    ("revenue_y2", "Umsatz aktuelles Jahr -2"), ("revenue_y3", "Umsatz aktuelles Jahr -3"),
    ("revenue_y4", "Umsatz aktuelles Jahr -4"),
    # enriched fields (see enrich/) — without these an export of a freshly
    # enriched market would carry none of the information that was just gathered
    ("customer_state", "Kundenstatus"),
    ("description", "Beschreibung (Website)"),
    ("assessment", "Einschätzung (KI, unbestätigt)"),
    ("products_str", "Produkte"),
    ("founded_year", "Gegründet"),
    ("employee_hint", "Betriebsgröße (Angabe)"),
    ("mentions_solarlux_str", "Nennt Solarlux"),
    ("competitor_brands_str", "Wettbewerber auf Website"),
    ("active_ads", "Aktive Anzeigen"),
    ("enrichment_status", "Anreicherungs-Status"),
]


def export_xlsx(filters: dict | None = None, ids: list[int] | None = None,
                sort: str | None = None, direction: str = "asc") -> bytes:
    """Export either an explicit selection (ids) or the whole filtered set, as a
    .xlsx. Carries the master data AND the enriched fields (description, the
    marked AI assessment, products, brand mentions, ad count) — an export of a
    freshly enriched market is otherwise missing everything just gathered."""
    import openpyxl
    from .models import CompanyEnrichment
    with SessionLocal() as s:
        stmt = select(Company)
        if ids is not None:
            stmt = stmt.where(Company.id.in_(ids))
        else:
            stmt = _apply_filters(stmt, filters or {})
        stmt = _apply_sort(stmt, sort, direction)
        companies = list(s.scalars(stmt))
        rows = [_to_dict(c) for c in companies]
        _attach_active_ads(s, rows)
        enr = {e.company_id: (e.fields or {}) for e in s.scalars(
            select(CompanyEnrichment).where(
                CompanyEnrichment.company_id.in_([c.id for c in companies])))} if companies else {}

    def _yesno(v):
        return "ja" if v is True else ("nein" if v is False else None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"
    ws.append([label for _, label in _EXPORT_COLUMNS])
    for d in rows:
        f = enr.get(d["id"], {})
        d = dict(d)
        # the AI assessment is exported in its own clearly-labelled column so it
        # can never be read as a verified fact
        d["assessment"] = f.get("assessment_de")
        d["products_str"] = ", ".join(d.get("products") or f.get("products") or [])
        d["mentions_solarlux_str"] = _yesno(f.get("mentions_solarlux"))
        d["competitor_brands_str"] = ", ".join(f.get("competitor_brands") or [])
        ws.append([d.get(field) for field, _ in _EXPORT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
